import datetime
import os
import sys
from selenium.webdriver.common.keys import Keys
from datetime import datetime, timedelta
from datetime import datetime, time as datetime_time
import random
import logging
import time, random, math
import traceback
import openpyxl
import pyperclip
import pandas as pd
import pyautogui
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException,ElementNotInteractableException
from selenium.common.exceptions import TimeoutException,StaleElementReferenceException,NoSuchElementException,ElementClickInterceptedException,WebDriverException
import shutil
from selenium.webdriver.common.action_chains import ActionChains
import re

# Define the base directory (adjust this based on your project's structure)
base_dir = os.path.dirname(os.path.abspath(__file__))  # Adjust '..' to navigate up as needed
screencap_dir= os.path.join(base_dir, '..', '..', 'Files')

#Added
#GLobal log action function
test_stats = {"passed": 0, "failed": 0, "total": 0}


def open_module(driver, wait, menu_title):
    menu = wait.until(EC.presence_of_element_located((By.XPATH, f"//a[@data-bs-title='{menu_title}']")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", menu)
    driver.execute_script("""
        const ev = new MouseEvent('click', {bubbles:true, cancelable:true, view: window});
        arguments[0].dispatchEvent(ev);
    """, menu)
    time.sleep(0.5)

def clean_text(text: str) -> str:
    return text.replace("\n", "").replace("\r", "").strip()

def suppress_output(func, *args, **kwargs):
    original_stdout = sys.stdout
    sys.stdout = open(os.devnull, 'w')
    try:
        return func(*args, **kwargs)
    finally:
        sys.stdout.close()
        sys.stdout = original_stdout



def log_action(message, success=None, log_file_path=None, error=False, overwrite=False):
    global test_stats
    
    # Determine success status
    if error:
        success = False
        
    elif success is None:
        # Auto-detect success based on error keywords in message
        error_keywords = ["error", "failed", "exception", "not found", "unable", "timeout"]
        success = not any(keyword in message.lower() for keyword in error_keywords)

    status = "PASSED" if success else "FAILED"
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Update test stats
    test_stats["total"] += 1
    if success:
        test_stats["passed"] += 1
    else:
        test_stats["failed"] += 1

    # Log message
    log_message = f"[{current_time}] {message} - {status}"
    
    # ✅ CRITICAL FIX: Print to stdout with flush=True so pytest can capture it
    print(log_message.strip(), file=sys.stdout, flush=True)

    # Write to log file if specified
    if log_file_path:
        os.makedirs(os.path.dirname(log_file_path), exist_ok=True)
        mode = 'w' if overwrite else 'a'
        try:
            with open(log_file_path, mode, encoding='utf-8') as f:
                f.write(f"[{current_time}] {message} - {status}\n")
                # Only write traceback if there's an active exception
                if error and traceback.format_exc().strip() != "NoneType: None":
                    f.write(traceback.format_exc() + "\n")
        except Exception as e:
            print(f"⚠️ Failed to write to log file: {e}", file=sys.stdout, flush=True)


def get_test_summary():
    if test_stats["total"] == 0:
        return "No tests were run."
    
    pass_percentage = (test_stats["passed"] / test_stats["total"]) * 100
    return (f"Test Summary: Total: {test_stats['total']}, Passed: {test_stats['passed']}, "
            f"Failed: {test_stats['failed']}, Pass%: {pass_percentage:.2f}%")

def log_test_summary(log_file_path=None):
    summary = get_test_summary()
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_message = f"[{current_time}] {summary}\n"

    if log_file_path:
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(summary_message)

    print(summary_message.strip())

def reset_test_stats():
    global test_stats
    test_stats = {"passed": 0, "failed": 0, "total": 0}


def log_error(error_message, log_file_path=None, driver=None, screencap_dir=None):
    # Setup logger
    if log_file_path:
        logging.basicConfig(
            filename=log_file_path,
            level=logging.ERROR,
            format='[%(asctime)s] %(levelname)s: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        logger = logging.getLogger(__name__)
        logger.error(error_message, exc_info=True)

    # Capture screenshot on error if driver and directory are defined
    if driver and screencap_dir:
        os.makedirs(screencap_dir, exist_ok=True)
        script_name = os.path.splitext(os.path.basename(log_file_path or ""))[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        screenshot_path = os.path.join(
            screencap_dir, f"{script_name}_Error_{timestamp}.png"
        )
        try:
            driver.save_screenshot(screenshot_path)
        except Exception as screenshot_error:
            logger.error(f"Screenshot failed: {screenshot_error}")

# Function to select a random option from a dropdown
def select_random_option(dropdown_id, driver, wait, log_file_path=None):
    max_attempts = 5  # Set a maximum number of attempts to avoid infinite loops
    attempts = 0
    selection_successful = False
    selected_text = None
    while attempts < max_attempts and not selection_successful:
        try:
            wait.until(EC.presence_of_element_located((By.ID, dropdown_id)))
            select = Select(driver.find_element(By.ID, dropdown_id))
            options = [option for option in select.options if not option.is_selected() and not option.get_attribute('disabled')]
            if options:
                random_option = random.choice(options)
                select.select_by_visible_text(random_option.text)
                selected_text = random_option.text
                selection_successful = True  # Set to True if selection is made without exceptions
                log_action(f"Selected random option from {dropdown_id}", log_file_path=log_file_path)
            else:
                log_action(f"No enabled options available in dropdown {dropdown_id}", log_file_path=log_file_path)
                selection_successful = True  # Break out of loop if no options

        except Exception as e:
            attempts += 1  # Increment the attempt counter
            log_action(f"Attempt {attempts} to select random option from {dropdown_id} failed: {str(e)}", log_file_path=log_file_path)

    if not selection_successful:
        log_error(f"Failed to select a random option from {dropdown_id} after multiple attempts", log_file_path=log_file_path)

    return selected_text

def select_random_prefix(driver, log_file_path, screenshots_folder):
    # Wait for the dropdown to be present
    prefix_dropdown = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "contactPrefix"))
    )
    
    
    select = Select(prefix_dropdown)
    options = [o.get_attribute("value") for o in select.options if o.get_attribute("value")]
    random_prefix = random.choice(options)
    select.select_by_value(random_prefix)
    log_action(f"Selected random prefix: {random_prefix}", log_file_path=log_file_path)
    time.sleep(1)
    driver.save_screenshot(os.path.join(screenshots_folder, f"Random_Prefix_{random_prefix}.png"))

    return random_prefix

# Adjustment type 
def select_adjustment_type(driver, adjustment_type):
 
    adjustment_dropdown = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "adjustmentType"))
    )
    
    select = Select(adjustment_dropdown)
    
    if adjustment_type.lower() == 'increase':
        select.select_by_value("1.0")
    elif adjustment_type.lower() == 'decrease':
        select.select_by_value("-1.0")
    else:
        raise ValueError(f"Invalid adjustment type: {adjustment_type}")
    
def select_tax_type(driver, tax_type="VAT", log_file_path=None):
    tax_select_element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "taxSelect"))
    )

    select = Select(tax_select_element)
    select.select_by_value(tax_type)

    selected_option = select.first_selected_option.text
    selected_clean = clean_text(selected_option).strip().splitlines()[0]

    if log_file_path:
        log_action(f"Selected Tax Type: {selected_clean}", log_file_path=log_file_path)

    WebDriverWait(driver, 10).until(lambda d: d.execute_script('return document.readyState') == 'complete')

    return selected_clean


# Function to select a random option from a dropdown by name
def select_random_option_name(dropdown_name, driver, wait, log_file_path=None):
    max_attempts = 5  # Set a maximum number of attempts to avoid infinite loops
    attempts = 0
    selection_successful = False
    selected_text = None
    
    while attempts < max_attempts and not selection_successful:
        try:
            wait.until(EC.presence_of_element_located((By.NAME, dropdown_name)))
            select = Select(driver.find_element(By.NAME, dropdown_name))
            options = [option for option in select.options if not option.is_selected() and not option.get_attribute('disabled')]
            if options:
                random_option = random.choice(options)
                select.select_by_visible_text(random_option.text)
                selected_text = random_option.text
                selection_successful = True  # Set to True if selection is made without exceptions
                log_action(f"Selected random option from {dropdown_name}", log_file_path=log_file_path)
            else:
                log_action(f"No enabled options available in dropdown {dropdown_name}", log_file_path=log_file_path)
                selection_successful = True  # Break out of loop if no options

        except Exception as e:
            attempts += 1  # Increment the attempt counter
            log_action(f"Attempt {attempts} to select random option from {dropdown_name} failed: {str(e)}", log_file_path=log_file_path)

    if not selection_successful:
        log_error(f"Failed to select a random option from {dropdown_name} after multiple attempts", log_file_path=log_file_path)

    return selected_text

def create_product(driver, wait, log_file_path):
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/header/div/nav/ul/li[2]/ul/li[4]/a')))
        driver.execute_script("arguments[0].click()", element)
        log_action("Clicked Product Management", log_file_path=log_file_path)

        element = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/header/div/div/div[2]/ul/li[1]/a")))
        driver.execute_script("arguments[0].click()", element)
        log_action("Clicked Create New Product", log_file_path=log_file_path)
        time.sleep(15)

def orderlist(driver, log_file_path):
    driver.find_element(By.XPATH, "//*[@id=\"UI_Assets-EYE-02\"]").click()
    log_action("Clicked Eye icon in orders", log_file_path=log_file_path)    
            # List of label IDs to check
    label_ids = [
                "orderNoLabel", "clientNameLabel", "mobileLabel", "emailLabel", 
                "addressLabel", "socmedLabel", "orderStatusLabel", "amountPaidLabel"
            ]

            # Flag to track if all labels have values
    all_have_values = True

            # Iterate through each label and check if it has a value
    for label_id in label_ids:
        label_element = driver.find_element(By.ID, label_id)
        label_text = label_element.text.strip()  # Get the text and remove surrounding whitespace
                
        if not label_text:  # If the label has no value
            all_have_values = False
            log_action("Label" + {label_id} + "does not exist", False, log_file_path=log_file_path)
            break

        # Final decision
        if all_have_values:
            log_action("All fields in order are loaded", log_file_path=log_file_path)
        else:
            log_action("Fail: One or more labels are missing values.", False, log_file_path=log_file_path)

def get_different_password(excel_path, current_password):
    # Read data from Excel
    data = pd.read_excel(excel_path, sheet_name='Admin')
    
    # Get a list of unique passwords that are different from the current one
    other_passwords = data[data['Password'] != current_password]['Password'].unique()
    
    # If there are other passwords available, return one randomly
    if len(other_passwords) > 0:
        return other_passwords[0]  # You can modify this to choose randomly if desired
    else:
        raise ValueError("No different password found in the Excel sheet.")
    
def update_excel(excel_path, username, new_password, log_file_path, driver):
    try:
        # Load the Excel workbook
        workbook = load_workbook(excel_path)
        sheet = workbook['Admin']  # Access the "Admin" sheet only

        # Find the row where the username matches
        for row in range(2, sheet.max_row + 1):  # Assuming row 1 contains headers
            cell_username = sheet.cell(row=row, column=1).value
            if cell_username == username:
                # Update the password in the same row
                sheet.cell(row=row, column=2).value = new_password
                break

        # Save only the "Admin" sheet without affecting other sheets
        workbook.save(excel_path)
        workbook.close()

        log_action(f"Updated password for user: {username}", log_file_path=log_file_path)

        driver.find_element(By.ID, "username").send_keys(username)
        log_action(f"Inputted username", log_file_path=log_file_path)

        driver.find_element(By.ID, "password").send_keys(new_password)
        log_action(f"Inputted updated password: {username}", log_file_path=log_file_path)

        driver.find_element(By.NAME, "submit").click()
        log_action(f"Clicked submit", log_file_path=log_file_path)
    
    except Exception as e:
        log_action(f"Error updating Excel: {str(e)}", log_file_path=log_file_path)

# SCD - ADD NEW PRODUCT
def product_data():
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\product_data.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    df = pd.read_excel(excel_path)
    row = df.sample(n=1).iloc[0]
    
    return {
        "Product Name": row.get("Product Name", ""),
        "SKU": row.get("SKU", ""),
        "Units of Measure": row.get("Units of Measure", ""),
        "Product Description": row.get("Product Description", ""),
        "Unit Price": row.get("Unit Price", ""),
        "Product Value": row.get("Product Value", ""),
        "Predefined Tags": row.get("Predefined Tags", ""),
        "Product Sub-Category": row.get("Product Sub-Category", ""),
        "Product Category": row.get("Product Category", ""),
        "Quantity": row.get("Quantity", "")
    }
def upload_image_product(
    driver,
    image_path=None,
    image_folder=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Images\Product"
):
    #  1. Pick a random image
    images = [os.path.join(image_folder, f) for f in os.listdir(image_folder)
              if f.lower().endswith((".png", ".jpg", ".jpeg"))]
    if not images:
        raise FileNotFoundError(f"No images found in {image_folder}")
    random_image = random.choice(images)

    try:
        #  2. Scroll to bottom to ensure visibility
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        #  3. Locate FilePond's input[type=file]
        upload_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input.filepond--browser[type='file']"))
        )

        #  4. Scroll element into view
        driver.execute_script("arguments[0].scrollIntoView(true);", upload_input)
        time.sleep(0.5)

        # 5. Send file path directly to FilePond input
        upload_input.send_keys(random_image)
        print(f"📸 Uploaded image successfully: {os.path.basename(random_image)}")

        # 6. Wait for FilePond preview or processing to finish
        time.sleep(2)

        return random_image

    except Exception as e:
        print(f"❌ Upload failed: {e}")
        driver.save_screenshot("upload_error.png")
        return None

# Manage Service Bulk
def search_and_select_client(driver, search_text, select_text):

    try:
        wait = WebDriverWait(driver, 15)
        
        # Find the client search input
        print(f"Searching for: {search_text}")
        search_input = wait.until(
            EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Search client']"))
        )
        
        # Clear and enter search text
        search_input.clear()
        time.sleep(0.3)
        search_input.send_keys(search_text)
        time.sleep(1)  # Wait for search results to appear
        
        # Wait for the list item to be visible and clickable
        print(f"Selecting: {select_text}")
        
        # Try multiple selection strategies
        try:
            # Strategy 1: Click the label directly
            client_label = wait.until(
                EC.element_to_be_clickable((By.XPATH, f"//label[contains(text(), '{select_text}')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", client_label)
            time.sleep(0.3)
            client_label.click()
            
        except (TimeoutException, ElementClickInterceptedException):
            # Strategy 2: Click the parent li element
            client_item = wait.until(
                EC.element_to_be_clickable((By.XPATH, f"//li[contains(@class, 'data-view-item')]//label[contains(text(), '{select_text}')]/parent::li"))
            )
            driver.execute_script("arguments[0].click();", client_item)
        
        # Verify selection
        time.sleep(0.5)
        selected = driver.find_element(By.XPATH, f"//li[contains(@class, 'selected')]//label[contains(text(), '{select_text}')]")
        if selected:
            print(f"✓ Successfully selected: {select_text}")
            return True
            
    except TimeoutException as e:
        print(f"✗ Timeout: Could not find element - {e}")
        return False
    except Exception as e:
        print(f"✗ Error selecting client: {e}")
        return False

def search_and_select_address(driver, search_text, select_text):
   
    try:
        wait = WebDriverWait(driver, 15)
        
        print(f"Searching for address: {search_text}")
        address_search = wait.until(
            EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Search address']"))
        )
        
        address_search.clear()
        time.sleep(0.3)
        address_search.send_keys(search_text)
        time.sleep(1)
        
        print(f"Selecting address: {select_text}")
        
        try:
            # Try clicking the label
            address_label = wait.until(
                EC.element_to_be_clickable((By.XPATH, f"//ul[@id='addressList']//label[contains(text(), '{select_text}')]"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", address_label)
            time.sleep(0.3)
            address_label.click()
            
        except (TimeoutException, ElementClickInterceptedException):
            # Use JavaScript click as fallback
            address_item = wait.until(
                EC.presence_of_element_located((By.XPATH, f"//ul[@id='addressList']//li[contains(., '{select_text}')]"))
            )
            driver.execute_script("arguments[0].click();", address_item)
        
        print(f"✓ Successfully selected address: {select_text}")
        return True
        
    except Exception as e:
        print(f"✗ Error selecting address: {e}")
        return False

def select_data(driver, wait, name, visible_text):
    try:
        dropdown = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.NAME, name))
        )
        select = Select(dropdown)
        select.select_by_visible_text(visible_text)
        print(f"✓ Selected '{visible_text}' from dropdown '{name}'")
    except Exception as e:
        print(f"✗ Failed to select '{visible_text}' from '{name}': {e}")

# Manage Order - Seller Center  
def get_latest_order_from_log(log_file_path):

    if not os.path.exists(log_file_path):
        print(f"Log file not found: {log_file_path}")
        return None
    
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Search backwards for the most recent entry
        for line in reversed(lines):
            # Use regex to extract service name and remove "- PASSED" if present
            match = re.search(r"Clicked on product dropdown item:\s*(.*?)(?:\s*-\s*PASSED.*)?$", line)
            if match:
                service_name = match.group(1).strip()
                # Additional cleanup - remove any remaining "- PASSED"
                service_name = service_name.replace(" - PASSED", "").strip()
                return service_name
        
        print(f"No 'Product Name' entry found in log.")
        return None
        
    except Exception as e:
        print(f"Error reading log file: {e}")
        return None


# Manage Product -  Seller Center
def get_latest_product_name_from_log(log_file_path):
   
    try:
        with open(log_file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        for line in reversed(lines):
            match = re.search(r"Entered Product Name:\s*(.*?)(?:\s*-\s*PASSED)?\s*$", line)
            if match:
                return match.group(1).strip()
    except Exception as e:
            print(f"Error reading product name from log: {e}")
    return None


def apply_price_adjustment(driver, amount, adj_type="increase", log_action=None, log_file_path=None):

    # --- ENTER AMOUNT ---
    amount_input = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.ID, "adjustmentAmount"))
    )
    amount_input.clear()
    amount_input.send_keys(str(amount))

    if log_action:
        log_action(f"Inputted Adjustment Amount: {amount}", log_file_path=log_file_path)

    # Wait until the amount is actually reflected
    WebDriverWait(driver, 10).until(
        lambda d: d.find_element(By.ID, "adjustmentAmount").get_attribute("value") == str(amount)
    )

    # --- SELECT ADJUSTMENT TYPE ---
    dropdown = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "adjustmentType"))
    )
    select = Select(dropdown)

    adj_type = adj_type.lower()
    if adj_type == "increase":
        select.select_by_value("1.0")
    elif adj_type == "decrease":
        select.select_by_value("-1.0")
    else:
        raise ValueError(f"Invalid adj_type: {adj_type}")

    if log_action:
        log_action(f"Selected Adjustment Type: {adj_type.capitalize()}", log_file_path=log_file_path)

    # --- WAIT FOR PRICE ADJUSTMENT TO UPDATE ---
    def price_adjustment_updated(d):
        try:
            price = d.find_element(By.ID, "priceAdjustment")
            clear_btn = d.find_element(By.ID, "clearPriceAdjustment")

            # Trigger conditions:
            if price.get_attribute("value") != "0.00":
                return True
            if price.get_attribute("disabled") is None:  # Enabled
                return True
            if clear_btn.get_attribute("disabled") is None:  # Clear button enabled
                return True
            return False
        except:
            return False

    WebDriverWait(driver, 20).until(price_adjustment_updated)

    if log_action:
        log_action("Price Adjustment UI Updated", log_file_path=log_file_path)

def apply_shipping_adjustment(driver, amount, adj_type="increase", log_action=None, log_file_path=None):

    # --- Enter Amount ---
    amount_input = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.ID, "shippingFeeAdjustment"))
    )
    amount_input.clear()
    amount_input.send_keys(str(amount))

    if log_action:
        log_action(f"Inputted Shipping Amount: {amount}", log_file_path=log_file_path)

    WebDriverWait(driver, 10).until(
        lambda d: d.find_element(By.ID, "shippingFeeAdjustment").get_attribute("value") == str(amount)
    )

  

def apply_adjustment(driver, amount, adj_type="increase", target="price", tax_type=None,
                     log_action=None, log_file_path=None, timeout=20, retries=3):
    """
    Robust universal function to apply Price, Shipping, or Tax adjustments.
    Includes retries and optional tax type selection for Tax adjustments.

    Parameters:
    - driver: Selenium WebDriver
    - amount: numeric value to input
    - adj_type: "increase" or "decrease"
    - target: "price", "shipping", or "tax"
    - tax_type: value to select in tax dropdown (only used if target="tax")
    - log_action: optional logging function
    - log_file_path: optional log file path
    - timeout: wait timeout in seconds
    - retries: number of retry attempts
    """

    # --- Map element IDs and Apply buttons based on target ---
    target = target.lower()
    if target == "price":
        field_id = "priceAdjustment"
        clear_btn_id = "clearPriceAdjustment"
        apply_btn_xpath = "//div[contains(@class,'price-adjustment-container')]//button[contains(text(),'Apply Price Adjustment')]"
    elif target == "shipping":
        field_id = "shippingFee"
        clear_btn_id = "clearShippingAdjustment"
        apply_btn_xpath = "//div[contains(@class,'shipping-container')]//button[contains(text(),'Apply Shipping Fee')]"
    elif target == "tax":
        field_id = "tax"
        clear_btn_id = "clearTaxAdjustment"
        apply_btn_xpath = "//div[contains(@class,'tax-container')]//button[contains(text(),'Apply Tax')]"
        tax_select_id = "taxSelect"  # Tax dropdown ID
    else:
        raise ValueError(f"Invalid target: {target}. Must be 'price', 'shipping', or 'tax'.")

    for attempt in range(1, retries + 1):
        try:
            # --- Click the Apply button ---
            apply_btn = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, apply_btn_xpath))
            )
            apply_btn.click()
            if log_action:
                log_action(f"[Attempt {attempt}] Clicked Apply {target.capitalize()} button", log_file_path=log_file_path)

            # --- Enter Amount ---
            amount_input = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.ID, field_id))
            )
            amount_input.clear()
            amount_input.send_keys(str(amount))
            if log_action:
                log_action(f"[Attempt {attempt}] Inputted {target.capitalize()} Amount: {amount}", log_file_path=log_file_path)

            WebDriverWait(driver, timeout).until(
                lambda d: d.find_element(By.ID, field_id).get_attribute("value") == str(amount)
            )

            # --- Select Adjustment Type ---
            dropdown = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.ID, "adjustmentType"))
            )
            select = Select(dropdown)
            if adj_type.lower() == "increase":
                select.select_by_value("1.0")
            elif adj_type.lower() == "decrease":
                select.select_by_value("-1.0")
            else:
                raise ValueError(f"Invalid adj_type: {adj_type}")

            if log_action:
                log_action(f"[Attempt {attempt}] Selected {target.capitalize()} Adjustment Type: {adj_type.capitalize()}", log_file_path=log_file_path)

            # --- Optional: Select Tax Type ---
            if target == "tax" and tax_type:
                try:
                    tax_dropdown = WebDriverWait(driver, timeout).until(
                        EC.presence_of_element_located((By.ID, tax_select_id))
                    )
                    tax_select = Select(tax_dropdown)
                    tax_select.select_by_value(tax_type)
                    selected = tax_select.first_selected_option.text
                    if log_action:
                        log_action(f"[Attempt {attempt}] Selected Tax Type: {selected}", log_file_path=log_file_path)
                except (StaleElementReferenceException, ElementNotInteractableException):
                    if attempt == retries:
                        raise
                    if log_action:
                        log_action(f"[Attempt {attempt}] Warning: Could not select tax type {tax_type}. Retrying...", log_file_path=log_file_path)
                    time.sleep(1)
                    continue

            # --- Wait for UI update ---
            def adjustment_updated(d):
                try:
                    field = d.find_element(By.ID, field_id)
                    clear_btn = d.find_element(By.ID, clear_btn_id)
                    if field.get_attribute("value") != "0.00":
                        return True
                    if field.get_attribute("disabled") is None:
                        return True
                    if clear_btn.get_attribute("disabled") is None:
                        return True
                    return False
                except:
                    return False

            WebDriverWait(driver, timeout).until(adjustment_updated)

            if log_action:
                log_action(f"[Attempt {attempt}] {target.capitalize()} Adjustment UI Updated", log_file_path=log_file_path)

            # Success, exit retry loop
            break

        except (StaleElementReferenceException, ElementClickInterceptedException, Exception) as e:
            if attempt == retries:
                raise
            if log_action:
                log_action(f"[Attempt {attempt}] Warning: {e}. Retrying in 1s...", log_file_path=log_file_path)
            else:
                print(f"[Attempt {attempt}] Warning: {e}. Retrying in 1s...")
            time.sleep(1)


# Manage Service Order -  Service Center
def get_latest_name_from_log_service(log_file_path):
    if not os.path.exists(log_file_path):
        print(f"Log file not found: {log_file_path}")
        return None
    
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Search backwards for the most recent entry
        for line in reversed(lines):
            # Use regex to extract service name and remove "- PASSED" if present
            match = re.search(r"Entered Service Name:\s*(.*?)(?:\s*-\s*PASSED.*)?$", line)
            if match:
                service_name = match.group(1).strip()
                # Additional cleanup - remove any remaining "- PASSED"
                service_name = service_name.replace(" - PASSED", "").strip()
                return service_name
        
        print(f"No 'Service Name' entry found in log.")
        return None
        
    except Exception as e:
        print(f"Error reading log file: {e}")
        return None

# Manage Service Order -  Service Center
def get_latest_name_from_log(log_file_path):
    if not os.path.exists(log_file_path):
        print(f"Log file not found: {log_file_path}")
        return None
    
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Search backwards for the most recent entry
        for line in reversed(lines):
            # Use regex to extract service name and remove "- PASSED" if present
            match = re.search(r"Selected Product:\s*(.*?)(?:\s*-\s*PASSED.*)?$", line)
            if match:
                service_name = match.group(1).strip()
                # Additional cleanup - remove any remaining "- PASSED"
                service_name = service_name.replace(" - PASSED", "").strip()
                return service_name
        
        print(f"No 'Service Name' entry found in log.")
        return None
        
    except Exception as e:
        print(f"Error reading log file: {e}")
        return None











# Manage Service -  Service Center
def get_latest_employee_name_from_log(log_file_path):

    if not os.path.exists(log_file_path):
        print(f"Log file not found: {log_file_path}")
        return None
    
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Search backwards for the most recent entry
        for line in reversed(lines):
            # Use regex to extract service name and remove "- PASSED" if present
            match = re.search(r"Entered email:\s*(.*?)(?:\s*-\s*PASSED.*)?$", line)
            if match:
                service_name = match.group(1).strip()
                # Additional cleanup - remove any remaining "- PASSED"
                service_name = service_name.replace(" - PASSED", "").strip()
                return service_name
        
        print(f"No 'Email' entry found in log.")
        return None
        
    except Exception as e:
        print(f"Error reading log file: {e}")
        return None

# Client List - Client Directory -  Business Hub
def get_latest_client_list_from_log(log_file_path):

    if not os.path.exists(log_file_path):
        print(f"Log file not found: {log_file_path}")
        return None
    
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Search backwards for the most recent entry
        for line in reversed(lines):
            # Use regex to extract service name and remove "- PASSED" if present
            match = re.search(r"Entered email address:\s*(.*?)(?:\s*-\s*PASSED.*)?$", line)
            if match:
                service_name = match.group(1).strip()
                # Additional cleanup - remove any remaining "- PASSED"
                service_name = service_name.replace(" - PASSED", "").strip()
                return service_name
        
        print(f"No 'Email Address' entry found in log.")
        return None
        
    except Exception as e:
        print(f"Error reading log file: {e}")
        return None

def fill_order_row(driver, wait, client_id, adjustment=None, shipping=None, tax=None):
    try:
        # 1️ Adjustment Amount
        if adjustment is not None:
            adj_input = wait.until(
                EC.visibility_of_element_located(
                    (By.CSS_SELECTOR, f"input.adjustment-input[data-client-id='{client_id}']")
                )
            )
            adj_input.clear()
            adj_input.send_keys(str(adjustment))
            print(f" Adjustment set to {adjustment} for client {client_id}")

        # 2 Shipping Amount
        if shipping is not None:
            ship_input = wait.until(
                EC.visibility_of_element_located(
                    (By.CSS_SELECTOR, f"input.shipping-input[data-client-id='{client_id}']")
                )
            )
            ship_input.clear()
            ship_input.send_keys(str(shipping))
            print(f" Shipping set to {shipping} for client {client_id}")

        # 3️ Tax Dropdown
        if tax is not None:
            # Open dropdown
            dropdown_btn = wait.until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, f"button.tax-dropdown-toggle[data-client-id='{client_id}']")
                )
            )
            dropdown_btn.click()

            # Click desired option
            option = wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, f"//a[@class='dropdown-item' and @data-value='{tax}' and @data-client-id='{client_id}']")
                )
            )
            option.click()
            print(f" Tax set to {tax} for client {client_id}")

    except Exception as e:
        print(f" Failed to fill order row for client {client_id}: {e}")

def select_time_booking(driver, log_file_path, name_or_selector, time_value, by="name", press_enter=False):
    try:
        # Determine locator strategy
        if by == "name":
            locator = (By.NAME, name_or_selector)
        elif by == "id":
            locator = (By.ID, name_or_selector)
        elif by == "css":
            locator = (By.CSS_SELECTOR, name_or_selector)
        else:
            locator = (By.NAME, name_or_selector)
        
        log_action(f"Looking for time input: {name_or_selector} (by={by})", log_file_path=log_file_path)
        
        # Wait for time input to be present with better error handling
        try:
            time_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(locator)
            )
            log_action(f"✓ Found time input element", log_file_path=log_file_path)
        except TimeoutException:
            log_action(f"✗ Timeout: Could not find time input with {by}='{name_or_selector}'", log_file_path=log_file_path)
            
            # Try to find ANY time input as fallback
            try:
                time_input = driver.find_element(By.CSS_SELECTOR, "input[type='time'], input[data-start-time], input[data-end-time]")
                log_action(f"✓ Found time input using fallback selector", log_file_path=log_file_path)
            except:
                log_action(f"✗ No time input found with any selector", log_file_path=log_file_path)
                return False
        
        # Scroll to element to ensure it's visible
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", time_input)
        time.sleep(0.5)
        
        # Check if element is visible and enabled
        if not time_input.is_displayed():
            log_action(f"⚠ Time input is not visible", log_file_path=log_file_path)
        if not time_input.is_enabled():
            log_action(f"⚠ Time input is disabled", log_file_path=log_file_path)
        
        # Convert time to 24-hour format for HTML5 input
        time_24hr = convert_to_24hr(time_value)
        log_action(f"Converted '{time_value}' to '{time_24hr}'", log_file_path=log_file_path)
        
        # Method 1: Try direct value setting (for HTML5 time inputs)
        try:
            js_success = driver.execute_script("""
                var input = arguments[0];
                var timeValue = arguments[1];
                input.value = timeValue;
                input.dispatchEvent(new Event('input', { bubbles: true }));
                input.dispatchEvent(new Event('change', { bubbles: true }));
                return input.value === timeValue;
            """, time_input, time_24hr)
            
            if js_success:
                log_action(f"✓ Set time '{time_value}' via JavaScript", log_file_path=log_file_path)
                time.sleep(0.5)
                return True
            else:
                log_action(f"⚠ JavaScript set failed, trying UI interaction", log_file_path=log_file_path)
        except Exception as js_error:
            log_action(f"⚠ JavaScript error: {js_error}", log_file_path=log_file_path)
        
        # Method 2: Try clicking and using send_keys
        try:
            time_input.clear()
            time_input.click()
            time.sleep(0.3)
            time_input.send_keys(time_24hr)
            log_action(f"✓ Set time '{time_value}' via send_keys", log_file_path=log_file_path)
            time.sleep(0.5)
            return True
        except Exception as sendkeys_error:
            log_action(f"⚠ send_keys failed: {sendkeys_error}", log_file_path=log_file_path)
        
        # Method 3: UI interaction for custom time picker
        log_action(f"⚠ Trying custom time picker UI interaction", log_file_path=log_file_path)
        
        # Click the time input to open custom picker
        time_input.click()
        time.sleep(0.8)
        
        # Parse time components
        time_parts = parse_time(time_value)
        hour = time_parts['hour']
        minute = time_parts['minute']
        period = time_parts['period']  # AM/PM
        
        log_action(f"Parsed time: hour={hour}, minute={minute}, period={period}", log_file_path=log_file_path)
        
        # Select hour
        try:
            # Try multiple selectors for hour
            hour_xpath_options = [
                f"//li[text()='{hour:02d}']",
                f"//li[text()='{hour}']",
                f"//div[contains(@class, 'hour')]//li[text()='{hour:02d}']",
                f"//div[contains(@class, 'hour')]//li[text()='{hour}']",
                f"//ul[contains(@class, 'hour')]//li[text()='{hour:02d}']",
            ]
            
            hour_selected = False
            for xpath in hour_xpath_options:
                try:
                    hour_option = WebDriverWait(driver, 2).until(
                        EC.element_to_be_clickable((By.XPATH, xpath))
                    )
                    hour_option.click()
                    log_action(f"✓ Selected hour: {hour}", log_file_path=log_file_path)
                    hour_selected = True
                    time.sleep(0.3)
                    break
                except:
                    continue
            
            if not hour_selected:
                log_action(f"⚠ Could not select hour {hour}", log_file_path=log_file_path)
        except Exception as hour_error:
            log_action(f"⚠ Hour selection error: {hour_error}", log_file_path=log_file_path)
        
        # Select minute
        try:
            minute_xpath_options = [
                f"//li[text()='{minute:02d}']",
                f"//li[text()='{minute}']",
                f"//div[contains(@class, 'minute')]//li[text()='{minute:02d}']",
                f"//ul[contains(@class, 'minute')]//li[text()='{minute:02d}']",
            ]
            
            minute_selected = False
            for xpath in minute_xpath_options:
                try:
                    minute_option = WebDriverWait(driver, 2).until(
                        EC.element_to_be_clickable((By.XPATH, xpath))
                    )
                    minute_option.click()
                    log_action(f"✓ Selected minute: {minute}", log_file_path=log_file_path)
                    minute_selected = True
                    time.sleep(0.3)
                    break
                except:
                    continue
            
            if not minute_selected:
                log_action(f"⚠ Could not select minute {minute}", log_file_path=log_file_path)
        except Exception as minute_error:
            log_action(f"⚠ Minute selection error: {minute_error}", log_file_path=log_file_path)
        
        # Select AM/PM if present
        if period:
            try:
                period_xpath_options = [
                    f"//button[text()='{period}']",
                    f"//button[contains(text(), '{period}')]",
                    f"//a[text()='{period}']",
                    f"//span[text()='{period}']",
                ]
                
                period_selected = False
                for xpath in period_xpath_options:
                    try:
                        period_button = WebDriverWait(driver, 2).until(
                            EC.element_to_be_clickable((By.XPATH, xpath))
                        )
                        period_button.click()
                        log_action(f"✓ Selected period: {period}", log_file_path=log_file_path)
                        period_selected = True
                        time.sleep(0.3)
                        break
                    except:
                        continue
                
                if not period_selected:
                    log_action(f"⚠ Could not select {period}", log_file_path=log_file_path)
            except Exception as period_error:
                log_action(f"⚠ Period selection error: {period_error}", log_file_path=log_file_path)
        
        log_action(f"✓ Completed time picker interaction for '{time_value}'", log_file_path=log_file_path)
        
        # Press ENTER if requested
        if press_enter:
            time.sleep(0.3)
            time_input.send_keys(Keys.ENTER)
            log_action(f"✓ Pressed ENTER after setting time", log_file_path=log_file_path)
            time.sleep(0.3)
        
        time.sleep(0.5)
        return True
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        log_action(f"✗ Failed to set time '{time_value}': {e}", log_file_path=log_file_path)
        log_action(f"Error details: {error_details}", log_file_path=log_file_path)
        return False

# Parse time string into components
def parse_time(time_string):
   
    import re
    
    # Remove extra spaces
    time_string = time_string.strip()
    
    # Check for AM/PM
    period = None
    if 'AM' in time_string.upper():
        period = 'AM'
        time_string = time_string.upper().replace('AM', '').strip()
    elif 'PM' in time_string.upper():
        period = 'PM'
        time_string = time_string.upper().replace('PM', '').strip()
    
    # Split hour and minute
    parts = time_string.split(':')
    hour = int(parts[0])
    minute = int(parts[1]) if len(parts) > 1 else 0
    
    return {
        'hour': hour,
        'minute': minute,
        'period': period
    }

# Convert time to 24-hour format (HH:MM) for HTML5 input
def convert_to_24hr(time_string):

    time_parts = parse_time(time_string)
    hour = time_parts['hour']
    minute = time_parts['minute']
    period = time_parts['period']
    
    # Convert to 24-hour format
    if period == 'PM' and hour != 12:
        hour += 12
    elif period == 'AM' and hour == 12:
        hour = 0
    
    return f"{hour:02d}:{minute:02d}"

# Booking SCD
def select_customer_from_booking(driver, wait, customer_name):
   
    try:
        # Click to open the Select2 dropdown
        select2_container = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".select2-selection.select2-selection--single"))
        )
        select2_container.click()
        print(f"✓ Opened customer dropdown")
        time.sleep(0.5)
        
        # Wait for dropdown results to appear
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".select2-results__options"))
        )
        
        # Find the customer in the list (scroll if needed)
        customer_found = False
        max_attempts = 10
        
        for attempt in range(max_attempts):
            try:
                # Try to find the customer in the current visible list
                customer_option = driver.find_element(
                    By.XPATH, 
                    f"//li[contains(@class, 'select2-results__option') and normalize-space(text())='{customer_name}']"
                )
                
                # Scroll to the customer option within the dropdown
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'nearest'});", customer_option)
                time.sleep(0.3)
                
                # Click the customer
                customer_option.click()
                customer_found = True
                print(f"✓ Selected customer: {customer_name}")
                break
                
            except NoSuchElementException:
                # Customer not visible yet, scroll down in the dropdown
                results_container = driver.find_element(By.CSS_SELECTOR, ".select2-results__options")
                
                # Scroll down within the dropdown
                driver.execute_script("arguments[0].scrollTop += 200;", results_container)
                time.sleep(0.3)
                print(f"⚠ Scrolling dropdown to find {customer_name}... (attempt {attempt + 1})")
        
        if not customer_found:
            print(f"✗ Customer '{customer_name}' not found after scrolling")
            return False
        
        time.sleep(1)  # Wait for client name to auto-populate
        
        # Verify client name was populated
        client_name_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[data-client-name]"))
        )
        
        # Scroll to client name field
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", client_name_input)
        time.sleep(0.3)
        
        populated_value = client_name_input.get_attribute('value')
        
        if populated_value:
            print(f"✓ Client name auto-populated: {populated_value}")
            return True
        else:
            print(f"⚠ Client name field is empty after selection")
            return False
        
    except Exception as e:
        print(f"✗ Failed to select customer from dropdown: {e}")
        return False


def force_load_branches(driver):
    """
    Aggressively force branch options to load by triggering all possible events
    """
    print("\n🔥 FORCING BRANCH LOAD...")
    
    try:
        # Get the customer select
        customer_select = driver.find_element(By.CSS_SELECTOR, "select[name='EndCustomerId']")
        branch_select = driver.find_element(By.CSS_SELECTOR, "select[data-branch-location]")
        
        # Execute comprehensive JavaScript to force loading
        result = driver.execute_script("""
            var customerSelect = arguments[0];
            var branchSelect = arguments[1];
            
            // Log current state
            console.log('Customer value:', customerSelect.value);
            console.log('Branch options before:', branchSelect.options.length);
            
            // Trigger every possible event on customer select
            var events = ['change', 'input', 'blur', 'click', 'focus', 'select'];
            events.forEach(function(eventName) {
                var event = new Event(eventName, { bubbles: true, cancelable: true });
                customerSelect.dispatchEvent(event);
            });
            
            // If jQuery exists, trigger jQuery events
            if (window.jQuery) {
                jQuery(customerSelect).trigger('change').trigger('select');
            }
            
            // Also try Select2 specific events if it exists
            if (window.jQuery && jQuery.fn.select2) {
                jQuery(customerSelect).trigger('select2:select');
            }
            
            // Focus on branch select to trigger any lazy loading
            branchSelect.focus();
            branchSelect.click();
            
            // Wait a moment then check
            setTimeout(function() {
                console.log('Branch options after:', branchSelect.options.length);
            }, 1000);
            
            return {
                customerValue: customerSelect.value,
                branchOptionsBefore: branchSelect.options.length
            };
        """, customer_select, branch_select)
        
        print(f"  Trigger result: {result}")
        print("  ⏳ Waiting 3 seconds for AJAX...")
        time.sleep(3)
        
        # Check if branches loaded
        options = branch_select.find_elements(By.TAG_NAME, "option")
        valid_options = [opt for opt in options if opt.get_attribute("value")]
        
        print(f"  📊 Result: {len(valid_options)} branches loaded")
        
        return len(valid_options) > 0
        
    except Exception as e:
        print(f"  ⚠️ Force trigger error: {e}")
        return False
        return False
        return False


def inject_and_select_branch(driver, branch_name,timeout=30):
   
    print(f"\n💥 NUCLEAR OPTION: Injecting branches")
    
    try:
        # Manually create the branch options
        result = driver.execute_script("""
            var branchSelect = document.querySelector('select[data-branch-location]');
            if (!branchSelect) return {success: false, error: 'Select not found'};
            
            // Clear existing options (except placeholder)
            while (branchSelect.options.length > 1) {
                branchSelect.remove(1);
            }
            
            // Manually add the branch options we know exist
            var branches = [
                {value: '1', text: 'HQ'},
                {value: '2', text: 'VBRANCH'},
                {value: '3', text: 'VMALL'}
            ];
            
            branches.forEach(function(branch) {
                var option = document.createElement('option');
                option.value = branch.value;
                option.text = branch.text;
                branchSelect.appendChild(option);
            });
            
            // Find and select the target branch
            var targetBranch = arguments[0];
            var targetOption = null;
            
            for (var i = 0; i < branchSelect.options.length; i++) {
                if (branchSelect.options[i].text === targetBranch) {
                    targetOption = branchSelect.options[i];
                    branchSelect.selectedIndex = i;
                    branchSelect.value = branchSelect.options[i].value;
                    break;
                }
            }
            
            if (!targetOption) {
                return {success: false, error: 'Target branch not found'};
            }
            
            // Trigger all events
            var events = ['change', 'input', 'blur'];
            events.forEach(function(eventName) {
                var event = new Event(eventName, { bubbles: true });
                branchSelect.dispatchEvent(event);
            });
            
            return {
                success: true,
                selectedValue: branchSelect.value,
                selectedText: branchSelect.options[branchSelect.selectedIndex].text,
                totalOptions: branchSelect.options.length
            };
        """, branch_name)
        
        print(f"  Nuclear result: {result}")
        
        if result.get('success'):
            print(f"✅ NUCLEAR SUCCESS: Injected and selected '{result.get('selectedText')}'")
            time.sleep(1)
            return True
        else:
            raise Exception(f"Nuclear option failed: {result.get('error')}")
            
    except Exception as e:
        print(f"💥 Nuclear option exploded: {e}")
        raise

def bypass_branch_dropdown(driver, branch_name="HQ"):

    print(f"\n🚀 BYPASS MODE: Setting branch to {branch_name}")
    
    branch_map = {
        "HQ": "1",
        "VBRANCH": "2", 
        "VMALL": "3"
    }
    
    branch_id = branch_map.get(branch_name)
    
    if not branch_id:
        raise Exception(f"Unknown branch: {branch_name}")
    
    # Set the select value directly
    driver.execute_script("""
        var select = document.querySelector('select[data-branch-location]');
        var input = document.querySelector('input[name="BranchId"]');
        
        // Set select value
        if (select) {
            select.value = arguments[0];
            var event = new Event('change', {bubbles: true});
            select.dispatchEvent(event);
        }
        
        // Also try to set/create hidden input
        if (!input) {
            input = document.createElement('input');
            input.type = 'hidden';
            input.name = 'BranchId';
            document.querySelector('#addNewAppointmentForm').appendChild(input);
        }
        
        if (input) {
            input.value = arguments[0];
        }
        
        return {
            selectSet: select ? select.value : null,
            inputSet: input ? input.value : null
        };
    """, branch_id)
    
    print(f"✅ Bypassed - Branch ID set to: {branch_id}")
    return True

def product_dropdown(driver, wait, select_id, value_to_select, log_file_path=None, screenshots_folder=None):
 
    try:
        # Wait until the select element is present
        element = wait.until(EC.presence_of_element_located((By.ID, select_id)))
        
        # Scroll into view in case it's off-screen
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, select_id)))

        # Create a Select instance
        select = Select(element)

        # Convert the value to string to avoid float errors
        value_to_select = str(value_to_select).strip()

        # Select by visible text (like “Food”, “Drinks”)
        select.select_by_visible_text(value_to_select)

        print(f"Selected '{value_to_select}' from dropdown '{select_id}'")

        # Optional logging & screenshot
        if log_file_path:
            log_action(f"Selected '{value_to_select}' in dropdown '{select_id}'", log_file_path=log_file_path)
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, f"Dropdown_{select_id}_{value_to_select}.png"))

        return True

    except Exception as e:
        print(f" Failed to select '{value_to_select}' from dropdown '{select_id}': {e}")
        return False
  

# def set_flatpickr_date(driver, wait, log_file_path, input_name, days_offset=2):
 
#     try:
#         # --- Generate target datetime ---
#         target_datetime = datetime.now() + timedelta(days=days_offset)

#         # Windows-compatible formatting for Flatpickr (no seconds)
#         date_value = target_datetime.strftime("%B %#d, %Y %I:%M %p")  # e.g. November 7, 2025 02:30 PM
#         date_part = target_datetime.strftime("%B %#d, %Y")  # For calendar aria-label matching

#         # Wait for the input to be clickable
#         date_input = wait.until(EC.element_to_be_clickable((By.NAME, input_name)))

#         # Scroll into view
#         driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", date_input)

#         # Click to open the calendar
#         driver.execute_script("arguments[0].click();", date_input)
#         time.sleep(0.5)

#         # Try clicking the correct day in the calendar
#         try:
#             calendar = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.dayContainer")))
#             all_days = calendar.find_elements(By.CSS_SELECTOR, "span.flatpickr-day")

#             clicked = False
#             for day in all_days:
#                 aria = day.get_attribute("aria-label")
#                 if aria == date_part and "flatpickr-disabled" not in day.get_attribute("class"):
#                     driver.execute_script("arguments[0].scrollIntoView(true);", day)
#                     driver.execute_script("arguments[0].click();", day)
#                     log_action(f"📅 Clicked date in calendar: {aria}", log_file_path=log_file_path)
#                     clicked = True
#                     break
#             if not clicked:
#                 log_action(f" Date {date_part} not found or disabled in calendar.", log_file_path=log_file_path)
#         except Exception as e:
#             log_action(f" Could not interact with calendar: {e}", log_file_path=log_file_path)

#         # Set full date/time via JS (reliable for Flatpickr)
#         driver.execute_script("arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",date_input,date_value)
#         log_action(f" Set '{input_name}' to {date_value}", log_file_path=log_file_path)

#     except Exception as e:
#         log_action(f" Failed to set '{input_name}': {e}", log_file_path=log_file_path)
#         raise




def click_menu_item(driver, menu_text, log_file_path, timeout=15):
    """Click a menu item by its text with multiple fallback selectors"""
    selectors = [
        (By.XPATH, f"//span[contains(text(),'{menu_text}')]/parent::a"),
        (By.XPATH, f"//a[contains(@data-bs-title,'{menu_text}')]"),
        (By.XPATH, f"//a[.//span[text()='{menu_text}']]"),
        (By.PARTIAL_LINK_TEXT, menu_text),
    ]
    
    for selector in selectors:
        try:
            element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(selector))
            driver.execute_script("arguments[0].click();", element)
            log_action(f"Clicked {menu_text}", log_file_path=log_file_path)
            return True
        except:
            continue
    
    raise Exception(f"Could not click menu item: {menu_text}")

def set_flatpickr_date(driver, wait, log_file_path, input_name, days_offset=2):
 
    try:
        # --- Generate target datetime ---
        target_datetime = datetime.now() + timedelta(days=days_offset)

        # Windows-compatible formatting for Flatpickr (no seconds)
        date_value = target_datetime.strftime("%m/%d/%Y %H:%M:%S")
        date_part = target_datetime.strftime("%B %#d, %Y")

        # Wait for the input to be clickable
        date_input = wait.until(EC.element_to_be_clickable((By.NAME, input_name)))

        # Scroll into view
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", date_input)

        # Click to open the calendar
        driver.execute_script("arguments[0].click();", date_input)
        time.sleep(0.5)

        # Try clicking the correct day in the calendar
        try:
            calendar = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.dayContainer")))
            all_days = calendar.find_elements(By.CSS_SELECTOR, "span.flatpickr-day")

            clicked = False
            for day in all_days:
                aria = day.get_attribute("aria-label")
                if aria == date_part and "flatpickr-disabled" not in day.get_attribute("class"):
                    driver.execute_script("arguments[0].scrollIntoView(true);", day)
                    driver.execute_script("arguments[0].click();", day)
                    log_action(f"📅 Clicked date in calendar: {aria}", log_file_path=log_file_path)
                    clicked = True
                    break
            if not clicked:
                log_action(f" Date {date_part} not found or disabled in calendar.", log_file_path=log_file_path)
        except Exception as e:
            log_action(f" Could not interact with calendar: {e}", log_file_path=log_file_path)

        # Set full date/time via JS (reliable for Flatpickr)
        driver.execute_script("arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));", date_input, date_value)
        log_action(f" Set '{input_name}' to {date_value}", log_file_path=log_file_path)

        # ✅ CLOSE the calendar after setting the date
        # Method 1: Click outside to close
        driver.execute_script("document.body.click();")
        time.sleep(0.3)
        
        # Method 2: Close via Flatpickr API (more reliable)
        driver.execute_script("""
            var fp = arguments[0]._flatpickr;
            if (fp) { fp.close(); }
        """, date_input)
        time.sleep(0.3)

    except Exception as e:
        log_action(f" Failed to set '{input_name}': {e}", log_file_path=log_file_path)
        raise


def set_start_date(driver, wait, log_file_path):
    set_flatpickr_date(driver, wait, log_file_path, "PriceAdjustments[0].EffectiveFr", days_offset=1)
    

def set_end_date(driver, wait, log_file_path):
    set_flatpickr_date(driver, wait, log_file_path, "PriceAdjustments[0].EffectiveTo", days_offset=2)
    
def set_post_promo_price(driver, wait, price_value, log_file_path):
  
    try:
        # Wait for the post promo price input
        price_input = wait.until(
            EC.element_to_be_clickable((By.NAME, "postPromoPrice"))
        )

        # Scroll into view
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", price_input)
        time.sleep(0.5)

        # Clear and input value
        price_input.clear()
        driver.execute_script("arguments[0].value = '';", price_input)  # ensure clean state
        price_input.send_keys(str(price_value))

        # Trigger any event listener
        driver.execute_script("arguments[0].dispatchEvent(new Event('input'));", price_input)

        log_action(f" Set Post Promo Price: {price_value}", log_file_path=log_file_path)

    except Exception as e:
        log_action(f" Failed to set Post Promo Price: {str(e)}", log_file_path=log_file_path)
        raise

# Options 2
def set_future_date(driver, wait, field_name="PriceAdjustments[0].EffectiveFr", days_ahead=1):

    try:
        # Wait until the date input is clickable
        date_input = wait.until(
            EC.element_to_be_clickable((By.NAME, field_name))
        )

        # Compute future date/time
        future_datetime = (datetime.now() + timedelta(days=days_ahead)).strftime("%m/%d/%Y %H:%M:%S")

        # Clear and input value
        driver.execute_script("arguments[0].value = '';", date_input)
        date_input.send_keys(future_datetime)

        print(f"[INFO] Set {field_name} to {future_datetime}")
        return future_datetime

    except Exception as e:
        print(f"[ERROR] Failed to set date for {field_name}: {e}")
        return None


def apply_tax_with_double_apply(driver, tax_type="VAT", tax_percentage="12", log_file_path=None, screenshots_folder=None):
    try:
        print("\n" + "="*50)
        print(f"APPLYING TAX: {tax_type} @ {tax_percentage}%")
        print("="*50)
        
        # === FIRST APPLY: Open Tax Modal ===
        print("\n🖱️ STEP 1: Opening tax modal (First Apply)...")
        apply_tax_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Apply Tax')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", apply_tax_btn)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", apply_tax_btn)
        log_action("Clicked 'Apply Tax' button (First Apply)", log_file_path=log_file_path)
        print("  ✅ Tax modal opened")
        
        # Wait for modal
        time.sleep(2)
        
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, 'Tax_Modal_Step1.png'))
        
        # === SELECT TAX TYPE ===
        print("\n📋 STEP 2: Selecting tax type...")
        
        # Find tax select (try multiple methods)
        tax_select = None
        try:
            tax_select = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "taxSelect"))
            )
            print("  ✅ Found tax select")
        except:
            # Try finding by any select in modal
            print("  ⚠️ Tax select not found by ID, trying alternatives...")
            selects = driver.find_elements(By.TAG_NAME, "select")
            for sel in selects:
                if sel.is_displayed():
                    tax_select = sel
                    print(f"  ✅ Found visible select: {sel.get_attribute('id') or 'no-id'}")
                    break
        
        if tax_select:
            select = Select(tax_select)
            options = select.options
            print(f"  📋 Tax options: {[opt.text for opt in options]}")
            
            try:
                select.select_by_visible_text(tax_type)
                log_action(f"Selected Tax Type: {tax_type}", log_file_path=log_file_path)
                print(f"  ✅ Selected: {tax_type}")
            except:
                print(f"  ⚠️ Could not select {tax_type}, trying first valid option...")
                for opt in options:
                    if opt.get_attribute('value'):
                        opt.click()
                        break
            
            # Trigger change event
            driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", tax_select)
            time.sleep(1)
        
        # === ENTER TAX PERCENTAGE ===
        print("\n⌨️ STEP 3: Entering tax percentage...")
        
        # Wait for modal to update after tax type selection
        time.sleep(2)
        
        # Find tax input - try multiple selectors
        tax_input = None
        input_selectors = [
            (By.ID, "taxAdjustment"),
            (By.CSS_SELECTOR, "input[type='number']"),
            (By.CSS_SELECTOR, "input[type='text']"),
            (By.XPATH, "//input[@id='taxAdjustment']"),
            (By.XPATH, "//input[contains(@id, 'tax')]"),
        ]
        
        for by, selector in input_selectors:
            try:
                tax_input = driver.find_element(by, selector)
                if tax_input.is_displayed():
                    print(f"  ✅ Found tax input: {selector}")
                    break
                else:
                    tax_input = None
            except:
                continue
        
        if not tax_input:
            print("  ⚠️ WARNING: Could not find tax input field")
            print("  Attempting to continue anyway...")
        else:
            log_action("Tax input field found", log_file_path=log_file_path)
            
            # Wait for input to be ready
            max_wait = 10
            for i in range(max_wait):
                try:
                    if tax_input.is_enabled() and not tax_input.get_attribute("readonly"):
                        print(f"  ✅ Input ready after {i+1} checks")
                        break
                except:
                    pass
                time.sleep(1)
            
            # Scroll into view
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tax_input)
            time.sleep(0.5)
            
            # Enter percentage - multiple methods
            input_success = False
            
            # Method 1: Standard
            try:
                tax_input.click()
                time.sleep(0.3)
                tax_input.clear()
                time.sleep(0.3)
                tax_input.send_keys(tax_percentage)
                time.sleep(0.5)
                
                verify = tax_input.get_attribute("value")
                if verify == tax_percentage:
                    print(f"  ✅ Entered {tax_percentage}% (Method 1)")
                    input_success = True
            except Exception as e:
                print(f"  ⚠️ Method 1 failed: {e}")
            
            # Method 2: JavaScript
            if not input_success:
                try:
                    # Re-find element in case it's stale
                    tax_input = driver.find_element(By.ID, "taxAdjustment")
                    
                    driver.execute_script("""
                        var input = arguments[0];
                        var value = arguments[1];
                        
                        // Clear and set value
                        input.value = '';
                        input.value = value;
                        
                        // Trigger events
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                        input.dispatchEvent(new Event('blur', { bubbles: true }));
                    """, tax_input, tax_percentage)
                    
                    time.sleep(0.5)
                    verify = tax_input.get_attribute("value")
                    if verify == tax_percentage:
                        print(f"  ✅ Entered {tax_percentage}% (Method 2 - JS)")
                        input_success = True
                except Exception as e:
                    print(f"  ⚠️ Method 2 failed: {e}")
            
            if input_success:
                log_action(f"Entered tax: {tax_percentage}%", log_file_path=log_file_path)
            else:
                print(f"  ⚠️ Could not verify tax percentage entry")
        
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, 'Tax_Modal_Step2.png'))
        
        time.sleep(1)
        
        # === SECOND APPLY: Apply Tax Percentage ===
        print("\n🖱️ STEP 4: Applying tax (Second Apply)...")
        
        # Find Apply button in modal
        apply_btns = driver.find_elements(By.XPATH, "//button[contains(text(), 'Apply') or contains(@class, 'apply')]")
        
        print(f"  Found {len(apply_btns)} 'Apply' button(s)")
        
        clicked = False
        for i, btn in enumerate(apply_btns):
            try:
                if btn.is_displayed() and btn.is_enabled():
                    btn_text = btn.text
                    print(f"  Attempting button {i+1}: '{btn_text}'")
                    
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", btn)
                    
                    print(f"  ✅ Clicked Apply button {i+1}")
                    log_action(f"Clicked Apply button (Second Apply)", log_file_path=log_file_path)
                    clicked = True
                    break
            except Exception as e:
                print(f"  ⚠️ Button {i+1} failed: {e}")
        
        if not clicked:
            print("  ⚠️ WARNING: Could not click any Apply button")
        
        # Wait for modal to process
        time.sleep(2)
        
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, 'Tax_Modal_Step3.png'))
        
        # === POSSIBLE THIRD STEP: Confirm/OK ===
        print("\n🔍 STEP 5: Checking for confirmation dialog...")
        
        try:
            # Check for OK/Confirm buttons
            confirm_btns = driver.find_elements(By.XPATH, "//button[contains(text(), 'OK') or contains(text(), 'Confirm') or contains(text(), 'Yes')]")
            
            if len(confirm_btns) > 0:
                print(f"  Found {len(confirm_btns)} confirmation button(s)")
                for btn in confirm_btns:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        print(f"  ✅ Clicked confirmation: {btn.text}")
                        log_action("Clicked confirmation button", log_file_path=log_file_path)
                        time.sleep(1)
                        break
            else:
                print("  ℹ️ No confirmation needed")
        except Exception as e:
            print(f"  ℹ️ No confirmation dialog: {e}")
        
        # === FINALIZE ===
        print("\n⏳ Finalizing...")
        time.sleep(2)
        
        # Wait for page to be ready
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, 'Tax_Final.png'))
        
        print("\n" + "="*50)
        print("✅ TAX APPLICATION COMPLETE")
        print("="*50 + "\n")
        
        log_action(f"Tax applied: {tax_type} @ {tax_percentage}%", log_file_path=log_file_path)
        
        return True
        
    except Exception as e:
        print(f"\n❌ TAX APPLICATION FAILED")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        
        if log_file_path:
            log_error(f"Tax failed: {str(e)}", log_file_path=log_file_path)
        
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, 'Tax_Error.png'))
            try:
                with open(os.path.join(screenshots_folder, 'Tax_Error.html'), 'w', encoding='utf-8') as f:
                    f.write(driver.page_source)
            except:
                pass
        
        raise




def wait_for_input_value_change(driver, locator, old_values, timeout=10):

    wait = WebDriverWait(driver, timeout)
    def _value_changed(drv):
        elem = drv.find_element(*locator)
        val = elem.get_attribute("value")
        # If input is disabled and it's tricky, you could enable it temporarily:
        # drv.execute_script("arguments[0].removeAttribute('disabled')", elem)
        return val not in old_values and val is not None
    wait.until(_value_changed)
    # Once condition is met, return the new value
    elem = driver.find_element(*locator)
    return elem.get_attribute("value")














#utility.py
def save_employee(username, employee_path, password, log_file_path):
    try:
        # Get the copied value from the clipboard
        copied_value = pyperclip.paste().strip()
        log_action("Copied value from clipboard retrieved", log_file_path)

        log_action(f"Excel file loaded from path: {employee_path}", log_file_path=log_file_path)
        workbook = openpyxl.load_workbook(employee_path)
        sheet = workbook.active  # Assuming data is saved in the first sheet

        # Find the next empty row in the sheet
        next_row = sheet.max_row + 1

        # Save the username and copied value to the next row
        sheet.cell(row=next_row, column=1, value=username)  # Assuming column 1 is for username
        sheet.cell(row=next_row, column=2, value=copied_value)  # Assuming column 2 is for the copied value

        log_action(f"Username '{username}' and copied value saved to row {next_row}", log_file_path=log_file_path)

        # Save the workbook
        workbook.save(employee_path)
        log_action(f"Employee data successfully saved to {employee_path}", log_file_path=log_file_path)
    
    except Exception as e:
        error_message = f"Failed to save employee: {str(e)}"
        log_error(error_message, log_file_path=log_file_path)
        raise

def get_future_time():
    now = datetime.now()  # Get the current time
    future_time = now + timedelta(minutes=5)  # Add 5 minutes
    return future_time.strftime("%m/%d/%Y %H:%M")

def get_end_time():
    now = datetime.now()  # Get the current time
    future_time = now + timedelta(minutes=5)  # Add 5 minutes
    future_time_plus_30 = future_time + timedelta(minutes=30)  # Add another 30 minutes
    return future_time_plus_30.strftime("%m/%d/%Y %H:%M")  # Format as MM/DD/YYYY HH:MM AM/PM

def find_element_by_id(wait, element_id, text):
    xpath = f"//*[contains(@id, '{element_id}') and contains(text(), '{text}')]"
    return wait.until(EC.presence_of_element_located((By.XPATH, xpath)))

def find_element_by_name(wait, element_name, text):
    xpath = f"//*[contains(@name, '{element_name}') and contains(text(), '{text}')]"
    return wait.until(EC.presence_of_element_located((By.XPATH, xpath)))



# Click an element
def click_element(wait, driver, by, value, log_file_path):
    try:
        element = wait.until(EC.element_to_be_clickable((by, value)))
        driver.execute_script("arguments[0].click()", element)
        log_action(f"Clicked element {value}", log_file_path=log_file_path)
    except Exception as e:
        error_message = f"Element not found or interaction failed: {repr(e)}"
        log_error(error_message, log_file_path=log_file_path, driver=driver)
        error_message = f"Element not found or interaction failed: {traceback.format_exc()}"
        print(error_message)

# Enter text in an element
def enter_text(wait, driver, by, value, text, log_file_path):
    try:
        element = wait.until(EC.presence_of_element_located((by, value)))
        element.clear()
        element.send_keys(text)
        log_action(f"Entered text '{text}' in element {value}", log_file_path=log_file_path)
    except Exception as e:
        error_message = f"Element not found or interaction failed: {repr(e)}"
        log_error(error_message, log_file_path=log_file_path, driver=driver)
        error_message = f"Element not found or interaction failed: {traceback.format_exc()}"
        print(error_message)

def click_button_text(wait, driver, item_text, log_file_path):
    try:
        element = wait.until(EC.presence_of_element_located((By.XPATH, f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{item_text.lower()}')")))
        driver.execute_script("arguments[0].click()", element)
        log_action(f"Selected item {item_text}", log_file_path=log_file_path)
    except Exception as e:
        error_message = f"Element not found or interaction failed: {repr(e)}"
        log_error(error_message, log_file_path=log_file_path, driver=driver)
        error_message = f"Element not found or interaction failed: {traceback.format_exc()}"
        print(error_message)

def select_item(wait, driver, element_id, item_text, log_file_path):
    try:
        element = wait.until(EC.element_to_be_clickable((By.ID, element_id)))
        driver.execute_script("arguments[0].click()", element)
        element = wait.until(EC.presence_of_element_located((By.XPATH, f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{item_text.lower()}') and contains(@class, 'list-group-item')]")))
        driver.execute_script("arguments[0].click()", element)
        log_action(f"Selected item {item_text}", log_file_path=log_file_path)
    except Exception as e:
        error_message = f"Element not found or interaction failed: {repr(e)}"
        log_error(error_message, log_file_path=log_file_path, driver=driver)
        error_message = f"Element not found or interaction failed: {traceback.format_exc()}"
        print(error_message)

# Product Management - Create New Bulk Product - Select Warehouse
def select_warehouse(driver, wait, select_id, value):
    try:
        wait.until(EC.presence_of_element_located((By.ID, select_id)))
        time.sleep(0.5)
        dropdown = wait.until(EC.element_to_be_clickable((By.ID, select_id)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown)
        time.sleep(0.5)

        # Always re-find the element to avoid stale references
        select = Select(driver.find_element(By.ID, select_id))
        select.select_by_visible_text(value)
        print(f" Selected '{value}' from '{select_id}'")
        time.sleep(1)
    except Exception as e:
        print(f" Failed to select '{value}' from '{select_id}': {e}")

# Manage Service Order
def search_select_order(driver,wait,log_file_path):

        search_field = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "searchTable")))
        ActionChains(driver).move_to_element(search_field).perform()
        WebDriverWait(driver, 5).until(lambda d: search_field.is_enabled())

        search_field.clear()
        search_field.send_keys("Unpaid")
        log_action("Searched: Unpaid", log_file_path=log_file_path)
        time.sleep(2)

        filter_payment = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "statusFilter")))
        Select(filter_payment).select_by_value("all")
        log_action("Selected 'All Orders' from Payment Status filter", log_file_path=log_file_path)

        service_status = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "orderStatusFilter")))
        Select(service_status).select_by_value("all")
        log_action("Selected 'All Orders' from Service Order Status filter", log_file_path=log_file_path)

        # Calculate date range (last 7 days)
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)

        date_range_text = f"{start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"

        date_range = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "daterange")))

        date_range.clear()
        date_range.send_keys(date_range_text)
        date_range.send_keys(Keys.RETURN)
        log_action(f"Applied Date Range filter: {date_range_text}", log_file_path=log_file_path)
            
        # =========================
        # Select Order and View Profile
        # =========================
        checkbox = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#manageOrderTable tbody tr:first-child td:first-child input[type="checkbox"]')))
        try:
            checkbox.click()
        except Exception:
            driver.execute_script("arguments[0].click();", checkbox)
            log_action("Checked first order row", log_file_path=log_file_path)
    
        time.sleep(5)
        


# Inventory Management

def select_data_by_text(log_file_path, wait, select_id, option_text_or_value, by="text"):
 
    try:
        select_element = wait.until(
            EC.presence_of_element_located((By.ID, select_id))
        )
        select = Select(select_element)
        
        if by == "value":
            # Select by value attribute
            select.select_by_value(option_text_or_value)
        elif by == "index":
            # Select by index (0-based)
            select.select_by_index(int(option_text_or_value))
        else:  # by == "text"
            # Check for duplicate options
            matching_options = [opt for opt in select.options 
                              if opt.text.strip() == option_text_or_value.strip()]
            
            if len(matching_options) > 1:
                
                print(f" Warning: Found {len(matching_options)} options with text '{option_text_or_value}'")
                print(f"   Using the first match with value: {matching_options[0].get_attribute('value')}")
                # Select the first matching option by its value
                select.select_by_value(matching_options[0].get_attribute('value'))
            elif len(matching_options) == 1:
                select.select_by_visible_text(option_text_or_value)
            else:
                raise NoSuchElementException(f"No option found with text: {option_text_or_value}")
        
        return select.first_selected_option
        
    except Exception as e:
        log_action(f"Failed to select '{option_text_or_value}' from '{select_id}': {str(e)}",log_file_path=log_file_path)
        raise

def handle_inventory_modal(driver, wait, action="update_quantity", log_file_path=None):
 
    try:
        # Wait for modal to appear
        modal = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
        )

        # Determine which button to click
        if action == "update_quantity":
            button = modal.find_element(By.CSS_SELECTOR, "button[data-update-quantity='true']")
            log_msg = "Clicked 'Update Quantity' in modal"
        elif action == "change_warehouse":
            button = modal.find_element(By.CSS_SELECTOR, "button[data-change-warehouse='true']")
            log_msg = "Clicked 'Change Warehouse' in modal"
        elif action == "cancel":
            button = modal.find_element(By.CSS_SELECTOR, "#closeExistingModalSwal")
            log_msg = "Clicked 'Cancel' in modal"
        else:
            raise ValueError(f"Unknown action: {action}")

        # Scroll button into view
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
        # Click the button
        try:
            button.click()
        except:
            driver.execute_script("arguments[0].click();", button)

        # Optional logging
        if log_file_path:
            with open(log_file_path, "a") as f:
                f.write(f"{log_msg}\n")
        else:
            print(log_msg)

        # Wait for modal to disappear
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element(modal)
        )
        time.sleep(1)  # small delay to ensure page updates

    except Exception as e:
        # If modal never appears, just continue
        if log_file_path:
            with open(log_file_path, "a") as f:
                f.write(f"No modal appeared. Continuing workflow. ({e})\n")
        else:
            print(f"No modal appeared. Continuing workflow. ({e})")



def select_service(wait, driver, element_id, item_text, log_file_path):
    try:
        element = wait.until(EC.element_to_be_clickable((By.ID, element_id)))
        driver.execute_script("arguments[0].click()", element)
        element = wait.until(EC.presence_of_element_located((By.XPATH, f"//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{item_text.lower()}') and contains(@class, 'list-group-item')]")))
        driver.execute_script("arguments[0].click()", element)
        log_action(f"Selected item {item_text}", log_file_path=log_file_path)
    except Exception as e:
        error_message = f"Element not found or interaction failed: {repr(e)}"
        log_error(error_message, log_file_path=log_file_path, driver=driver)
        error_message = f"Element not found or interaction failed: {traceback.format_exc()}"
        print(error_message)

def select_item_qr(wait, driver, element_id, item_text, log_file_path):
    try:
        element = wait.until(EC.element_to_be_clickable((By.ID, element_id)))
        driver.execute_script("arguments[0].click()", element)
        element = wait.until(EC.presence_of_element_located((By.XPATH, f"//div[contains(@class, 'product-item')]//span[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{item_text}')]")))
        driver.execute_script("arguments[0].click()", element)
        log_action(f"Selected item {item_text}", log_file_path=log_file_path)
    except Exception as e:
        error_message = f"Element not found or interaction failed: {repr(e)}"
        log_error(error_message, log_file_path=log_file_path, driver=driver)
        error_message = f"Element not found or interaction failed: {traceback.format_exc()}"
        print(error_message)

def select_random_list_item(wait, driver, list_xpath, log_file_path):
    try:
        # Get all list items inside the specified list
        list_items = wait.until(EC.presence_of_all_elements_located((By.XPATH, f"{list_xpath}/li")))

        if list_items:
            random_item = random.choice(list_items)
            random_item_text = random_item.text.strip()
            random_item.click()
            log_action(f"Selected item: {random_item_text}", log_file_path=log_file_path)
            return random_item_text
        else:
            log_action(f"No options found in the list: {list_xpath}", log_file_path=log_file_path)
            return None
    except Exception as e:
        log_error(f"Error selecting item from {list_xpath}: {str(e)}", log_file_path=log_file_path, driver=driver)

def handle_reference_number(channel_element, driver, wait, log_file_path):
    selected_channel = channel_element.get_attribute("value").lower()

    if selected_channel in ["QRPH", "CREDIT/DEBIT CARD", "BANK TRANSFER/DEPOSIT"]:
        reference_field = find_element_by_id(wait, "reference", "")
        reference_field.send_keys("REF12345678")  # Replace with dynamic value if needed
        log_action(f"Entered Reference Number for channel: {selected_channel}", log_file_path=log_file_path)
    else:
        log_action(f"Channel selected: {selected_channel} (No reference required)", log_file_path=log_file_path)

#Updated
def go_next(driver, wait, log_action, log_file_path, max_retries=3):
    next_xpath = "//*[@id='nextPage']"
    retries = 0

    while True:
        try:
            btn = wait.until(EC.element_to_be_clickable((By.XPATH, next_xpath)))
            
            # Scroll down so the button appears in view
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.5)  # pause so the scroll is visible
            
            driver.execute_script("arguments[0].click();", btn)
            log_action("Clicked Next", log_file_path)
            
            time.sleep(0.5)  # pause to show the click effect
            driver.execute_script("window.scrollTo(0, 0);")  # scroll back up

            # Wait until the button goes stale to confirm navigation
            wait.until(EC.staleness_of(btn))
            log_action("Navigated to next page", log_file_path)
            
            retries = 0  # reset retry counter
            time.sleep(1)  # allow page settle

        except TimeoutException:
            log_action("No more pages (Next missing or not clickable)", log_file_path)
            break

        except StaleElementReferenceException:
            retries += 1
            if retries <= max_retries:
                log_action(f"Stale element on Next click; retry {retries}/{max_retries}", log_file_path)
                time.sleep(1)
                continue
            else:
                log_action("Failed clicking Next after retries", log_file_path)
                break

def save_screenshot(driver, module_name, base_dir='screenshots'):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    module_dir = os.path.join(base_dir, module_name)
    os.makedirs(module_dir, exist_ok=True)
    filename = f"{module_name}_{timestamp}.png"
    filepath = os.path.join(module_dir, filename)
    driver.save_screenshot(filepath)
    return filepath

def clear_terminal():
    if os.name == 'nt':   # Windows
        os.system('cls')
    else:                 # Linux / macOS
        os.system('clear')

def is_nan_value(val):
    """Return True if val is a numeric NaN, or string 'nan' (case insensitive)."""
    # Numeric NaNs
    if isinstance(val, float):
        if math.isnan(val):
            return True
    # Possibly other numeric types? You can extend if needed.

    # If it's a string, check if it equals 'nan' (or maybe contains 'nan' etc.)
    if isinstance(val, str):
        # strip whitespace and compare lowercase
        if val.strip().lower() == 'nan':
            return True

    return False

def human_like_typing(element, text, min_delay=0.05, max_delay=0.15):

    # Always interact with the element (click and clear)
    try:
        element.click()
    except Exception:
        pass

    time.sleep(random.uniform(0.1, 0.3))

    # Always clear existing text
    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.DELETE)
    time.sleep(random.uniform(0.1, 0.3))

    # Always type something (guaranteed to have data)

    for char in text:
        element.send_keys(char)
        delay = random.uniform(min_delay, max_delay)
        time.sleep(delay)

    time.sleep(random.uniform(0.2, 0.5))
    return text

def safe_click(driver, wait, locator, description, log_file_path, report=None, module="CDirectory", test_name=None, back_after=False):
  
    test_start = time.time()
    try:
        element = wait.until(EC.element_to_be_clickable(locator))
        driver.execute_script("arguments[0].click()", element)
        log_action(f"Clicked {description}", log_file_path=log_file_path)
        if back_after:
            driver.back()
            time.sleep(1)
        if report:
            report.add_test_result(
                module_name=module,
                test_name=test_name or description,
                status="PASSED",
                description=f"Successfully clicked {description}",
                duration=time.time() - test_start
            )
        return True
    except Exception as e:
        if report:
            report.add_test_result(
                module_name=module,
                test_name=test_name or description,
                status="FAILED",
                description=f"Failed to click {description}",
                error_message=str(e),
                duration=time.time() - test_start
            )
        raise

def random_search(excel_path, sheet_name="SearchTerms", column_name="SearchTerm"):

    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    except Exception as e:
        raise RuntimeError(f"Could not read Excel file {excel_path}: {str(e)}")
    if column_name not in df.columns:
        raise ValueError(f"Column {column_name} not found in sheet {sheet_name}")
    
    # Drop NaNs, convert to str, strip whitespace
    terms = df[column_name].dropna().astype(str).str.strip().tolist()
    # Optional: filter out empty strings
    terms = [t for t in terms if t]
    if not terms:
        raise ValueError("No valid search terms found in Excel file")
    return terms

def random_client_type_select(driver, wait):
    try:
        # Wait until the select dropdown is clickable / present
        select_elem = wait.until(EC.element_to_be_clickable((By.NAME, "statusFilter")))
        
        # Wrap it in Selenium’s Select class
        sel = Select(select_elem)
        
        # Get all available options (texts)
        options = sel.options
        visible_texts = [opt.text.strip() for opt in options if opt.text.strip()]
        if not visible_texts:
            return None
        
        # Choose randomly
        choice = random.choice(visible_texts)
        
        # Select by visible text
        sel.select_by_visible_text(choice)
        
        return choice

    except Exception as e:
        # Optionally print or log the error
        print(f"random_client_type_select failed: {e}")
        return None

#Employee    
def random_employee():
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\employee_testdata.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    df = pd.read_excel(excel_path)
    row = df.sample(n=1).iloc[0]
    
    return {
        "First Name": row.get("First Name", ""),
        "Middle Name": row.get("Middle Name", ""),  # will be "" if not in Excel
        "Last Name": row.get("Last Name", ""),
        "Birthday": row.get("Birthday", ""),
        "Civil Status": row.get("Civil Status", ""),
        "Citizenship": row.get("Citizenship", ""),
        "TIN": row.get("TIN", ""),
        "Number of Dependencies": row.get("Number of Dependencies", 0),
        "Mobile Number": row.get("Mobile Number", ""),
        "Email": row.get("Email", "")
    }

# def upload_image(driver, button_id="uploadPictureId",
#                             image_folder=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Images\One Piece"):
#     # Collect random image
#     images = [os.path.join(image_folder, f) for f in os.listdir(image_folder)
#               if f.lower().endswith((".png", ".jpg", ".jpeg"))]
#     if not images:
#         raise FileNotFoundError(f"No images found in {image_folder}")
#     random_image = random.choice(images)

#     # Click the upload button to open dialog
#     upload_btn = driver.find_element(By.ID, button_id)
#     driver.execute_script("arguments[0].click()", upload_btn)
#     time.sleep(2)  # wait for dialog to appear

#     # Type path + press Enter
#     pyautogui.write(random_image)
#     pyautogui.press("enter")

#     return random_image



def upload_image_by_name(driver, employee_name,
                         button_id="uploadPictureId",
                         image_folder=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Images\One Piece"):
    """
    Upload an image matching the employee's first and last name.
    """
    import pyautogui
    
    # Normalize employee name: lowercase, remove spaces
    normalized_name = employee_name.lower().replace(" ", "")
    
    # Get all image files
    images = [f for f in os.listdir(image_folder) if f.lower().endswith((".png", ".jpg", ".jpeg"))]

    # Find matching images (normalize filenames the same way)
    matched_images = [os.path.join(image_folder, f) 
                      for f in images 
                      if normalized_name in f.lower().replace(" ", "").replace("_", "")]

    if not matched_images:
        available = os.listdir(image_folder)
        raise FileNotFoundError(f"No images found for '{employee_name}'. Available images: {available}")

    # Pick the first match
    employee_image = matched_images[0]

    # Upload
    upload_btn = driver.find_element(By.ID, button_id)
    driver.execute_script("arguments[0].click()", upload_btn)
    time.sleep(2)  # wait for dialog

    pyautogui.write(employee_image)
    pyautogui.press("enter")

    return employee_image


def navigate_flyout(driver, link_text, flyout_title=None, timeout=10):
    # Wait for flyout panel to be visible
    flyout_panel = WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "div.flyout-panel"))
    )

    # If flyout_title is specified, ensure this panel matches the title
    if flyout_title:
        title_elem = flyout_panel.find_element(By.CSS_SELECTOR, "h1[data-flyout-title]")
        if flyout_title.strip() not in title_elem.text.strip():
            raise Exception(f"Flyout with title '{flyout_title}' not found")

    # Find the link inside the flyout by text
    link_elem = flyout_panel.find_element(By.XPATH, f".//li/a[.//span[text()='{link_text}']]")

    # Scroll into view and click
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", link_elem)
    driver.execute_script("arguments[0].click();", link_elem)



def Main_Dashboard(driver,log_file_path,screenshots_folder):
    main_dashboard = WebDriverWait(driver, 30).until(
    EC.element_to_be_clickable((
        By.XPATH,
        "//a[@href='/Home' "
        "and @data-bs-title='Main Dashboard' "
        "and @data-bs-toggle='tooltip' "
        "and @data-bs-custom-class='ob-side-menu-tooltip' "
        "and @data-sub-menu='' "
        "and .//span[text()='Main Dashboard']]"
    )))

    driver.execute_script("arguments[0].click();", main_dashboard)
    log_action("Clicked Main Dashboard", log_file_path=log_file_path)

    WebDriverWait(driver, 30).until(lambda d: d.execute_script("return document.readyState") == "complete")
    driver.save_screenshot(os.path.join(screenshots_folder, "MainDashboard.png"))

def click_service_management(driver, timeout=30, screenshot_path="service_mgmt_debug.png"):
    wait = WebDriverWait(driver, timeout)
    strategies = [
        # CSS by data attribute (fast & simple)
        ("css", "a[data-bs-title='Service Management']"),
        # XPath: attribute on <a> and descendant span contains text (robust)
        ("xpath", "//a[@data-bs-title='Service Management' and @data-bs-toggle='tooltip']//span[contains(normalize-space(.), 'Service Management')]"),
        # XPath: any element with that text anywhere inside link (very permissive)
        ("xpath", "//a[contains(., 'Service Management')]"),
        # XPath: exact normalized text in span (strict)
        ("xpath", "//a[@data-bs-title='Service Management']//span[normalize-space()='Service Management']"),
        # JS query for data attribute + text check (fallback)
        ("js", "document.querySelectorAll(\"a[data-bs-title='Service Management']\");")
    ]

    last_err = None

    # helper to try clicking element
    def try_click(el):
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
            wait.until(EC.element_to_be_clickable((By.XPATH, "(.//*[(name() = 'dummy')])[1]")), timeout=0)  # noop to use wait object (safe)
        except Exception:
            pass
        try:
            el.click()
            return True
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", el)
                return True
            except Exception as e:
                raise e

    # 1) Try direct strategies
    for kind, selector in strategies:
        try:
            if kind == "css":
                el = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
                if try_click(el):
                    print(f"[OK] Clicked via CSS: {selector}")
                    return True
            elif kind == "xpath":
                el = wait.until(EC.presence_of_element_located((By.XPATH, selector)))
                if try_click(el):
                    print(f"[OK] Clicked via XPath: {selector}")
                    return True
            elif kind == "js":
                # run script to find the best candidate and click via JS
                js = """
                const nodes = document.querySelectorAll("a[data-bs-title='Service Management']");
                for (const n of nodes) {
                    if (n.innerText && n.innerText.trim().includes('Service Management')) {
                        n.scrollIntoView({block:'center'});
                        n.click();
                        return true;
                    }
                }
                return nodes.length > 0;
                """
                clicked = driver.execute_script(js)
                if clicked:
                    print("[OK] Clicked via JS selector")
                    return True
        except Exception as e:
            last_err = e
            # store debug HTML for this selector
            try:
                snippet = driver.execute_script("return (arguments[0] || document.body).outerHTML;", el) if 'el' in locals() else driver.page_source[:5000]
                print(f"[WARN] strategy {kind} failed: {e}\nHTML snippet preview:\n{snippet[:1000]}")
            except Exception:
                pass

    # 2) Check if inside iframe — try switching to any iframe that contains the text
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for i, frm in enumerate(iframes):
            try:
                driver.switch_to.frame(frm)
                found = driver.find_elements(By.XPATH, "//*[contains(normalize-space(.), 'Service Management')]")
                if found:
                    print(f"[INFO] Found 'Service Management' inside iframe index {i}. Attempting click inside iframe.")
                    el = found[0]
                    try:
                        if try_click(el):
                            print(f"[OK] Clicked inside iframe index {i}")
                            driver.switch_to.default_content()
                            return True
                    except Exception as e:
                        last_err = e
                driver.switch_to.default_content()
            except Exception:
                driver.switch_to.default_content()
                continue
    except Exception as e:
        last_err = e

    # 3) Shadow DOM hint: we can't pierce shadow DOM easily — try JS to query deeper
    try:
        js_deep = """
        const q = (root) => {
          const els = [];
          function walk(node) {
            if (!node) return;
            if (node.nodeType === 1 && node.textContent && node.textContent.includes('Service Management')) els.push(node);
            const shadow = node.shadowRoot;
            if (shadow) walk(shadow);
            for (const child of node.children) walk(child);
          }
          walk(document);
          return els.slice(0,5).map(e => ({tag:e.tagName, txt: e.innerText ? e.innerText.trim().slice(0,100) : ''}));
        };
        return q(document);
        """
        shadow_matches = driver.execute_script(js_deep)
        if shadow_matches:
            print("[INFO] Potential matches inside shadow DOM or nested DOM (displaying upto 5):")
            for m in shadow_matches:
                print(m)
            print("[ACTION] If element is inside Shadow DOM, use JS to click it or a specialized helper (see notes).")
    except Exception:
        pass

    # take screenshot for offline debugging
    try:
        driver.save_screenshot(screenshot_path)
        print(f"[DEBUG] Screenshot saved to {screenshot_path}")
    except Exception:
        pass

    # final fallback: raise with helpful info
    print("[ERROR] All strategies failed. Last exception:")
    traceback.print_exception(type(last_err), last_err, last_err.__traceback__)
    raise RuntimeError("Unable to locate/click Service Management link — see debug output and screenshot.")




def random_warehouse():
    warehouse_data = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\warehouse_data.xlsx"
    
    if not os.path.exists(warehouse_data):
        raise FileNotFoundError(f"Excel file not found at: {warehouse_data}")
    
    # Read the Excel file
    df = pd.read_excel(warehouse_data)
    
    # Pick a random row
    row = df.sample(n=1).iloc[0]
    
    # Return dictionary with warehouse info
    return {
        "Warehouse Name": row["Warehouse Name"],
        "Contact Person": row["Contact Person"],
        "Contact No.": row["Contact No."],
        "Location Name": row["Location Name"],
        "House/Floor/Unit No.": row["House/Floor/Unit No."],
        "Block/Building/Street": row["Block/Building/Street"],
        "Country": row["Country"],
        "Province": row["Province"],
        "City/Municipality": row["City/Municipality"],
        "Barangay": row["Barangay"],
        "Postal Code": row["Postal Code"]
    }
def upload_image_warehouse(driver, button_id="Photo",
                            image_folder=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Images\Warehouse"):
    # Collect random image
    images = [os.path.join(image_folder, f) for f in os.listdir(image_folder)
              if f.lower().endswith((".png", ".jpg", ".jpeg"))]
    if not images:
        raise FileNotFoundError(f"No images found in {image_folder}")
    random_image = random.choice(images)

    # Click the upload button to open dialog
    upload_btn = driver.find_element(By.ID, button_id)
    driver.execute_script("arguments[0].click()", upload_btn)
    time.sleep(2)  # wait for dialog to appear

    # Type path + press Enter
    pyautogui.write(random_image)
    pyautogui.press("enter")

    return random_image

#Warehouse dropdown
def select_dropdown(driver, wait, container_id, value):
    dropdown = driver.find_element(By.ID, container_id)
    dropdown.click()
    search_input = wait.until(
        lambda d: d.find_element(By.XPATH, "//input[@class='select2-search__field']")
    )
    search_input.clear()
    search_input.send_keys(value)
    search_input.send_keys(Keys.ENTER)

def parse_table(driver, table_xpath, log_file_path, table_name):
    from selenium.common.exceptions import (
        StaleElementReferenceException,
        TimeoutException,
        NoSuchElementException
    )
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time, traceback

    max_retries = 3

    for attempt in range(max_retries):
        try:
            # --- Wait for the table body to exist before accessing ---
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, table_xpath))
            )

            table_body = driver.find_element(By.XPATH, table_xpath)
            rows = table_body.find_elements(By.TAG_NAME, "tr")

            # --- If table is empty, log and return ---
            if not rows:
                print(f"[DEBUG] Table is empty ({table_name}).")
                log_action(f"Table is empty ({table_name})", log_file_path=log_file_path)
                return []

            valid_rows = []
            for i, row in enumerate(rows, start=1):
                try:
                    # Re-find row each iteration (avoid stale issues)
                    table_body = driver.find_element(By.XPATH, table_xpath)
                    current_row = table_body.find_elements(By.TAG_NAME, "tr")[i - 1]
                    cols = current_row.find_elements(By.TAG_NAME, "td")

                    row_data = []
                    for col in cols:
                        try:
                            text = col.text.strip()
                            if text:
                                row_data.append(text)
                        except StaleElementReferenceException:
                            continue

                    if row_data:
                        valid_rows.append(row_data)
                        row_text = " | ".join(row_data)
                        print(f"[DEBUG] {table_name} - Row {i}: {row_text}")
                        log_action(f"{table_name} - Row {i}: {row_text}", log_file_path=log_file_path)

                except (StaleElementReferenceException, IndexError):
                    continue

            print(f"[DEBUG] Table has {len(valid_rows)} valid rows ({table_name}).")
            log_action(f"Table has {len(valid_rows)} valid rows ({table_name})", log_file_path=log_file_path)
            return valid_rows

        except TimeoutException:
            # Table didn’t appear at all
            msg = f"Timeout waiting for table {table_name} (XPath: {table_xpath})"
            log_error(msg, log_file_path=log_file_path)
            print(f"[DEBUG] {msg}")
            return []

        except NoSuchElementException:
            # Table element missing entirely
            msg = f"{table_name} not found (XPath: {table_xpath})"
            log_error(msg, log_file_path=log_file_path)
            print(f"[DEBUG] {msg}")
            return []

        except StaleElementReferenceException:
            if attempt < max_retries - 1:
                print(f"[DEBUG] Stale element on attempt {attempt + 1}, retrying...")
                time.sleep(0.5)
                continue
            else:
                msg = f"Stale element in {table_name} after {max_retries} retries"
                log_error(msg, log_file_path=log_file_path)
                print(f"[DEBUG] {msg}")
                return []

        except Exception:
            msg = f"Error parsing {table_name}: {traceback.format_exc()}"
            log_error(msg, log_file_path=log_file_path)
            print(f"[DEBUG] {msg}")
            return []

    return []  # fallback if all retries fail

# Seller Center Product
def select_product(driver, product_name, 
                                  wait_time=10, 
                                  screenshots_folder=None, 
                                  log_file_path=None):
    try:
        # Step 1: Type product name in search box
        search_input = WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.NAME, "ORDER_productSearchInput"))
        )
        search_input.clear()
        human_like_typing(search_input, product_name)
        
        if log_file_path:
            log_action(f"Selected Product: {product_name}", log_file_path=log_file_path)
        
        time.sleep(1)  # Brief pause for dropdown to render
        
        try:
            dropdown_item = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, '#ORDER_productDataList li, [role="option"]')
                )
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", dropdown_item)
            driver.execute_script("arguments[0].click();", dropdown_item)
            
        except:
            # Strategy 2: Find by text containing product name
            dropdown_item = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable(
                    (By.XPATH, f'//li[contains(text(), "{product_name}")]')
                )
            )
            driver.execute_script("arguments[0].click();", dropdown_item)
        
        if log_file_path:
            log_action(f"Clicked on product dropdown item: {product_name}", log_file_path=log_file_path)
        
        # Step 4: Wait for selection to complete
        time.sleep(1)
        WebDriverWait(driver, 30).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        # Step 5: Take screenshot if folder provided
        if screenshots_folder:
            time.sleep(10)
            driver.save_screenshot(
                os.path.join(screenshots_folder, f"Product_Selected_{product_name.replace(' ', '_')}.png")
            )
        
        if log_file_path:
            log_action(f"Product confirmed: {product_name}", log_file_path=log_file_path)
        
        return True
        
    except Exception as e:
        if log_file_path:
            log_action(f"Error selecting product '{product_name}': {str(e)}", log_file_path=log_file_path)
        
        # Take error screenshot
        if screenshots_folder:
            try:
                driver.save_screenshot(
                    os.path.join(screenshots_folder, f"ERROR_Product_Selection_{product_name.replace(' ', '_')}.png")
                )
            except:
                pass
        
        return False
    
# Service Center Services
def select_services(driver, service_name, 
                                  wait_time=10, 
                                  screenshots_folder=None, 
                                  log_file_path=None):
    try:
        # Step 1: Type product name in search box
        search_input = WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.NAME, "SERVICES_productSearchInput"))
        )
        search_input.clear()
        human_like_typing(search_input, service_name)
        
        if log_file_path:
            log_action(f"Selected Product: {service_name}", log_file_path=log_file_path)
        
        time.sleep(1)  # Brief pause for dropdown to render
        
        try:
            dropdown_item = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, 'By.CSS_SELECTOR, "#SERVICES_productDataList, .product-data-list')
                )
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", dropdown_item)
            driver.execute_script("arguments[0].click();", dropdown_item)
            
        except:
            # Strategy 2: Find by text containing product name
            dropdown_item = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable(
                    (By.XPATH, f'//li[contains(text(), "{service_name}")]')
                )
            )
            driver.execute_script("arguments[0].click();", dropdown_item)
        
        if log_file_path:
            log_action(f"Clicked on service dropdown item: {service_name}", log_file_path=log_file_path)
        
        # Step 4: Wait for selection to complete
        time.sleep(1)
        WebDriverWait(driver, 30).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        # Step 5: Take screenshot if folder provided
        if screenshots_folder:
            time.sleep(10)
            driver.save_screenshot(
                os.path.join(screenshots_folder, f"Service_Selected_{service_name.replace(' ', '_')}.png")
            )
        
        if log_file_path:
            log_action(f"service confirmed: {service_name}", log_file_path=log_file_path)
        
        return True
        
    except Exception as e:
        if log_file_path:
            log_action(f"Error selecting service '{service_name}': {str(e)}", log_file_path=log_file_path)
        
        # Take error screenshot
        if screenshots_folder:
            try:
                driver.save_screenshot(
                    os.path.join(screenshots_folder, f"ERROR_Service_Selection_{service_name.replace(' ', '_')}.png")
                )
            except:
                pass
        
        return False

def dashboard(driver, wait, log_file_path):
        element = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/header/div/nav/ul/li[1]/a")))
        driver.execute_script("arguments[0].click()", element)
        log_action("Clicked Main Dashboard on navbar", log_file_path=log_file_path)


# Create New Service - Service Management
def random_service():
    service_data_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\services_data.xlsx"
    
    if not os.path.exists(service_data_path):
        raise FileNotFoundError(f"Excel file not found at: {service_data_path}")
    
    # Read the Excel file
    df = pd.read_excel(service_data_path)
    
    # Define all required columns
    required_columns = [
        "Service Name", 
        "SKU", 
        "Units of Measure", 
        "Service Description", 
        "Location", 
        "Unit Price", 
        "Service Tags", 
        "Service Value", 
        "Service Sub-Category"
    ]
    
    # Keep only rows where ALL required columns have non-empty values
    valid_df = df.copy()
    
    for col in required_columns:
        if col in valid_df.columns:
            # Remove rows where column is NaN, empty string, or just whitespace
            valid_df = valid_df[
                valid_df[col].notna() & 
                (valid_df[col].astype(str).str.strip() != "") &
                (valid_df[col].astype(str).str.strip() != "nan")
            ]
    
    # Check if we have any valid rows left
    if valid_df.empty:
        raise ValueError(
            f"❌ No valid service data found in Excel!\n"
            f"Total rows in file: {len(df)}\n"
            f"Valid rows with all fields filled: 0\n"
            f"Please ensure all required columns have values: {required_columns}"
        )
    
    print(f"✅ Found {len(valid_df)} valid services out of {len(df)} total rows")
    
    # Pick a random row from the valid data
    row = valid_df.sample(n=1).iloc[0]
    
    # Helper function to safely convert to string
    def safe_str(value):
        """Convert value to string and strip whitespace"""
        return str(value).strip()
    
    # Trim Service Description to 50 characters
    service_description = safe_str(row["Service Description"])[:50]
    
    # Return dictionary with service info
    return {
        "Service Name": safe_str(row["Service Name"]),
        "SKU": safe_str(row["SKU"]),
        "Units of Measure": safe_str(row["Units of Measure"]),
        "Service Description": service_description,
        "Location": safe_str(row["Location"]),
        "Unit Price": safe_str(row["Unit Price"]),
        "Service Value": safe_str(row["Service Value"]),
        "Service Tags": safe_str(row["Service Tags"]),
        "Service Sub-Category": safe_str(row["Service Sub-Category"])
    }


# --- FIND TARGET ROW MSERVICE---
def find_service_row(driver, service_name):
    rows = driver.find_elements(By.CSS_SELECTOR, "#servicemanagemenTable tbody tr")
    for row in rows:
        if service_name.lower() in row.text.lower():
            return row
        return None    

# Employee Management
def find_employee_row(driver, employee_email, timeout=10, log_file_path=None):
    try:
        # After search filtering, wait for at least one row to be present
        rows_locator = (By.CSS_SELECTOR, "#employeeListTable tbody tr")
        WebDriverWait(driver, timeout).until(EC.presence_of_all_elements_located(rows_locator))
        
        # Get all visible rows
        rows = driver.find_elements(*rows_locator)
        
        if log_file_path:
            log_action(f"Found {len(rows)} row(s) after search filter", log_file_path=log_file_path)
        
        # Check if we have rows
        if not rows:
            if log_file_path:
                log_action("No rows found in table", log_file_path=log_file_path)
            return None
        
        # Check for "no results" message
        first_row = rows[0]
        first_row_text = first_row.text.lower()
        if "no" in first_row_text and ("result" in first_row_text or "found" in first_row_text):
            if log_file_path:
                log_action(f"'No results' message found: {first_row.text}", log_file_path=log_file_path)
            return None
        
        # If search is working correctly, first row should be our employee
        if log_file_path:
            log_action(f"Returning first row as target employee", log_file_path=log_file_path)
        
        return first_row
        
    except Exception as e:
        if log_file_path:
            log_error(f"Error finding employee row: {str(e)}", log_file_path=log_file_path, driver=driver)

# Employee Management
def search_and_select_employee(driver, wait, employee_email, log_file_path, screenshots_folder, max_retries=3):
    
    for attempt in range(max_retries):
        try:
            log_action(f"Search attempt {attempt + 1}/{max_retries} for email: {employee_email}", 
                      log_file_path=log_file_path)
            
            # Refresh the page on retry attempts to get fresh data
            if attempt > 0:
                log_action(f"Refreshing page before retry {attempt + 1}", log_file_path=log_file_path)
                driver.refresh()
                WebDriverWait(driver, 30).until(lambda d: d.execute_script('return document.readyState') == 'complete')
                time.sleep(2)
                # Wait for table to reload
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//*[@id='employeeListTable']/tbody/tr"))
                )
            
            # Search for employee
            search_box = wait.until(EC.presence_of_element_located((By.ID, "searchTable")))
            search_box.clear()
            human_like_typing(search_box, employee_email)
            time.sleep(3)  # Increased wait for table filtering
            
            # Debug: Count visible rows
            all_rows = driver.find_elements(By.XPATH, "//*[@id='employeeListTable']/tbody/tr")
            log_action(f"Found {len(all_rows)} row(s) in table", log_file_path=log_file_path)
            
            driver.save_screenshot(os.path.join(screenshots_folder, f"Search_Attempt_{attempt + 1}.png"))
            
            # Try to find the employee
            target_row = find_employee_row(driver, employee_email, log_file_path=log_file_path)
            
            if target_row:
                log_action(f"Employee found on attempt {attempt + 1}", log_file_path=log_file_path)
                
                # Click checkbox
                checkbox = target_row.find_element(By.CSS_SELECTOR, "th .form-check-input[type='checkbox']")
                driver.execute_script("arguments[0].click();", checkbox)
                log_action("Selected employee checkbox", log_file_path=log_file_path)
                
                return target_row
            
            # If not found and not last attempt, wait before retry
            if attempt < max_retries - 1:
                log_action(f"Employee not found, waiting 5 seconds before retry...", log_file_path=log_file_path)
                time.sleep(5)
        
        except Exception as e:
            if attempt == max_retries - 1:
                # Last attempt failed, raise the error
                raise
            else:
                log_action(f"Error on attempt {attempt + 1}: {str(e)}", log_file_path=log_file_path)
                time.sleep(5)
    
    # If we get here, all retries failed
    driver.save_screenshot(os.path.join(screenshots_folder, f"Email_not_found_{employee_email}_FINAL.png"))
    raise Exception(
        f"Email '{employee_email}' not found in table after {max_retries} attempts. "
        f"Employee was created but may not be synced to the list view yet."
    )


# Client Directory
def find_client_row(driver, client_email):
    rows = driver.find_elements(By.CSS_SELECTOR, "#clientDirectoryTable tbody tr")
    for row in rows:
        if client_email in row.text:
            return row
    return None

def upload_image_service(driver, log_file_path,
                         input_locator,  # Locator (tuple) for the file input
                         image_folder=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Images\Warehouse",
                         timeout=10):
    """
    Uploads a random image via a file input.

    :param input_locator: A selenium locator tuple, e.g. (By.CSS_SELECTOR, "input[type='file']")
    """
    # Validate folder
    if not os.path.isdir(image_folder):
        raise FileNotFoundError(f"Image folder not found: {image_folder}")

    # List image files
    images = [
        os.path.join(image_folder, f)
        for f in os.listdir(image_folder)
        if f.lower().endswith((".png", ".jpg", ".jpeg"))
    ]
    if not images:
        raise FileNotFoundError(f"No images found in {image_folder}")

    random_image = random.choice(images)

    try:
        # Wait for the file input to be present
        file_input = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(input_locator)
        )

        # If it's hidden, make it visible via JS
        driver.execute_script("arguments[0].style.display = 'block';", file_input)
        driver.execute_script("arguments[0].style.visibility = 'visible';", file_input)

        # Upload the file by sending the path
        file_input.send_keys(random_image)

        log_action(f"upload_image_service: uploaded {random_image}", log_file_path=log_file_path)
        return random_image

    except Exception as e:
        err = f"upload_image_service failed with image {random_image}: {traceback.format_exc()}"
        log_error(err, log_file_path=log_file_path, driver=driver)
        raise

def upload_file_service(driver, log_file_path, 
                        button_id=None,
                        file_path=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\Service Batch Upload.xlsx",
                        timeout=10):

    # Check if file_path is a file or directory
    if os.path.isfile(file_path):
        selected_file = file_path
    elif os.path.isdir(file_path):
        # If it's a directory, find supported files
        files = [
            os.path.join(file_path, f)
            for f in os.listdir(file_path)
            if f.lower().endswith((".pdf", ".docx", ".txt", ".xlsx", ".xls"))
        ]
        if not files:
            raise FileNotFoundError(f"No supported files found in {file_path}")
        selected_file = random.choice(files)
    else:
        raise FileNotFoundError(f"File or directory not found: {file_path}")

    # Verify the selected file exists
    if not os.path.exists(selected_file):
        raise FileNotFoundError(f"Selected file not found: {selected_file}")

    try:
        # Try to find and use file input directly first (most reliable method)
        try:
            file_input = driver.find_element(By.XPATH, "//input[@type='file']")
            file_input.send_keys(selected_file)
            log_action(f"upload_file_service: uploaded {selected_file} via file input", log_file_path=log_file_path)
            return selected_file
        except Exception:
            pass

        # If button_id is provided, try clicking it
        if button_id:
            try:
                upload_btn = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((By.ID, button_id))
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", upload_btn)
                driver.execute_script("arguments[0].click()", upload_btn)
                log_action(f"Clicked upload button with ID: {button_id}", log_file_path=log_file_path)
            except Exception:
                pass

        # Try clicking the "Click to upload" span
        try:
            upload_span = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, "//span[@class='click' and normalize-space(text())='Click to upload']"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", upload_span)
            try:
                upload_span.click()
            except Exception:
                driver.execute_script("arguments[0].click();", upload_span)
            log_action("Clicked 'Click to upload' span", log_file_path=log_file_path)
        except Exception:
            # Try alternative upload trigger elements
            alternative_triggers = [
                "//div[contains(@class, 'upload')]",
                "//button[contains(text(), 'Browse')]",
                "//div[contains(@class, 'filepond')]",
                "//*[contains(text(), 'upload') or contains(text(), 'Upload')]"
            ]
            
            trigger_found = False
            for trigger_xpath in alternative_triggers:
                try:
                    trigger_element = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, trigger_xpath))
                    )
                    driver.execute_script("arguments[0].click();", trigger_element)
                    log_action(f"Clicked upload trigger: {trigger_xpath}", log_file_path=log_file_path)
                    trigger_found = True
                    break
                except Exception:
                    continue
            
            if not trigger_found:
                raise Exception("Could not find any upload trigger element")

        # Wait for file dialog and use pyautogui
        time.sleep(2)
        pyautogui.write(selected_file)
        time.sleep(1)
        pyautogui.press("enter")
        time.sleep(1)

        log_action(f"upload_file_service: uploaded {selected_file} via file dialog", log_file_path=log_file_path)
        return selected_file

    except Exception as e:
        err = f"upload_file_service failed with file {selected_file}: {traceback.format_exc()}"
        log_error(err, log_file_path=log_file_path, driver=driver)
        raise

#Update Contact & Shipping    
def update_profile_data():
    # file path — change to where you saved the Excel
    sample_data_file = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\update_profile_data.xlsx"
    
    if not os.path.exists(sample_data_file):
        raise FileNotFoundError(f"Excel file not found at: {sample_data_file}")
    
    # Read the Excel file into DataFrame
    df = pd.read_excel(sample_data_file)
    
    # Optional: check that expected columns are there
    expected_cols = [
        "Mobile No.",
        "Country",
        "Province",
        "City/Municipality",
        "Barangay",
        "House/Floor/Unit No.",
        "Block/Building/Street",
        "Postal Code"
    ]
    for col in expected_cols:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found in Excel file")
    
    # Pick a random row
    row = df.sample(n=1).iloc[0]
    
    # Return dictionary with address data
    return {
        "Mobile No.": row["Mobile No."],
        "Country": row["Country"],
        "Province": row["Province"],
        "City/Municipality": row["City/Municipality"],
        "Barangay": row["Barangay"],
        "House/Floor/Unit No.": row["House/Floor/Unit No."],
        "Block/Building/Street": row["Block/Building/Street"],
        "Postal Code": str(row["Postal Code"])  # convert to string if needed
    }

# Setup Restaurant Table
def new_table(excel_path, row_number=0):
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\table_data.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    df = pd.read_excel(excel_path)
    row = df.sample(n=1).iloc[0]
    
    # Return as dictionary with cleaned values
    return {
        'Table': row.get(['Table #/Name']),
        'Pax': row.get(['No. of Pax']),
        'Table Floor/Area': row.get(['Table Floor/Area'])
    }

def read_excel_data(file_path, sheet_name=None):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df.fillna("").to_dict(orient="records")


# Client Directory
def add_new_client(driver, log_file_path, screenshots_folder):
    
     # file path — change to where you saved the Excel
    sample_data_file = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\new_client.xlsx"
    
    if not os.path.exists(sample_data_file):
        raise FileNotFoundError(f"Excel file not found at: {sample_data_file}")
    
    # Read the Excel file into DataFrame
    df = pd.read_excel(sample_data_file)
    
    # Optional: check that expected columns are there
    expected_cols = [
        "Prefix",
        "First Name",
        "Last Name",
        "Company Name (Optional)",
        "Email Address (Optional)",
        "Mobile (Optional)",
    ]
    for col in expected_cols:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found in Excel file")
    
    # Pick a random row
    row = df.sample(n=1).iloc[0]
    excel_prefix = row["Prefix"]  # Take prefix from Excel
    
    # Select the prefix in the dropdown
    prefix_dropdown = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.ID, "contactPrefix"))
    )
    select = Select(prefix_dropdown)
    # Make sure the Excel value exists in the dropdown
    option_values = [o.get_attribute("value") for o in select.options]
    if excel_prefix not in option_values:
        raise ValueError(f"Excel prefix '{excel_prefix}' not found in dropdown options")
    
    select.select_by_value(excel_prefix)
    log_action(f"Selected prefix from Excel: {excel_prefix}", log_file_path=log_file_path)
    time.sleep(1)
    driver.save_screenshot(os.path.join(screenshots_folder, f"Prefix_{excel_prefix}.png"))
    
    # Return a dictionary with client data
    return {
        "Prefix": excel_prefix,
        "First Name": row["First Name"],
        "Last Name": row["Last Name"],
        "Company Name (Optional)": row["Company Name (Optional)"],
        "Email Address (Optional)": row["Email Address (Optional)"],
         "Mobile (Optional)": str(row["Mobile (Optional)"])  
    }


# Client List
def add_new_address(driver, log_file_path, screenshots_folder):
    # Excel file path
    sample_data_file = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\client_address.xlsx"
    
    if not os.path.exists(sample_data_file):
        raise FileNotFoundError(f"Excel file not found at: {sample_data_file}")
    
    # Read Excel
    df = pd.read_excel(sample_data_file)
    
    # Expected columns
    expected_cols = [
        "HOUSE / FLOOR / UNIT NO",
        "BLOCK / BLDG / STREET",
        "COUNTRY",
        "PROVINCE",
        "CITY/MUNICIPALITY",
        "BARANGAY",
        "POSTAL CODE",
        "OTHERS"
    ]
    for col in expected_cols:
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found in Excel file")
    
    # Pick a random row
    row = df.sample(n=1).iloc[0]
    
    # Wait for modal/form to appear
    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".modal-body")))
    log_action("Modal visible for adding address", log_file_path=log_file_path)
    
    
    # Return the filled data for logging
    return {
        "HOUSE / FLOOR / UNIT NO": row["HOUSE / FLOOR / UNIT NO"],
        "BLOCK / BLDG / STREET": row["BLOCK / BLDG / STREET"],
        "COUNTRY": row["COUNTRY"],
        "PROVINCE": row["PROVINCE"],
        "CITY/MUNICIPALITY": row["CITY/MUNICIPALITY"],
        "BARANGAY": row["BARANGAY"],
        "POSTAL CODE": str(row["POSTAL CODE"]),
        "OTHERS": row["OTHERS"] if not pd.isna(row["OTHERS"]) else ""
    }
    


# Client List
def fill_up_address(driver, trigger_locator, option_text, timeout=10, max_retries=3):
    wait = WebDriverWait(driver, timeout)

    for attempt in range(max_retries):
        try:
            log_action(f"Opening Select2 dropdown via {trigger_locator} (Attempt {attempt + 1}/{max_retries})")
            
            # Click the dropdown trigger
            trigger = wait.until(EC.element_to_be_clickable(trigger_locator))
            trigger.click()
            time.sleep(0.3)  # Increased wait for dropdown to fully open

            # Wait for search input
            search_input = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input.select2-search__field"))
            )
            search_input.clear()
            time.sleep(0.1)

            # Type the text
            search_input.send_keys(option_text)
            log_action(f"Typed '{option_text}' into search field")
            time.sleep(0.5)  # Wait for dropdown results to load

            # Wait for the dropdown option to appear
            option_locator = (
                By.XPATH,
                f"//li[contains(@class,'select2-results__option')][normalize-space()='{option_text}']"
            )
            
            try:
                option_elem = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located(option_locator)
                )
                log_action(f"Found option in dropdown: '{option_text}'")
            except:
                log_error(f"Option '{option_text}' not found in dropdown, trying Enter key")
                search_input.send_keys(Keys.ENTER)
                time.sleep(0.3)
                
                # Verify if Enter worked
                try:
                    # Check if dropdown closed (search input should be gone)
                    WebDriverWait(driver, 2).until(
                        EC.invisibility_of_element_located((By.CSS_SELECTOR, "input.select2-search__field"))
                    )
                    log_action(f"Option selected via Enter key")
                    return True
                except:
                    log_error(f"Enter key did not select option")
                    if attempt < max_retries - 1:
                        continue
                    return False

            # Click the option
            try:
                option_elem.click()
                log_action(f"Clicked option '{option_text}'")
            except:
                log_action(f"Click failed, using Enter fallback")
                search_input.send_keys(Keys.ENTER)

            # Wait for dropdown to close
            time.sleep(0.3)
            
            # VERIFY SELECTION
            try:
                # Wait for the dropdown to close
                WebDriverWait(driver, 3).until(
                    EC.invisibility_of_element_located((By.CSS_SELECTOR, "input.select2-search__field"))
                )
                
                # Get the selected value from trigger
                time.sleep(0.2)
                trigger_element = driver.find_element(*trigger_locator)
                selected_value = trigger_element.text.strip()
                
                log_action(f"Trigger shows: '{selected_value}'")
                
                # Check if selection matches (flexible matching)
                if option_text.lower() in selected_value.lower() or selected_value.lower() in option_text.lower():
                    log_action(f"✓ VERIFIED: '{option_text}' selected successfully")
                    return True
                else:
                    log_error(f"Selection mismatch: Expected '{option_text}', got '{selected_value}'")
                    if attempt < max_retries - 1:
                        log_action("Retrying selection...")
                        time.sleep(0.5)
                        continue
                    return False
                    
            except Exception as verify_error:
                log_error(f"Verification error: {verify_error}")
                if attempt < max_retries - 1:
                    continue
                return False

        except Exception as e:
            log_error(f"Error on attempt {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                log_action("Retrying...")
                time.sleep(0.5)
            else:
                log_error(f"All {max_retries} attempts failed for '{option_text}'")
                return False

    return False

# Menu Data
def menu_data():
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\menu_data.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    df = pd.read_excel(excel_path)
    row = df.sample(n=1).iloc[0]
    
    return {
        "Menu Name": row.get("Menu Name", ""),
        "SKU": row.get("SKU", ""),
        "Units of Measure": row.get("Units of Measure", ""),
        "Menu Description": row.get("Menu Description", ""),
        "Unit Price": row.get("Unit Price", ""),
        "Menu Value": row.get("Menu Value", ""),
        "Menu Tags": row.get("Menu Tags", ""),
        "Menu Sub-Category": row.get("Menu Sub-Category", ""),
        "Menu Category": row.get("Menu Category", ""),

    }
def upload_image_restaurant(
    driver,
    image_path=None,
    image_folder=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Images\Restaurant"
):
    #  1. Pick a random image
    images = [os.path.join(image_folder, f) for f in os.listdir(image_folder)
              if f.lower().endswith((".png", ".jpg", ".jpeg"))]
    if not images:
        raise FileNotFoundError(f"No images found in {image_folder}")
    random_image = random.choice(images)

    try:
        #  2. Scroll to bottom to ensure visibility
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        #  3. Locate FilePond's input[type=file]
        upload_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input.filepond--browser[type='file']"))
        )

        #  4. Scroll element into view
        driver.execute_script("arguments[0].scrollIntoView(true);", upload_input)
        time.sleep(0.5)

        # 5. Send file path directly to FilePond input
        upload_input.send_keys(random_image)
        print(f"📸 Uploaded image successfully: {os.path.basename(random_image)}")

        # 6. Wait for FilePond preview or processing to finish
        time.sleep(2)

        return random_image

    except Exception as e:
        print(f"❌ Upload failed: {e}")
        driver.save_screenshot("upload_error.png")
        return None

#-----BOOKING UTILITY-----#
def parse_preferred_time(val):
    # If it's already a datetime.time object
    if isinstance(val, datetime_time):
        hour = val.hour
        minute = val.minute
        am_pm = "AM" if hour < 12 else "PM"
        # Convert to 12-hour format
        if hour == 0:
            hour = 12
        elif hour > 12:
            hour = hour - 12
        return f"{hour:02d}", f"{minute:02d}", am_pm
    
    # If it's a string
    if isinstance(val, str):
        val = val.strip().upper()
        
        # Handle formats like "18:00", "6:00 PM", "1800"
        if ":" in val:
            # Format: "18:00" or "6:00 PM"
            time_part = val.split()[0]  # Get time before space
            hour_str, minute_str = time_part.split(":")
            hour = int(hour_str)
            minute = int(minute_str)
            
            # Check if AM/PM is specified
            if "PM" in val or "AM" in val:
                am_pm = "PM" if "PM" in val else "AM"
            else:
                # Assume 24-hour format, convert to 12-hour
                am_pm = "AM" if hour < 12 else "PM"
                if hour == 0:
                    hour = 12
                elif hour > 12:
                    hour = hour - 12
            
            return f"{hour:02d}", f"{minute:02d}", am_pm
    
    # Default fallback
    return "06", "00", "PM"

def appointment_data():
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\appointment_data.xlsx"

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")

    df = pd.read_excel(excel_path, engine='openpyxl')  # ensure engine supports time types
    row = df.sample(n=1).iloc[0]

    pref_time = row.get("Preferred Time", "")
    hour, minute, am_pm = parse_preferred_time(pref_time)

    return {
        "Customer Name": row.get("Customer Name", ""),
        "Table #": row.get("Table #", ""),
        "No. of Pax": row.get("No. of Pax", ""),
        "Preferred Time": pref_time,
        "Hour": hour,
        "Minute": minute,
        "AM_PM": am_pm,
        "Notes": row.get("Notes", ""),
    }
def set_preferred_time(driver, time_input_selector, preferred_time, timeout=10):
    hour, minute, am_pm = parse_preferred_time(preferred_time)
    if hour is None:
        print(f"[set_preferred_time] Could not parse time: {preferred_time}")
        return False

    wait = WebDriverWait(driver, timeout)

    try:
        time_input = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, time_input_selector)))
    except Exception as e:
        print(f"[set_preferred_time] Time input not found or not clickable: {e}")
        return False

    # Try clear
    try:
        time_input.clear()
    except Exception:
        pass

    # Try typing in
    try:
        human_like_typing(time_input, f"{hour}:{minute} {am_pm}")
    except Exception as e:
        print(f"[set_preferred_time] send_keys typing failed: {e}")

    # Maybe JS fallback — in case input is readonly/masked
    try:
        driver.execute_script(
            "arguments[0].value = arguments[1];"
            "arguments[0].dispatchEvent(new Event('input', { bubbles: true }));"
            "arguments[0].dispatchEvent(new Event('change', { bubbles: true }));",
            time_input,
            f"{hour}:{minute} {am_pm}"
        )
    except Exception as e:
        print(f"[set_preferred_time] JS fallback failed: {e}")

    # Optionally blur / tab out so that change is recognized
    try:
        driver.execute_script("arguments[0].blur();", time_input)
    except Exception:
        pass


from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException

def select_branch(driver, wait, element_id, option_value):
    dropdown_element = wait.until(EC.presence_of_element_located((By.ID, element_id)))
    select = Select(dropdown_element)

    try:
        # Try selecting by visible text first
        select.select_by_visible_text(option_value)
    except NoSuchElementException:
        try:
            # If not found, try selecting by value
            select.select_by_value(option_value)
        except NoSuchElementException:
            raise ValueError(f"Option '{option_value}' not found in dropdown {element_id}")
        
# Seller Center > Create Order
def random_client_product():
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\client_testdata_product.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    df = pd.read_excel(excel_path)
    row = df.sample(n=1).iloc[0]
    
    return {
        "First Name": row.get("First Name", ""),
        "Middle Name": row.get("Middle Name", ""),  # will be "" if not in Excel
        "Last Name": row.get("Last Name", ""),
        "Company Name": row.get("Company Name", ""),
        "Mobile Number": row.get("Mobile Number", ""),
        "Email": row.get("Email", ""),
        "Due Date": row.get("Due Date", ""),
        "Country": row.get("Country", ""),
        "Province": row.get("Province", ""),
        "City/Municipality": row.get("City/Municipality", ""),
        "Barangay": row.get("Barangay", ""),
        "House/Floor/Unit No.": row.get("House/Floor/Unit No.", ""),
        "Block/Building/Street": row.get("Block/Building/Street", ""),
        "Order Notes": row.get("Order Notes", ""),
        "Postal Code": row.get("Postal Code", ""),
        "Product": row.get("Product", ""),
    }

# Service Center > Create Service Order
def random_client_service():
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\client_testdata_service.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    df = pd.read_excel(excel_path)
    row = df.sample(n=1).iloc[0]
    
    return {
        "First Name": row.get("First Name", ""),
        "Middle Name": row.get("Middle Name", ""),  # will be "" if not in Excel
        "Last Name": row.get("Last Name", ""),
        "Company Name": row.get("Company Name", ""),
        "Mobile Number": row.get("Mobile Number", ""),
        "Email": row.get("Email", ""),
        "Due Date": row.get("Due Date", ""),
        "Country": row.get("Country", ""),
        "Province": row.get("Province", ""),
        "City/Municipality": row.get("City/Municipality", ""),
        "Barangay": row.get("Barangay", ""),
        "House/Floor/Unit No.": row.get("House/Floor/Unit No.", ""),
        "Block/Building/Street": row.get("Block/Building/Street", ""),
        "Order Notes": row.get("Order Notes", ""),
        "Postal Code": row.get("Postal Code", ""),
        "Product": row.get("Product", ""),
    }


def upload_file_bulk_employee(driver, log_file_path, 
                        button_id=None,
                        file_path=r"D:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\SCD_BULK_ENDCUSTOMER.xlsx",
                        timeout=10):

    # Check if file_path is a file or directory
    if os.path.isfile(file_path):
        selected_file = file_path
    elif os.path.isdir(file_path):
        # If it's a directory, find supported files
        files = [
            os.path.join(file_path, f)
            for f in os.listdir(file_path)
            if f.lower().endswith((".pdf", ".docx", ".txt", ".xlsx", ".xls"))
        ]
        if not files:
            raise FileNotFoundError(f"No supported files found in {file_path}")
        selected_file = random.choice(files)
    else:
        raise FileNotFoundError(f"File or directory not found: {file_path}")

    # Verify the selected file exists
    if not os.path.exists(selected_file):
        raise FileNotFoundError(f"Selected file not found: {selected_file}")

    try:
        # Try to find and use file input directly first (most reliable method)
        try:
            file_input = driver.find_element(By.XPATH, "//input[@type='file']")
            file_input.send_keys(selected_file)
            log_action(f"upload_file_service: uploaded {selected_file} via file input", log_file_path=log_file_path)
            return selected_file
        except Exception:
            pass

        # If button_id is provided, try clicking it
        if button_id:
            try:
                upload_btn = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((By.ID, button_id))
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", upload_btn)
                driver.execute_script("arguments[0].click()", upload_btn)
                log_action(f"Clicked upload button with ID: {button_id}", log_file_path=log_file_path)
            except Exception:
                pass

        # Try clicking the "Click to upload" span
        try:
            upload_span = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, "//span[@class='click' and normalize-space(text())='Click to upload']"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", upload_span)
            try:
                upload_span.click()
            except Exception:
                driver.execute_script("arguments[0].click();", upload_span)
            log_action("Clicked 'Click to upload' span", log_file_path=log_file_path)
        except Exception:
            # Try alternative upload trigger elements
            alternative_triggers = [
                "//div[contains(@class, 'upload')]",
                "//button[contains(text(), 'Browse')]",
                "//div[contains(@class, 'filepond')]",
                "//*[contains(text(), 'upload') or contains(text(), 'Upload')]"
            ]
            
            trigger_found = False
            for trigger_xpath in alternative_triggers:
                try:
                    trigger_element = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, trigger_xpath))
                    )
                    driver.execute_script("arguments[0].click();", trigger_element)
                    log_action(f"Clicked upload trigger: {trigger_xpath}", log_file_path=log_file_path)
                    trigger_found = True
                    break
                except Exception:
                    continue
            
            if not trigger_found:
                raise Exception("Could not find any upload trigger element")

        # Wait for file dialog and use pyautogui
        time.sleep(2)
        pyautogui.write(selected_file)
        time.sleep(1)
        pyautogui.press("enter")
        time.sleep(1)

        log_action(f"upload_file_service: uploaded {selected_file} via file dialog", log_file_path=log_file_path)
        return selected_file

    except Exception as e:
        err = f"upload_file_service failed with file {selected_file}: {traceback.format_exc()}"
        log_error(err, log_file_path=log_file_path, driver=driver)
        raise


def choose_gray_table(driver, timeout=10, table_number=None):

    print(f"🔍 Looking for gray tables (specific: {table_number or 'any'})...")
    
    WebDriverWait(driver, timeout).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.table-tile"))
    )
    
    # Give tables time to update their color attributes
    time.sleep(2)

    tiles = driver.find_elements(By.CSS_SELECTOR, "div.table-tile")
    print(f"📊 Found {len(tiles)} total table tiles")
    
    gray_tables = []
    all_tables_debug = []

    for idx, tile in enumerate(tiles):
        try:
            color_attr = (tile.get_attribute("data-tile-color") or "").strip().lower()
            
            # Try to get table number
            try:
                num_elem = tile.find_element(By.CLASS_NAME, "table-tile__number")
                num = num_elem.text.strip()
            except Exception:
                # Fallback: try other possible locations
                try:
                    num = tile.get_attribute("data-table-number") or "Unknown"
                except:
                    num = f"Unknown_{idx}"
            
            # Store for debugging
            all_tables_debug.append(f"Table {num}: {color_attr}")
            
            if color_attr == "gray":
                gray_tables.append((tile, num))
                print(f"  ✅ Gray table found: {num}")
            
        except Exception as e:
            print(f"  ⚠️ Error processing table {idx}: {e}")
            continue

    # Debug output
    print(f"\n📋 All tables status:")
    for status in all_tables_debug:
        print(f"  {status}")
    
    print(f"\n✅ Gray (available) tables: {len(gray_tables)}")
    if gray_tables:
        gray_numbers = [num for _, num in gray_tables]
        print(f"   Numbers: {gray_numbers}")

    if not gray_tables:
        raise ValueError("❌ No gray-colored tables found. All tables may be occupied.")

    # 🎯 If a specific table number is given, find it
    if table_number:
        table_number_str = str(table_number).strip()
        print(f"\n🎯 Searching for specific table: '{table_number_str}'")
        
        for tile, num in gray_tables:
            # Flexible matching: strip whitespace and compare
            if num.strip() == table_number_str:
                print(f"✅ Found matching table: {num}")
                return tile, num
        
        # Not found - provide helpful error
        available = [num for _, num in gray_tables]
        error_msg = (
            f"❌ Gray table '{table_number}' not found.\n"
            f"   Available gray tables: {available}\n"
            f"   Hint: Table {table_number} may be occupied or doesn't exist."
        )
        raise ValueError(error_msg)

    # Otherwise, pick random gray table
    chosen_tile, num = random.choice(gray_tables)
    print(f"✅ Randomly selected gray table: {num}")
    return chosen_tile, num

def get_table_number(driver, color="gray", timeout=10, retries=3, log_file_path=None):
    for attempt in range(1, retries + 1):
        try:
            table_element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, f'div.table-tile[data-tile-color="{color}"]')
                )
            )
            # Re-check if the element is still attached to the DOM
            table_number = table_element.get_attribute("data-tile-number")

            if table_number:
                msg = f" Located {color.upper()} table with number: {table_number}"
                print(msg)
                if log_file_path:
                    from Utility import log_action
                    log_action(msg, log_file_path)
                return table_number

        except (StaleElementReferenceException, NoSuchElementException):
            print(f" Attempt {attempt}/{retries}: Table element for color '{color}' became stale or missing. Retrying...")
            time.sleep(1)
            continue
        except TimeoutException:
            print(f"Timeout: Unable to locate {color} table after {timeout} seconds (Attempt {attempt}/{retries})")
            time.sleep(1)
            continue

    # If all retries fail
    error_msg = f"Failed to locate table with color '{color}' after {retries} attempts."
    print(error_msg)
    if log_file_path:
        from Utility import log_error
        log_error(error_msg, log_file_path)
    raise TimeoutException(error_msg)


def extract_table_number(table_element):

    try:
        number_elem = table_element.find_element(By.CLASS_NAME, "table-tile__number")
        table_number = number_elem.text.strip()
        return table_number if table_number else table_element
    except StaleElementReferenceException:
        # Element became stale, try fallback attribute
        try:
            table_number = table_element.get_attribute("data-tile-number")
            return table_number if table_number else table_element
        except Exception:
            return table_element
    except Exception:
        # Element not found or other error, try fallback attribute
        try:
            table_number = table_element.get_attribute("data-tile-number")
            return table_number if table_number else table_element
        except Exception:
            return table_element


def additional_order_blue(driver, gray_table_number,timeout=30):
    gray_table_number = str(gray_table_number).strip()

    def table_is_ready(driver):
        try:
            tiles = driver.find_elements(By.CSS_SELECTOR, "div.table-tile")
            for t in tiles:
                color_attr = (t.get_attribute("data-tile-color") or "").strip().lower()
                number_elem = t.find_element(By.CLASS_NAME, "table-tile__number")
                table_number = number_elem.text.strip()
                if color_attr == "blue" and table_number == gray_table_number:
                    return t 
        except Exception:
            return False
        return False

    try:
        return WebDriverWait(driver, timeout).until(table_is_ready)
    except TimeoutException:
        raise ValueError(f"No blue table found with table number {gray_table_number} within {timeout} seconds")

def close_table(driver, table_number, timeout=30):
    table_number = str(table_number).strip()

    def table_is_ready(driver):
        try:
            tiles = driver.find_elements(By.CSS_SELECTOR, "div.table-tile")
            for t in tiles:
                color_attr = (t.get_attribute("data-tile-color") or "").strip().lower()
                number_elem = t.find_element(By.CLASS_NAME, "table-tile__number")
                current_table_number = number_elem.text.strip()
                
                if color_attr == "green" and current_table_number == table_number:
                    return t 
        except Exception:
            return False
        return False

    try:
        print(f"🔍 Waiting for green table with number {table_number}...")
        table_element = WebDriverWait(driver, timeout).until(table_is_ready)
        print(f"Green table {table_number} found!")
        return table_element
    except TimeoutException:
        raise ValueError(f"No green table found with table number {table_number} within {timeout} seconds")
    
def click_table_kitchen(driver, table_number, timeout=10):
    table_number = str(table_number).strip()

    def table_is_clickable(drv):
        tiles = drv.find_elements(By.CSS_SELECTOR, "div.table-tile")
        for t in tiles:
            try:
                number_elem = t.find_element(By.CLASS_NAME, "table-tile__number")
                current_number = number_elem.text.strip()
                if current_number == table_number:
                    return t
            except (NoSuchElementException, StaleElementReferenceException):
                continue
        return False

    try:
        table_tile = WebDriverWait(driver, timeout).until(table_is_clickable)
        color_attr = (table_tile.get_attribute("data-tile-color") or "").strip().lower()
        table_tile.click()
        print(f"Clicked Table {table_number} (color: {color_attr})")
        return table_tile
    except TimeoutException:
        raise ValueError(f"Table with number {table_number} not found within {timeout} seconds")
    
def find_billing_table_id(driver, target_table_number, max_retries=3):

    target_table_number = str(target_table_number).strip()
    
    for retry in range(max_retries):
        try:
            # Get fresh list of billing cards on each retry
            billing_cards = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.billing-card"))
            )
            
            print(f" Scanning {len(billing_cards)} billing cards to match table number: {target_table_number}")
            
            # Get all billing IDs first to avoid stale references during iteration
            billing_data = []
            for index, card in enumerate(billing_cards):
                try:
                    billing_id = card.get_attribute("data-tablebilling-id")
                    if billing_id:
                        billing_data.append((index, billing_id))
                except StaleElementReferenceException:
                    continue  # Skip this one and continue with others
            
            # Now iterate through billing IDs and check table numbers
            for index, billing_id in billing_data:
                try:
                    # Re-locate the specific card by its billing ID
                    card = driver.find_element(
                        By.CSS_SELECTOR, 
                        f'div.billing-card[data-tablebilling-id="{billing_id}"]'
                    )
                    
                    # Try to find the table number within this card
                    try:
                        # Adjust selector based on your HTML structure
                        # Common patterns:
                        table_elem = card.find_element(By.CSS_SELECTOR, ".table-number, .table-name, [class*='table']")
                        table_text = table_elem.text.strip()
                        
                        # Extract number from text (handles "Table 36", "T36", "36", etc.)
                        import re
                        numbers = re.findall(r'\d+', table_text)
                        
                        if numbers and numbers[0] == target_table_number:
                            print(f" ✓ Found match! Table {target_table_number} -> Billing ID: {billing_id}")
                            return billing_id
                            
                    except NoSuchElementException:
                        # This card doesn't have a table number element
                        continue
                        
                except StaleElementReferenceException:
                    # Element became stale, but we have the billing_id, so we can continue
                    continue
                except Exception as e:
                    # Other errors, log but continue
                    continue
            
            # If we get here, no match was found on this attempt
            if retry < max_retries - 1:
                print(f" ⚠ No match found on attempt {retry + 1}, retrying...")
                time.sleep(1)
            else:
                print(f" No billing card matched table number: {target_table_number}")
                return None
                
        except Exception as e:
            if retry < max_retries - 1:
                print(f" ⚠ Error during scan (attempt {retry + 1}): {str(e)}")
                time.sleep(1)
            else:
                print(f" Error finding billing card: {str(e)}")
                return None
    
    return None

def billing_tbl(driver, gray_table_number, timeout=15, retries=3):
    gray_table_number = str(gray_table_number).strip()
    
    print(f"\n{'='*60}")
    print(f" Starting billing process for Table {gray_table_number}")
    print(f"{'='*60}\n")

    try:
        # Step 1: Wait for billing cards to load
        print(" Step 1: Waiting for billing cards to load...")
        WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.billing-card"))
        )
        print(" Billing cards loaded")
        
        # Step 2: Find the correct billing card ID
        print(f"\n Step 2: Finding billing card for table {gray_table_number}...")
        billing_id = find_billing_table_id(driver, gray_table_number)
        
        if not billing_id:
            raise Exception(f"No billing card found for table {gray_table_number}")
        
        print(f" Proceeding with billing-card ID: {billing_id}")
        
        # Step 3: Define condition to check if card is ready
        def blue_table_is_ready(driver):
            try:
                element = driver.find_element(
                    By.CSS_SELECTOR, 
                    f'div.billing-card[data-tablebilling-id="{billing_id}"]'
                )
                if element.is_displayed() and element.is_enabled():
                    return element
            except Exception:
                return False
            return False
        
        # Step 4: Click billing card with retry mechanism
        print(f"\n Step 3: Attempting to click billing card (max {retries} retries)...")
        
        for attempt in range(1, retries + 1):
            try:
                blue_card = WebDriverWait(driver, timeout).until(blue_table_is_ready)
                
                # Scroll into view
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", blue_card)
                time.sleep(1)
                
                # Click
                try:
                    blue_card.click()
                    print(f" Clicked billing card (Table {gray_table_number}, Billing ID {billing_id})")
                except Exception:
                    driver.execute_script("arguments[0].click();", blue_card)
                    print(f" Clicked billing card (Table {gray_table_number}, Billing ID {billing_id})")
                
                break
                
            except TimeoutException:
                print(f"⚠ Retry {attempt}/{retries}: Billing card not ready yet...")
                if attempt == retries:
                    raise TimeoutException(
                        f"Billing card for table {gray_table_number} not clickable after {retries} retries."
                    )
                time.sleep(2)
        
        # Step 5: Wait for network to settle
        print("\n Step 4: Waiting for page to settle...")
        time.sleep(2)
        
        # Step 6: Wait for page readiness
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        print("✓ Page ready")
        
        print(f"\n{'='*60}")
        print(f" Billing card clicked for Table {gray_table_number}")
        print(f"{'='*60}\n")
        
        return billing_id  # Return billing_id for further use
        
    except Exception as e:
        error_msg = f" Billing process failed for Table {gray_table_number}: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        raise

def click_adjustments(driver, timeout=10):
    # Wait for the “View Adjustments” link to be clickable
    view_adj = WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((By.ID, "viewAdjustments"))
    )
    # Click it
    view_adj.click()
    
    # Now wait for the adjustment list <ul> to be present (or visible)
    adj_list = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "ul.viewAdjustmentList"))
    )
    
    # Optionally, wait until it is visible (displayed)
    WebDriverWait(driver, timeout).until(
        EC.visibility_of(adj_list)
    )
    
    # Now you can assert
    assert adj_list.is_displayed(), "Adjustment list should now be visible"
    
    # Optionally, check one of its children (li) is visible or has text
    first_item = adj_list.find_element(By.TAG_NAME, "li")
    assert first_item.is_displayed(), "At least one adjustment item should be visible"
    # Optionally check the text content
    text = first_item.text
    assert "senior discount" in text.lower(), f"Unexpected adjustment text: {text}"
    
    return adj_list  # or return True / return the list element
    
def wait_for_adjustment_list(driver, timeout=10):
    
    try:
        print("Waiting for adjustments to load...")
        
        # Wait for adjustment items to appear
        WebDriverWait(driver, timeout).until(
            lambda d: len(d.find_elements(
                By.CSS_SELECTOR, 
                "li[data-billing-adjustment-seqnbr]"
            )) > 0
        )
        
        adjustments = driver.find_elements(
            By.CSS_SELECTOR, 
            "li[data-billing-adjustment-seqnbr]"
        )
        
        print(f"✓ Found {len(adjustments)} adjustment(s)")
        
        # Get details of all adjustments
        adjustment_details = []
        for adj in adjustments:
            try:
                note = adj.find_element(By.CSS_SELECTOR, "[data-billing-note]").text
                amount = adj.find_element(By.CSS_SELECTOR, "[data-billing-adjustment-amount]").text
                sign = adj.find_element(By.CSS_SELECTOR, "[data-billing-adjustment-sign]").text
                
                adjustment_details.append({
                    "note": note,
                    "amount": amount,
                    "sign": sign
                })
            except:
                pass
        
        return True, adjustment_details
        
    except TimeoutException:
        print("✗ Timeout: No adjustments found")
        return False, []
    except Exception as e:
        print(f"✗ Error: {e}")
        return False, []

def click_paynow_button(driver, timeout=10):    
    try:
        print("Looking for PayNow button...")
        
        # Try multiple selectors for PayNow button
        selectors = [
            (By.ID, "qrPhPayment"),
            (By.CSS_SELECTOR, "button#qrPhPayment"),
            (By.XPATH, "//button[@id='qrPhPayment']"),
            (By.XPATH, "//button[contains(text(), 'Pay Now')]"),
        ]
        
        element = None
        for by, selector in selectors:
            try:
                element = WebDriverWait(driver, timeout).until(
                    EC.element_to_be_clickable((by, selector))
                )
                print(f"✓ Found PayNow button with: {selector}")
                break
            except:
                continue
        
        if not element:
            print("✗ PayNow button not found")
            return False
        
        # Scroll into view
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        time.sleep(0.5)
        
        # Click
        driver.execute_script("arguments[0].click();", element)
        print("✓ Clicked PayNow button")
        
        return True
        
    except Exception as e:
        print(f"✗ Error: {e}")
        return False



def click_view_orders(driver, billing_id=None, timeout=10):
    try:
        print("⏳ Looking for View Orders button...")
        
        if billing_id:
            # Target specific View Orders button by billing ID
            view_orders_btn = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR, 
                    f'button.accordion-button[data-bs-target="#collapseOrders-{billing_id}"]'
                ))
            )
            print(f" Found View Orders button for billing ID: {billing_id}")
        else:
            # Click first available View Orders button
            view_orders_btn = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR, 
                    'button.accordion-button.btn-order-details'
                ))
            )
            billing_id = view_orders_btn.get_attribute('data-bs-target').replace('#collapseOrders-', '')
            print(f" Found View Orders button for billing ID: {billing_id}")
        
        # Get order number
        order_no = view_orders_btn.get_attribute('data-orderno')
            
        # Scroll into view
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", view_orders_btn)
        time.sleep(0.5)
        
        # Check if already expanded
        is_expanded = view_orders_btn.get_attribute('aria-expanded')
        if is_expanded == 'true':
            return billing_id
        
        # Click to expand
        try:
            view_orders_btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", view_orders_btn)
        
        time.sleep(1)
        
        # Wait for collapse to expand
        collapse_id = f"collapseOrders-{billing_id}"
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.ID, collapse_id))
        )
        print(" View Orders accordion expanded")
        
        return billing_id
        
    except Exception as e:
        print(f" Error clicking View Orders: {e}")
        print(traceback.format_exc())
        raise
        
def wait_for_collapse_content(driver, billing_id, timeout=10):
   
    try:
        collapse_id = f"collapseOrders-{billing_id}"
        print(f" Waiting for collapse content: {collapse_id}")
        
        collapse_content = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.ID, collapse_id))
        )
        
        print(" Collapse content is visible")
        return collapse_content
        

    except Exception as e:
        print(f" Error waiting for collapse content: {e}")
        raise

def click_bill_out(driver, timeout=15):
    print(" Waiting for and clicking 'Bill Out' button...")

    try:
        bill_out_btn = WebDriverWait(driver, timeout).until(
            lambda d: next(
                (
                    btn for btn in d.find_elements(
                        By.XPATH,
                        "//button[normalize-space(text())='Bill Out' or "
                        "normalize-space(text())='BILL OUT' or "
                        "normalize-space(text())='bill out']"
                    )
                    if btn.is_displayed() and btn.is_enabled()
                ),
                None
            )
        )

        if bill_out_btn:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", bill_out_btn)
            bill_out_btn.click()
            print(" Bill Out' button clicked successfully.")
        else:
            raise TimeoutException(" 'Bill Out' button not visible or clickable.")

    except Exception as e:
        print(f" Unable to click 'Bill Out' button: {e}")
        print(" Debugging info:")
        print(f"   Current URL: {driver.current_url}")
        print(f"   Page title: {driver.title}")
        print("   Possible cause: View Orders panel overlaying or modal delayed.")


def wait_for_payment_modal(driver, timeout=15):
    print(" Waiting for payment modal to appear...")
    
    try:
        # Wait for any modal with 'show' class (Bootstrap 5 shows modals this way)
        modal = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, ".modal.show"))
        )
        
        modal_id = modal.get_attribute('id')
        modal_classes = modal.get_attribute('class')
    
        print(f" Payment modal is visible!")
        print(f"   Modal ID: {modal_id}")
        print(f"   Modal classes: {modal_classes}")
        
        # Verify VAT text is visible (optional check)
        try:
            vat_element = modal.find_element(By.XPATH, ".//span[@class='fw-bold' and contains(text(), 'VAT')]")
            if vat_element.is_displayed():
                print(f" VAT text visible: {vat_element.text}")
        except:
            print(" VAT text not found in modal (might be in parent page)")
        
        return modal, modal_id
        
    except TimeoutException:
        print(" Payment modal did not appear within timeout")
        
        # Debug: Show all modals on page
        print("\n Debug - All modal elements found:")
        all_modals = driver.find_elements(By.CSS_SELECTOR, "[class*='modal']")
        for m in all_modals:
            print(f"   - ID: {m.get_attribute('id')}, Class: {m.get_attribute('class')}, Displayed: {m.is_displayed()}")
        
        return None, None
        
    except Exception as e:
        print(f" Error waiting for modal: {e}")
        return None, None


def close_modal(driver, modal_id=None, timeout=15):
    
    if not modal_id:
        # Find any visible modal
        print("⏳ Looking for any visible modal...")
        try:
            modals = driver.find_elements(By.CSS_SELECTOR, ".modal.show")
            if modals:
                modal_id = modals[0].get_attribute('id')
                print(f" Found modal: {modal_id}")
            else:
                print(" No visible modal found")
                return False
        except:
            print(" Could not find any modal")
            return False
    
    print(f" Attempting to close modal '{modal_id}'...")
    
    # Strategy 1: Find and click close button
    close_button_selectors = [
        (By.CSS_SELECTOR, f"#{modal_id} button[data-bs-dismiss='modal']"),
        (By.CSS_SELECTOR, f"#{modal_id} .btn-close"),
        (By.CSS_SELECTOR, f"#{modal_id} button.btn-close"),
        (By.CSS_SELECTOR, f"#{modal_id} button[aria-label='Close']"),
        (By.XPATH, f"//div[@id='{modal_id}']//button[@data-bs-dismiss='modal']"),
        (By.XPATH, f"//div[@id='{modal_id}']//button[contains(@class, 'btn-close')]"),
    ]
    
    for i, (by, selector) in enumerate(close_button_selectors, 1):
        try:
            print(f"   Trying selector {i}/{len(close_button_selectors)}...")
            close_btn = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((by, selector))
            )
            
            # Scroll into view
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", close_btn)
            time.sleep(0.3)
            
            # Try JavaScript click
            driver.execute_script("arguments[0].click();", close_btn)
            print(f" Close button clicked successfully!")
            
            # Wait for modal to disappear
            time.sleep(1)
            
            # Check if modal is gone
            try:
                modal_elem = driver.find_element(By.ID, modal_id)
                if not modal_elem.is_displayed():
                    print(" Modal closed successfully!")
                    return True
                else:
                    print(" Modal still visible after clicking close button")
            except:
                # Modal element not found = successfully closed
                print(" Modal closed successfully!")
                return True
                
        except TimeoutException:
            continue
        except Exception as e:
            print(f"   Error with selector: {e}")
            continue
    
    # Strategy 2: Press ESC key
    print(" Close button click didn't work, trying ESC key...")
    try:
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
        time.sleep(1)
        
        # Check if modal closed
        try:
            modal_elem = driver.find_element(By.ID, modal_id)
            if not modal_elem.is_displayed():
                print(" Modal closed with ESC key!")
                return True
        except:
            print(" Modal closed with ESC key!")
            return True
            
    except Exception as e:
        print(f" ESC key failed: {e}")
    
    # Strategy 3: Click modal backdrop
    print(" Trying to click modal backdrop...")
    try:
        backdrop = driver.find_element(By.CSS_SELECTOR, ".modal-backdrop")
        backdrop.click()
        time.sleep(1)
        print(" Clicked backdrop")
        return True
    except:
        print(" Could not click backdrop")
    
    # Strategy 4: Force close with JavaScript
    print(" Trying to force close...")
    try:
        driver.execute_script(f"""
            var modal = document.getElementById('{modal_id}');
            if (modal) {{
                var bsModal = bootstrap.Modal.getInstance(modal);
                if (bsModal) {{
                    bsModal.hide();
                }} else {{
                    modal.classList.remove('show');
                    modal.style.display = 'none';
                    document.body.classList.remove('modal-open');
                    var backdrop = document.querySelector('.modal-backdrop');
                    if (backdrop) backdrop.remove();
                }}
            }}
        """)
        time.sleep(1)
        print(" Modal force")
        return True
    except Exception as e:
        print(f"  Force close failed: {e}")
    
    # If all strategies failed
    print(" All close strategies failed")
    driver.save_screenshot(f"modal_close_failed_{int(time.time())}.png")
    
    # Final debug info
    print("\n Final Debug Info:")
    try:
        modal_elem = driver.find_element(By.ID, modal_id)
        print(f"   Modal HTML (first 1000 chars):")
        print(f"   {modal_elem.get_attribute('outerHTML')[:1000]}")
    except:
        print("   Could not get modal HTML")
    
    return False


# def click_pay_now(driver, timeout=15):
#     print("💰 Waiting for and clicking 'Pay Now' button...")

#     try:
#         pay_now_btn = WebDriverWait(driver, timeout).until(
#             lambda d: next(
#                 (
#                     btn for btn in d.find_elements(
#                         By.XPATH,
#                         "//button[normalize-space(text())='Pay Now' or "
#                         "normalize-space(text())='PAY NOW' or "
#                         "normalize-space(text())='pay now']"
#                     )
#                     if btn.is_displayed() and btn.is_enabled()
#                 ),
#                 None
#             )
#         )

#         if pay_now_btn:
#             driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", pay_now_btn)
#             pay_now_btn.click()
#             print(" 'Pay Now' button clicked successfully.")
#         else:
#             raise TimeoutException("❌ 'Pay Now' button not visible or clickable.")

#     except Exception as e:
#         print(f" Unable to click 'Pay Now' button: {e}")
#         print(" Debugging info:")
#         print(f"   Current URL: {driver.current_url}")
#         print(f"   Page title: {driver.title}")
#         print("   Possible cause: Payment modal not rendered or overlay blocking interaction.")
#         raise TimeoutException("Pay Now button not found or clickable.") from e

def click_pay_now(driver, timeout=15):
    """
    Click the Pay Now button after handling any blocking modals
    """
    print("💰 Waiting for and clicking 'Pay Now' button...")

    try:
        # Step 1: Handle SweetAlert modal if present
        try:
            print("🔍 Checking for blocking SweetAlert modals...")
            swal_modal = WebDriverWait(driver, 3).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
            )
            
            if swal_modal.is_displayed():
                print("⚠️ SweetAlert modal detected - attempting to dismiss...")
                
                # Try to get modal text for debugging
                try:
                    title = driver.find_element(By.ID, "swal2-title").text
                    content = driver.find_element(By.ID, "swal2-html-container").text
                    print(f"📋 Modal Title: {title}")
                    print(f"📋 Modal Content: {content}")
                except:
                    pass
                
                # Find and click the confirm button
                confirm_selectors = [
                    (By.CSS_SELECTOR, ".swal2-confirm"),
                    (By.CSS_SELECTOR, "button.swal2-confirm"),
                    (By.XPATH, "//button[contains(@class, 'swal2-confirm')]"),
                    (By.XPATH, "//button[contains(text(), 'OK')]"),
                    (By.XPATH, "//button[contains(text(), 'Yes')]"),
                    (By.XPATH, "//button[contains(text(), 'Confirm')]"),
                ]
                
                clicked = False
                for selector in confirm_selectors:
                    try:
                        confirm_btn = driver.find_element(*selector)
                        if confirm_btn.is_displayed():
                            driver.execute_script("arguments[0].click();", confirm_btn)
                            print("✅ Dismissed SweetAlert modal")
                            time.sleep(2)  # Wait for modal to close
                            clicked = True
                            break
                    except:
                        continue
                
                if not clicked:
                    print("⚠️ Could not find confirm button, trying ESC key...")
                    from selenium.webdriver.common.keys import Keys
                    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                    time.sleep(1)
                
                # Wait for modal to disappear
                WebDriverWait(driver, 5).until(
                    EC.invisibility_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
                )
                print("✅ Modal dismissed successfully")
                
        except TimeoutException:
            # No modal found, continue normally
            print("✅ No blocking modal detected")
        except Exception as modal_error:
            print(f"⚠️ Modal handling error (continuing anyway): {modal_error}")
        
        # Step 2: Now find and click Pay Now button
        print("🔍 Looking for Pay Now button...")
        
        pay_now_btn = WebDriverWait(driver, timeout).until(
            lambda d: next(
                (
                    btn for btn in d.find_elements(
                        By.XPATH,
                        "//button[normalize-space(text())='Pay Now' or "
                        "normalize-space(text())='PAY NOW' or "
                        "normalize-space(text())='pay now']"
                    )
                    if btn.is_displayed() and btn.is_enabled()
                ),
                None
            )
        )

        if pay_now_btn:
            # Scroll into view
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", pay_now_btn)
            time.sleep(0.5)
            
            # Try multiple click methods for reliability
            try:
                # Method 1: JavaScript click (most reliable)
                driver.execute_script("arguments[0].click();", pay_now_btn)
                print("✅ 'Pay Now' button clicked successfully")
            except Exception as e1:
                print(f"⚠️ click failed, trying regular click: {e1}")
                try:
                    # Method 2: Regular click
                    pay_now_btn.click()
                    print("✅ 'Pay Now' button clicked successfully (regular)")
                except Exception as e2:
                    print(f"⚠️ Regular click failed, trying ActionChains: {e2}")
                    # Method 3: ActionChains
                    from selenium.webdriver.common.action_chains import ActionChains
                    actions = ActionChains(driver)
                    actions.move_to_element(pay_now_btn).click().perform()
                    print("✅ 'Pay Now' button clicked successfully (ActionChains)")
            
            time.sleep(1)
            return True
        else:
            raise TimeoutException("❌ 'Pay Now' button not visible or clickable.")

    except Exception as e:
        print(f"❌ Unable to click 'Pay Now' button: {e}")
        print("🔍 Debugging info:")
        print(f"   Current URL: {driver.current_url}")
        print(f"   Page title: {driver.title}")
        
        # Check for any visible modals
        try:
            modals = driver.find_elements(By.CSS_SELECTOR, ".modal.show, .swal2-show, [role='dialog']")
            if modals:
                print(f"   ⚠️ Found {len(modals)} visible modal(s):")
                for i, modal in enumerate(modals):
                    if modal.is_displayed():
                        print(f"      Modal {i+1}: {modal.get_attribute('class')}")
        except:
            pass
        
        # Check Pay Now button state
        try:
            pay_btns = driver.find_elements(By.XPATH, "//button[contains(text(), 'Pay Now')]")
            print(f"   Pay Now buttons found: {len(pay_btns)}")
            if pay_btns:
                btn = pay_btns[0]
                print(f"   Button displayed: {btn.is_displayed()}")
                print(f"   Button enabled: {btn.is_enabled()}")
        except:
            pass
        
        print("   Possible causes:")
        print("   - Payment modal not rendered")
        print("   - Overlay/modal blocking interaction")
        print("   - Button not yet enabled")
        
        raise TimeoutException("Pay Now button not found or clickable.") from e

















def select_senior_discount(driver, timeout=60):
    try:
        print(" Waiting for discount dropdown...")
        discount_dropdown = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "discount"))
        )
        
        # Use Select class for dropdown
        select = Select(discount_dropdown)
        select.select_by_value("senior discount")
        
        print(" Selected 'Senior discount'")
        time.sleep(0.5)
        return True
        
    except Exception as e:
        print(f" Error selecting senior discount: {e}")

        return False

def select_discount_option(driver, discount_type, screenshots_folder, timeout=60, take_screenshot=True):
    # Map discount types to their values
    discount_map = {
        "senior": "senior discount",
        "senior discount": "senior discount",
        "pwd": "pwd discount",
        "pwd discount": "pwd discount",
        "others": "others",
        "add discount": "others",
        "charges": "charges",
        "add charges": "charges"
    }
    
    try:
        print(f"⏳ Waiting for discount dropdown...")
        
        # Wait for dropdown to be present
        discount_dropdown = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "discount"))
        )
        
        # Get the value to select
        discount_value = discount_map.get(discount_type.lower())
        
        if not discount_value:
            print(f"❌ Invalid discount type: {discount_type}")
            print(f"   Valid options: {list(discount_map.keys())}")
            return False
        
        # Take screenshot before selection
        if take_screenshot:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_name = f"discount_before_{timestamp}.png"
            screenshot_path = os.path.join(screenshots_folder, screenshot_name)
            driver.save_screenshot(screenshot_path)
            print(f"📸 Screenshot before: {screenshot_path}")
        
        # Use Select class for dropdown
        select = Select(discount_dropdown)
        select.select_by_value(discount_value)
        
        # Get the selected option text
        selected_option = select.first_selected_option
        selected_text = selected_option.text
        
        print(f"✅ Selected: {selected_text}")
        
        time.sleep(1)
        
        # Take screenshot after selection
        if take_screenshot:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_name = f"discount_{discount_type.replace(' ', '_')}_selected_{timestamp}.png"
            screenshot_path = os.path.join(screenshots_folder, screenshot_name)
            driver.save_screenshot(screenshot_path)
            print(f"📸 Screenshot after: {screenshot_path}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error selecting discount: {e}")
        if take_screenshot:
            screenshot_name = f"discount_error_{int(time.time())}.png"
            screenshot_path = os.path.join(screenshots_folder, screenshot_name)
            driver.save_screenshot(screenshot_path)
        return False    


def select_all_discounts(driver, screenshots_folder, delay_between=3):

    discount_options = ["senior", "pwd", "others", "charges"]
    
    print("\n" + "="*60)
    print("📸 Cycling through all discount options...")
    print("="*60 + "\n")
    
    for discount in discount_options:
        print(f"\n--- Selecting: {discount.upper()} ---")
        
        success = select_discount_option(driver, discount, screenshots_folder)
        
        if success:
            print(f"✅ {discount.upper()} selected successfully")
        else:
            print(f"❌ Failed to select {discount.upper()}")
        
        # Wait before next selection
        if discount != discount_options[-1]:  # Don't wait after last option
            print(f"⏳ Waiting {delay_between} seconds before next selection...")
            time.sleep(delay_between)
    
    print("\n" + "="*60)
    print("✅ All discount options captured!")
    print("="*60 + "\n")





def input_discount_amount(driver, amount="20.00", timeout=15):
    try:
        print(" Waiting for amount input field...")
        amount_input = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, "discountAmount"))
        )
        
        # Clear and input amount
        amount_input.clear()
        amount_input.send_keys(amount)
        
        print(f" Entered discount amount: {amount}")
        time.sleep(0.5)
        return True
        
    except Exception as e:
        print(f" Error entering discount amount: {e}")
        return False
    
def update_bill(driver, timeout=15):
    try:
        print(" Waiting for 'Update Bill' button...")
        update_btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, "updateBill"))
        )
        
        # Get current total before update
        try:
            current_total = driver.find_element(By.CSS_SELECTOR, "[data-billing-total]").text
            print(f"   Current total: ₱{current_total}")
        except:
            current_total = None
        
        # Click update button
        driver.execute_script("arguments[0].click();", update_btn)
        print(" Clicked 'Update Bill' button")
        
        # Wait for bill to update (wait for total to change or for a success message)
        time.sleep(2)
        
        # Verify bill updated
        try:
            new_total = driver.find_element(By.CSS_SELECTOR, "[data-billing-total]").text
            print(f"   New total: ₱{new_total}")
            
            if current_total and new_total != current_total:
                print(" Bill updated successfully (total changed)")
            else:
                print(" Bill update completed")
        except:
            print(" Bill update completed")
        
        return True
        
    except Exception as e:
        print(f" Error updating bill: {e}")
        return False
    
def reprint_billout(driver, timeout=15):
    try:
        print(" Waiting for 'Reprint BillOut' button...")
        reprint_btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, "reprintBillOut"))
        )
        
        # Click reprint button
        driver.execute_script("arguments[0].click();", reprint_btn)
        print(" Clicked 'Reprint BillOut' button")
     
        
    except Exception as e:
        print(f" Error in reprint billout: {e}")
        return False, None

def select_qr_payment(driver, timeout=15):
    try:
        # Select QR PH radio button
        print(" Waiting for QR PH radio button...")
        qr_radio = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, "qrph-radio"))
        )
        
        # Click radio button
        driver.execute_script("arguments[0].click();", qr_radio)
        print(" Selected 'QR PH' payment option")
        time.sleep(0.5)
        
    except Exception as e:
        print(f" Error in QR payment selection: {e}")
        return False
    
def select_other_payment_method(driver, timeout=15):
    try:
        # Select Other payment radio button
        print(" Waiting for Other Payment Mode button...")
        other_radio = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, "other-radio"))
        )
        
        # Click radio button
        driver.execute_script("arguments[0].click();", other_radio)
        print(" Selected payment option")
        time.sleep(0.5)
        
    except Exception as e:
        print(f" Error in payment selection: {e}")
        return False
    






    
def pay_now(driver):
    pay_now_btn = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'qrPhPayment')))
    pay_now_btn.click()
    print("✓ Pay Now button clicked")
    WebDriverWait(driver, 10).until(lambda d: d.execute_script('return document.readyState') == 'complete')

def confirm_payment(driver, timeout=10):
    print(" Confirming payment...")
    proceed_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Yes, proceed')]"))
)
    proceed_button.click()
    print(" Payment confirmed.")

def wait_for_qr_and_screenshot(driver, timeout=30):
    try:
        print(" Waiting for QR code to appear...")
        
        # Try multiple selectors for QR code
        qr_selectors = [
            (By.CSS_SELECTOR, "img[alt*='QR']"),
            (By.CSS_SELECTOR, "img[src*='qr']"),
            (By.CSS_SELECTOR, ".qr-code"),
            (By.ID, "qrCode"),
            (By.CSS_SELECTOR, "canvas"),  # QR codes often rendered in canvas
            (By.XPATH, "//div[contains(@class, 'qr')]"),
        ]
        
        qr_element = None
        for by, selector in qr_selectors:
            try:
                qr_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((by, selector))
                )
                print(f" QR code found using: {selector}")
                break
            except TimeoutException:
                continue
        
        if not qr_element:
            print(" QR code not found with specific selector, trying generic modal wait...")
            time.sleep(3)  # Give it time to render
        
        # Take screenshot
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        screenshot_name = f"qr_code_{timestamp}.png"
        driver.save_screenshot(screenshot_name)
        print(f"📸 Screenshot saved: {screenshot_name}")
        
        return True, screenshot_name
        
    except Exception as e:
        print(f" Error waiting for QR code: {e}")
        # Still take a screenshot for debugging
        screenshot_name = f"qr_error_{int(time.time())}.png"
        driver.save_screenshot(screenshot_name)
        return False, screenshot_name
def wait_for_payment_modal(driver, timeout=30):
    try:
        print("⏳ Waiting for modal to appear...")
        
        # Try multiple selectors for modal
        modal_selectors = [
            (By.CSS_SELECTOR, ".modal.show"),
            (By.CSS_SELECTOR, ".modal.fade.show"),
            (By.CSS_SELECTOR, "div[role='dialog']"),
            (By.CSS_SELECTOR, ".modal-content"),
            (By.CSS_SELECTOR, ".modal-dialog"),
            (By.ID, "receivePaymentModal"),
            (By.CSS_SELECTOR, "[class*='modal'][style*='display: block']"),
            (By.XPATH, "//div[contains(@class, 'modal') and contains(@style, 'display')]"),
        ]
        
        modal_element = None
        for by, selector in modal_selectors:
            try:
                modal_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((by, selector))
                )
                print(f"✅ Modal found using: {selector}")
                break
            except TimeoutException:
                continue
        
        if not modal_element:
            print("⚠️ Modal not found with specific selector, trying generic wait...")
            time.sleep(3)  # Give it time to render
        
        # Additional wait for modal to be fully visible
        try:
            WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, ".modal-body"))
            )
            print("✅ Modal body is visible")
        except:
            print("⚠️ Modal body visibility check failed, continuing anyway...")
        
        # Take screenshot
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        screenshot_name = f"modal_{timestamp}.png"
        driver.save_screenshot(screenshot_name)
        print(f"📸 Screenshot saved: {screenshot_name}")
        
        return True, screenshot_name
        
    except Exception as e:
        print(f"❌ Error waiting for modal: {e}")
        # Still take a screenshot for debugging
        screenshot_name = f"modal_error_{int(time.time())}.png"
        driver.save_screenshot(screenshot_name)
        return False, screenshot_name    
    
def check_payment_and_confirm(driver, timeout=30):
    try:
        # Click Check Payment button
        print("⏳ Looking for 'Check Payment' button...")
        
        check_payment_selectors = [
            (By.ID, "checkPayment"),
            (By.XPATH, "//button[contains(text(), 'Check Payment')]"),
            (By.XPATH, "//button[contains(text(), 'Verify Payment')]"),
            (By.CSS_SELECTOR, "button.btn-check-payment"),
        ]
        
        check_btn = None
        for by, selector in check_payment_selectors:
            try:
                check_btn = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((by, selector))
                )
                print(f" Found 'Check Payment' button")
                break
            except TimeoutException:
                continue
        
        if not check_btn:
            raise TimeoutException("Check Payment button not found")
        
        driver.execute_script("arguments[0].click();", check_btn)
        print(" Clicked 'Check Payment' button")
        
        # Wait for success message/modal
        print(" Waiting for payment confirmation message...")
        time.sleep(3)
        
        # Look for OK/Close button in success message
        print(" Looking for 'OK' button...")
        
        ok_selectors = [
            (By.XPATH, "//button[contains(text(), 'OK')]"),
            (By.XPATH, "//button[contains(text(), 'Close')]"),
            (By.XPATH, "//button[contains(text(), 'Done')]"),
            (By.CSS_SELECTOR, "button.swal2-confirm"),
            (By.CSS_SELECTOR, "button.btn-primary"),
            (By.CSS_SELECTOR, "button.btn-success"),
        ]
        
        ok_btn = None
        for by, selector in ok_selectors:
            try:
                ok_btn = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((by, selector))
                )
                print(f" Found 'OK' button")
                break
            except TimeoutException:
                continue
        
        if ok_btn:
            driver.execute_script("arguments[0].click();", ok_btn)
            print(" Clicked 'OK' button")
            time.sleep(1)
        else:
            print(" No 'OK' button found (modal might have auto-closed)")
        
        return True
        
    except Exception as e:
        print(f" Error in check payment: {e}")
        driver.save_screenshot(f"check_payment_error_{int(time.time())}.png")
        return False
    
    
def close_modal_paynow(driver):
    
    methods = [
        ("XPath class contains", lambda: driver.find_element(By.XPATH, "//button[contains(@class, 'btn-close')]").click()),
        ("CSS class selector", lambda: driver.find_element(By.CSS_SELECTOR, "button.btn-close").click()),
        ("XPath data-bs-dismiss", lambda: driver.find_element(By.XPATH, "//button[contains(@data-bs-dismiss, 'modal')]").click()),
        ("JavaScript click XPath", lambda: driver.execute_script("arguments[0].click();", driver.find_element(By.XPATH, "//button[contains(@class, 'btn-close')]"))),
        ("JavaScript click CSS", lambda: driver.execute_script("arguments[0].click();", driver.find_element(By.CSS_SELECTOR, "button.btn-close"))),
        ("ESC key", lambda: driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)),
        ("Bootstrap modal.hide()", lambda: driver.execute_script("bootstrap.Modal.getInstance(document.querySelector('.modal.show')).hide();")),
        ("jQuery hide", lambda: driver.execute_script("$('.modal').modal('hide');")),
    ]
    
    for description, method in methods:
        try:
            method()
            print(f" Modal closed using: {description}")
            time.sleep(1)
            return True
        except Exception as e:
            print(f" {description} failed: {e.__class__.__name__}")
            continue
    
    print(" All methods failed")
    driver.save_screenshot("modal_close_failed.png")
    return False

def collapse_open_orders(driver):
    try:
        open_panels = driver.find_elements(By.CSS_SELECTOR, "div.collapse.show")
        if open_panels:
            for panel in open_panels:
                driver.execute_script("arguments[0].classList.remove('show');", panel)
            print(" Collapsed all expanded 'View Orders' sections.")
            time.sleep(1)
        else:
            print(" No expanded 'View Orders' sections found.")
    except Exception as e:
        print(f" Unable to collapse open orders: {e}")

def wait_for_overlay_to_close(driver, timeout=5):
    try:
        WebDriverWait(driver, timeout).until_not(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".modal-backdrop.show"))
        )
    except TimeoutException:
        pass  # It's fine if no overlay is found



       
# Discount
def get_random_discount_from_excel(file_path):

    df = pd.read_excel(file_path)

    if df.empty:
        raise ValueError("Excel discount file is empty.")

    chosen_row = df.sample(n=1).iloc[0]  # Randomly pick one
    discount_type = str(chosen_row["DiscountType"]).strip()
    discount_amount = str(chosen_row["DiscountAmount"]).strip()

    return discount_type, discount_amount

def apply_random_discount(driver, wait, file_path, log_file_path=None):
    try:
        #  Step 1: Pick random discount from Excel
        discount_type, discount_amount = get_random_discount_from_excel(file_path)

        #  Step 2: Select discount from dropdown
        select_dropdown(driver, wait, By.ID, "discount", discount_type)
        log_action(f"Randomly selected discount: {discount_type}", log_file_path)

        # Step 3: Enter discount amount
        discount_amount_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "discountAmount"))
        )
        discount_amount_field.clear()
        discount_amount_field.send_keys(discount_amount)
        log_action(f"Entered discount amount: {discount_amount}", log_file_path)

        # Wait for UI update confirmation (optional)
        WebDriverWait(driver, 3).until(
            EC.text_to_be_present_in_element_value((By.ID, "discountAmount"), discount_amount)
        )

    except Exception:
        log_error(f"Failed to apply random discount: {traceback.format_exc()}", log_file_path)

def click_table_status(driver, log_file_path, screenshots_folder, 
                       element_xpath=None, element_id=None, screenshot_name="status"):
    try:
        # Find element
        if element_id:
            elements = driver.find_elements(By.ID, element_id)
        elif element_xpath:
            elements = driver.find_elements(By.XPATH, element_xpath)
        else:
            raise ValueError("Either element_id or element_xpath must be provided")

        if elements and elements[0].is_displayed():
            el = elements[0]
            driver.execute_script("arguments[0].click()", el)
            log_action(f"Click {screenshot_name}", log_file_path=log_file_path)

            # Wait for page load and tables presence
            WebDriverWait(driver, 10).until(lambda d: d.execute_script('return document.readyState') == 'complete')
            WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.ID, 'restaurantTables')))
            time.sleep(3)
            driver.save_screenshot(os.path.join(screenshots_folder, f"{screenshot_name}.png"))

        else:
            # Check empty state
            empty_elements = driver.find_elements(By.XPATH, '//*[@id="tables-empty-state"]/div/strong')
            if empty_elements and empty_elements[0].is_displayed():
                log_action(f"No tables available: {empty_elements[0].text}", log_file_path=log_file_path)
            else:
                log_action(f"QA-SCDWEB-PB42-S2: {screenshot_name} not available", log_file_path=log_file_path)

    except (TimeoutException, NoSuchElementException) as e:
        log_action(f"QA-SCDWEB-PB42-S2: Failed to click {screenshot_name} - {str(e)}", log_file_path=log_file_path)

    
def wait_for_tile_color_by_number(driver, table_num, target_color, timeout=10):

    xpath = (
        f"//div[contains(@class, 'table-tile')]"
        f"[.//div[contains(@class, 'table-tile__number') and normalize-space()='{table_num}']]"
        f"[@data-tile-color='{target_color}']"
    )
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, xpath))
    )

def gray_table_number(driver, timeout=10, log_file_path=None):
 
    try:
        # Wait for tables to be present
        WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#restaurantTables .table-card"))
        )
        
        tables = driver.find_elements(By.CSS_SELECTOR, "#restaurantTables .table-card")
        for table in tables:
            color = table.get_attribute("data-color")  # assuming color stored in attribute
            number = table.get_attribute("data-number")  # assuming table number stored in attribute
            if color.lower() == "gray":
                if log_file_path:
                    log_action(f"Found gray table: {number}", log_file_path=log_file_path)
                return int(number)
        # If no gray table found
        if log_file_path:
            log_action("No gray table found", log_file_path=log_file_path)
        return None

    except TimeoutException:
        if log_file_path:
            log_action("Tables not loaded in time", log_file_path=log_file_path)
        return None
    
def clear_folder(folder_path=None, logs_folder=None, screenshots_folder=None):
    # Pick the first valid argument
    target = None
    if folder_path:
        target = folder_path
    elif logs_folder:
        target = logs_folder
    elif screenshots_folder:
        target = screenshots_folder

    if not target:
        raise ValueError("No folder path provided.")

    try:
        if os.path.exists(target):
            shutil.rmtree(target)  # delete entire folder
        os.makedirs(target, exist_ok=True)  # recreate empty folder
        print(f"[INFO] Cleared folder: {target}")
        return True
    except Exception as e:
        print(f"[ERROR] Failed to clear folder {target}. Reason: {e}")
        return False

def page_loaded(locator):
    def _predicate(driver):
        ready = driver.execute_script("return document.readyState") == 'complete'
        if not ready:
            return False
        try:
            elem = driver.find_element(*locator)
            return elem.is_displayed()  # or other criteria
        except:
            return False
    return _predicate

# Printer Setup
def printer():
    excel_path = r"d:\Bastion\SQA_QualityEnterprise\SoftwareTesting\Projects\Common\Automation\OPI\Files\Testdata\printer.xlsx"
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    df = pd.read_excel(excel_path)
    row = df.sample(n=1).iloc[0]
    
    return {
        "Kitchen Printer": row.get("Kitchen Printer", ""),
        "Drinks Printer": row.get("Drinks Printer", ""),  # will be "" if not in Excel
        "Bar Printer": row.get("Bar Printer", "")
    }
def set_preferred_time_booking(driver, time_input_selector, preferred_time, timeout=10):

    try:
        from datetime import datetime

        print(f" Setting time: {preferred_time}")

        #  Convert 12-hour format to 24-hour format
        try:
            time_obj = datetime.strptime(preferred_time.strip(), "%I:%M %p")
            time_24hr = time_obj.strftime("%H:%M")
            print(f"   Converted to 24-hour format: {time_24hr}")
        except ValueError:
            print(f" Invalid time format: {preferred_time}. Expected format: 'HH:MM AM/PM'")
            return False

        #  Wait for input to be clickable (not just present)
        time_input = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, time_input_selector))
        )
        print(" Time input located")

        #  Scroll into view and ensure visible
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", time_input)
        driver.execute_script("arguments[0].focus();", time_input)
        time.sleep(0.3)

        if not time_input.is_enabled():
            print(" Time input field is disabled.")
            return False

        #  Try JS method (best for <input type='time'>)
        driver.execute_script(f"arguments[0].value = '{time_24hr}';", time_input)
        driver.execute_script("""
            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('blur', { bubbles: true }));
        """, time_input)

        time.sleep(0.5)
        current_value = time_input.get_attribute("value")

        if current_value == time_24hr:
            print(f" Time successfully set via JS: {current_value} ({preferred_time})")
            return True

        #  JS failed → fallback to send_keys
        print(f" JS method mismatch. Retrying with send_keys...")

        time_input.clear()
        time_input.click()
        time.sleep(0.3)
        time_input.send_keys(time_24hr)
        time_input.send_keys(Keys.TAB)
        time.sleep(0.5)

        current_value = time_input.get_attribute("value")
        if current_value == time_24hr:
            print(f" Time successfully set via send_keys: {current_value}")
            return True
        else:
            print(f" Failed to set time. Current value: '{current_value}' (expected {time_24hr})")
            return False

    except TimeoutException:
        print(f" Timeout: No element found using selector '{time_input_selector}'")
        return False

    except Exception as e:
        print(f" Unexpected error while setting time: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def Restaurant_Home_Dashboard(driver,log_file_path):
    # Click Restaurant Center
        Restaurant_Center = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/header/div/nav/ul/li[4]/a/span[2]"))
        )
        driver.execute_script("arguments[0].click();", Restaurant_Center)
        log_action("Clicked Restaurant Center", log_file_path=log_file_path)
        WebDriverWait(driver, 20).until(lambda d: d.execute_script('return document.readyState') == 'complete')

        # Wait for basic elements to appear
        WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="nav-home-restaurant"]/div[2]/div[1]/div[1]/div')))
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'ongoingOrders')))
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, 'appointmentForToday')))

def resto_dropdown(driver, wait, select_id, value_to_select, log_file_path=None, screenshots_folder=None):
 
    try:
        # Wait until the select element is present
        element = wait.until(EC.presence_of_element_located((By.ID, select_id)))
        
        # Scroll into view in case it's off-screen
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, select_id)))

        # Create a Select instance
        select = Select(element)

        # Convert the value to string to avoid float errors
        value_to_select = str(value_to_select).strip()

        # Select by visible text (like “Food”, “Drinks”)
        select.select_by_visible_text(value_to_select)

        print(f"Selected '{value_to_select}' from dropdown '{select_id}'")

        # Optional logging & screenshot
        if log_file_path:
            log_action(f"Selected '{value_to_select}' in dropdown '{select_id}'", log_file_path=log_file_path)
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, f"Dropdown_{select_id}_{value_to_select}.png"))

        return True

    except Exception as e:
        print(f" Failed to select '{value_to_select}' from dropdown '{select_id}': {e}")
        return False
    
def get_visible_menu_categories(driver, log_file_path=None, timeout=10):
    try:
        # Wait for the <ul> container
        menu_container = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "subcategory-tabs"))
        )

        menu_items = menu_container.find_elements(By.CSS_SELECTOR, "li.nav-item")
        visible_categories = []

        for li in menu_items:
            try:
                name = li.find_element(By.TAG_NAME, "a").text.strip()
                style = (li.get_attribute("style") or "").lower()
                is_visible = "display: none" not in style
                state = "visible" if is_visible else "hidden"

                log_action(f" Menu Category: '{name}' ({state})", log_file_path)
                if is_visible:
                    visible_categories.append(name)
            except Exception as inner_e:
                log_error(f" Error reading menu item: {inner_e}", log_file_path)

        log_action(f" Visible menu categories detected: {visible_categories}", log_file_path)
        return visible_categories

    except TimeoutException:
        log_error(" Timeout: subcategory-tabs not found on the page.", log_file_path)
        return []
    
def wait_and_click_ok(driver, timeout=15):
    try:
        # Wait for a modal or dialog to appear first
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[contains(@class,'modal') or contains(@role,'dialog')]")
            )
        )

        # Then wait for OK/Ok button
        ok_button = WebDriverWait(
            driver, timeout, poll_frequency=0.5,
            ignored_exceptions=[StaleElementReferenceException]
        ).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[normalize-space()='OK' or normalize-space()='Ok']")
            )
        )

        driver.execute_script("arguments[0].scrollIntoView(true);", ok_button)
        time.sleep(0.3)
        ok_button.click()
      

    except TimeoutException:
        print(" Timeout: OK button not found within time limit.")
    except ElementClickInterceptedException:
        print(" OK button found but blocked by another overlay, retrying...")
        driver.execute_script("arguments[0].click();", ok_button)
    except Exception as e:
        print(f" Unexpected error clicking OK button: {e}")

def is_item_in_dessert_flexible(driver, product_name, timeout=10,
                                container_selector="div.subcategory-products",
                                item_selector="div.card-button",
                                inactive_class="inactive"):
    # Wait for the container to appear
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, container_selector))
        )
    except TimeoutException:
        return False

    # Prepare lowercase version for matching
    target = product_name.strip().lower()

    # Find all candidate items which are not marked inactive
    # Using CSS to filter out inactive class
    css = f"{container_selector} {item_selector}:not(.{inactive_class})"
    try:
        elems = driver.find_elements(By.CSS_SELECTOR, css)
    except StaleElementReferenceException:
        elems = driver.find_elements(By.CSS_SELECTOR, css)

    for el in elems:
        # Get the data-product-name attribute (if exists)
        name_attr = el.get_attribute("data-product-name")
        if name_attr:
            name_lower = name_attr.strip().lower()
            # Check if the target is a substring (so partial)
            if target in name_lower:
                return True

        # If no data attribute, fallback to visible text in the element or its children
        try:
            visible_text = el.text  # this is all the text inside the element
        except StaleElementReferenceException:
            continue

        if visible_text and target in visible_text.strip().lower():
            return True

    return False

# -----Verify all billing details are present and visible----- #
def verify_billing_details(driver, log_file_path):

    try:
        # Wait for container
        billing_container = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.billing-details"))
        )
        
        # Check if container is displayed
        assert billing_container.is_displayed(), "Billing details container not visible"
        
        # Verify Order Number
        order_no = WebDriverWait(billing_container, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.selected-orderno span"))
        )
        assert order_no.text, "Order number is empty"
        print(f" Order #: {order_no.text}")
        
        # Get all divs and verify date/time
        all_divs = billing_container.find_elements(By.TAG_NAME, "div")
        
        # Find date (contains "Date:")
        date_found = any("Date:" in div.text for div in all_divs)
        assert date_found, "Date not found"
        print(" Date field present")
        
        # Find time (contains "Time:")
        time_found = any("Time:" in div.text for div in all_divs)
        assert time_found, "Time not found"
        print(" Time field present")
        
        log_action("QA-SCDWEB-PB45-S4: Display table card details - PASSED", log_file_path=log_file_path)
        return True
        
    except Exception as e:
        print(f" Billing details verification failed: {str(e)}")
        return False

def verify_legend_tooltip_text(driver, color, expected_text):
    
    tile = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, f"//span[@class='legend-tile' and @data-tile-color='{color}']"))
    )
    
    # Scroll into view
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tile)
    time.sleep(0.5)
    
    print(f"\n=== Debugging {color} tooltip ===")
    
    # Check if Bootstrap is available
    bootstrap_available = driver.execute_script("return typeof bootstrap !== 'undefined';")
    print(f"Bootstrap available: {bootstrap_available}")
    
    if bootstrap_available:
        # Check Bootstrap version
        bootstrap_version = driver.execute_script("return bootstrap.Tooltip ? 'v5+' : 'unknown';")
        print(f"Bootstrap version: {bootstrap_version}")
    
    # Try to initialize and show tooltip
    result = driver.execute_script("""
        var element = arguments[0];
        
        try {
            // Check if bootstrap is available
            if (typeof bootstrap === 'undefined') {
                return {success: false, error: 'Bootstrap not available'};
            }
            
            // Initialize tooltip
            if (!bootstrap.Tooltip.getInstance(element)) {
                new bootstrap.Tooltip(element);
            }
            
            // Show tooltip
            var tooltipInstance = bootstrap.Tooltip.getInstance(element);
            tooltipInstance.show();
            
            return {success: true, error: null};
        } catch (e) {
            return {success: false, error: e.toString()};
        }
    """, tile)
    
    print(f"Tooltip show result: {result}")
    
    time.sleep(2)
    
    # Look for ANY element with tooltip in class or id
    all_tooltips = driver.find_elements(By.XPATH, "//*[contains(@class, 'tooltip') or contains(@id, 'tooltip')]")
    print(f"Found {len(all_tooltips)} elements with 'tooltip' in class/id")
    
    for i, tt in enumerate(all_tooltips):
        print(f"  Element {i}: tag={tt.tag_name}, class='{tt.get_attribute('class')}', visible={tt.is_displayed()}, text='{tt.text}'")
    
    # Take screenshot
    driver.save_screenshot(f"tooltip_debug_{color}.png")
    
    # Try to get text from aria-label as fallback
    aria_text = tile.get_attribute("aria-label")
    print(f"aria-label text: {aria_text}")
    
    return False

def remove_adjustment(driver, seq_nbr, note):
    selector = f"button.deleteInvoiceAdjustment[data-adjustment-seq-nbr='{seq_nbr}'][data-adjustment-note='{note}']"
    
    remove_btn = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
    )
    remove_btn.click()
    print(f" Removed adjustment: {note} (seq: {seq_nbr})")

def close_paynow_modal(driver, timeout=10):
    print("Closing PayNow modal with ESC key...")
    try:
        # Press ESC
        ActionChains(driver).send_keys(Keys.ESCAPE).perform()
        
        # Wait for modal to close
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".modal.show"))
        )
        
        print(" Modal closed properly")
        return True
    except Exception as e:
        print(f" Error: {e}")
        return False
    

def close_reprint_billout_modal(driver, timeout=10):
    wait = WebDriverWait(driver, timeout)
    
    try:
        close_btn = wait.until(EC.element_to_be_clickable((
            By.CSS_SELECTOR,
            ".modal-content.billout-modal .modal-header .btn-close"
        )))
        close_btn.click()
    
        # Wait until the modal-content disappears
        wait.until(EC.invisibility_of_element_located((
            By.CSS_SELECTOR,
            ".modal-content.billout-modal"
        )))
        return True
    except Exception as e:
        print("Could not click close button:", e)


    try:
        body = driver.find_element(By.TAG_NAME, "body")
        body.send_keys(Keys.ESCAPE)
        # Wait for the modal to be gone
        wait.until(EC.invisibility_of_element_located((
            By.CSS_SELECTOR,
            ".modal-content.billout-modal"
        )))
        return True
    except Exception as e:
        print("Failed:", e)

    try:
        modal_elem = driver.find_element(By.CSS_SELECTOR, ".modal-content.billout-modal")
        driver.execute_script("arguments[0].style.display = 'none';", modal_elem)
        return True
    except Exception as e:
        print(" Failed:", e)

    return False

def close_qr_modal(driver, timeout=10):
    try:
        print("Closing QR Code modal...")
        
        # Click close button
        close_btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".modal.show .btn-close"))
        )
        driver.execute_script("arguments[0].click();", close_btn)
        
        # Wait for modal to close
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".modal.show"))
        )
        
        print(" QR Code modal closed")
        return True
        
    except Exception as e:
        print(f" Error: {e}")
        return False
    
def expand_adjustments(driver, log_file_path, screenshots_folder, wait_time=15):
  
    try:
        log_action("Attempting to click View Adjustments", log_file_path=log_file_path)
        
        # Wait for page ready
        WebDriverWait(driver, wait_time).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        time.sleep(1)
        
        driver.save_screenshot(os.path.join(screenshots_folder, "BeforeClickAdjustments.png"))
        
        # Find ALL elements with ID 'viewAdjustments'
        all_adjustments = driver.find_elements(By.ID, "viewAdjustments")
        log_action(f"Found {len(all_adjustments)} elements with ID 'viewAdjustments'", 
                  log_file_path=log_file_path)
        
        # Find the visible and clickable one
        click_adjustments = None
        for idx, elem in enumerate(all_adjustments):
            try:
                if elem.is_displayed() and elem.is_enabled():
                    click_adjustments = elem
                    log_action(f"Found visible element at index {idx}", log_file_path=log_file_path)
                    break
            except:
                continue
        
        if not click_adjustments:
            raise Exception(f"No visible 'viewAdjustments' element found among {len(all_adjustments)} instances")
        
        # Scroll into view
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", 
            click_adjustments
        )
        time.sleep(0.5)
        
        # Click using JavaScript
        driver.execute_script("arguments[0].click();", click_adjustments)
        log_action("Clicked View Adjustments", log_file_path=log_file_path)
        
        # Wait for page ready after click
        WebDriverWait(driver, wait_time).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        time.sleep(1)
        
        driver.save_screenshot(os.path.join(screenshots_folder, "AfterClickAdjustments.png"))
        
        log_action("expand_adjustments completed successfully", log_file_path=log_file_path)
        return True
    
    except Exception as e:
        error_type = type(e).__name__
        error_msg = str(e).strip() or f"{error_type} occurred with no message"
        
        log_action(f"Exception in expand_adjustments: {error_type}", log_file_path=log_file_path)
        log_action(f"Error message: {error_msg}", log_file_path=log_file_path)
        log_action(f"Full traceback:\n{traceback.format_exc()}", log_file_path=log_file_path)
        
        driver.save_screenshot(os.path.join(screenshots_folder, "Error_ClickAdjustments.png"))
        raise Exception(f"expand_adjustments failed: {error_type} - {error_msg}")
    

def click_next_serving(driver, gray_table_number, log_file_path, screenshots_folder, timeout=30):
  
    gray_table_number = str(gray_table_number).strip()
    
    def button_is_ready(driver):
        try:
            btn = driver.find_element(By.ID, "btnNextServing")
            if btn.is_displayed() and btn.is_enabled():
                return btn
        except Exception:
            return False
        return False

    try:
        log_action(f"Waiting for 'Next Serving' button for table {gray_table_number}", 
                  log_file_path=log_file_path)
        
        button = WebDriverWait(driver, timeout).until(button_is_ready)
        
        log_action(f"'Next Serving' button found for table {gray_table_number}", 
                  log_file_path=log_file_path)
        
        # Scroll and click
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", button)
        
        log_action(f"'Next Serving' button clicked for table {gray_table_number}", 
                  log_file_path=log_file_path)
        driver.save_screenshot(os.path.join(screenshots_folder, f"AfterNextServing_Table{gray_table_number}.png"))
        
        return button
        
    except TimeoutException:
        error_msg = f"'Next Serving' button not found for table {gray_table_number} within {timeout} seconds"
        log_action(error_msg, log_file_path=log_file_path)
        driver.save_screenshot(os.path.join(screenshots_folder, f"Error_NextServing_Table{gray_table_number}.png"))
        raise ValueError(error_msg)
    
    except Exception as e:
        error_msg = f"Error clicking 'Next Serving' button for table {gray_table_number}: {str(e)}"
        log_action(error_msg, log_file_path=log_file_path)
        driver.save_screenshot(os.path.join(screenshots_folder, f"Error_NextServing_Table{gray_table_number}.png"))
        raise


def handle_payment_selection(driver, screenshots_folder, timeout=20):
    try:
        print("⏳ Waiting for payment channel dropdown...")
        
        # Wait for the dropdown to be present
        channel_dropdown = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "channel"))
        )
        
        print("\n" + "="*60)
        print("📸 Cycling through all payment methods...")
        print("="*60 + "\n")
        
        # Get all payment method options
        select = Select(channel_dropdown)
        all_options = select.options
        
        # Filter out the default/placeholder option
        payment_methods = []
        for option in all_options:
            value = option.get_attribute("value")
            if value:  # Skip empty value options
                payment_methods.append({
                    'value': value,
                    'text': option.text
                })
        
        # Cycle through each payment method and take screenshots
        for idx, method in enumerate(payment_methods, 1):
            print(f"\n--- {idx}/{len(payment_methods)}: {method['text']} ---")
            
            try:
                # Select the payment method
                select.select_by_value(method['value'])
                print(f"✅ Selected: {method['text']}")
                
                # Wait for dynamic fields to update
                WebDriverWait(driver, 5).until(
                    lambda d: d.execute_script('return document.readyState') == 'complete'
                )
                
                # Take screenshot
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                # Clean the method name for filename
                method_clean = method['value'].replace('/', '-').replace(' ', '_')
                screenshot_name = f"payment_{method_clean}_{timestamp}.png"
                screenshot_path = os.path.join(screenshots_folder, screenshot_name)
                driver.save_screenshot(screenshot_path)
                print(f"📸 Screenshot saved: {screenshot_path}")
                
                # Test validation: Try to click Receive Payment without filling reference
                print("🧪 Testing validation...")
                try:
                    receive_payment_btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.ID, "submitOtherPayment"))
                    )
                    
                    # Try JavaScript click to avoid interception
                    driver.execute_script("arguments[0].click();", receive_payment_btn)
                    
                    # Wait for any response
                    WebDriverWait(driver, 3).until(
                        lambda d: d.execute_script('return document.readyState') == 'complete'
                    )
                    
                    # Check for validation error
                    try:
                        error_msg = driver.find_element(By.CSS_SELECTOR, ".invalid-feedback.d-block")
                        if error_msg.is_displayed():
                            error_text = error_msg.text
                            print(f"⚠️ Validation error detected: {error_text}")
                            
                            # Screenshot validation error
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            screenshot_name = f"validation_error_{method_clean}_{timestamp}.png"
                            screenshot_path = os.path.join(screenshots_folder, screenshot_name)
                            driver.save_screenshot(screenshot_path)
                            print(f"📸 Validation error screenshot: {screenshot_path}")
                    except:
                        print("✅ No validation errors")
                    
                    # Close any open SweetAlert dialogs
                    try:
                        swal_dialog = driver.find_element(By.CSS_SELECTOR, ".swal2-popup.swal2-show")
                        if swal_dialog.is_displayed():
                            print("🔄 Closing validation dialog...")
                            # Try to click cancel/close button
                            try:
                                cancel_btn = driver.find_element(By.CSS_SELECTOR, "button.swal2-cancel")
                                cancel_btn.click()
                            except:
                                # Press ESC key to close
                                driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                            
                            # Wait for dialog to close
                            WebDriverWait(driver, 5).until_not(
                                EC.presence_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
                            )
                            print("✅ Dialog closed")
                    except:
                        pass
                        
                except Exception as e:
                    print(f"⚠️ Could not test validation: {e}")
                    
                    # Try to close any open dialogs
                    try:
                        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                        WebDriverWait(driver, 3).until_not(
                            EC.presence_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
                        )
                    except:
                        pass
                
                # Wait between payment methods
                print(f"⏳ Waiting for next selection...")
                WebDriverWait(driver, 5).until(
                    lambda d: d.execute_script('return document.readyState') == 'complete'
                )
                    
            except Exception as e:
                print(f"⚠️ Error with {method['text']}: {e}")
                continue
        
        # Ensure all dialogs are closed before proceeding
        print("\n🔄 Ensuring all dialogs are closed...")
        for attempt in range(3):  # Try multiple times
            try:
                swal_dialog = driver.find_element(By.CSS_SELECTOR, ".swal2-popup.swal2-show")
                if swal_dialog.is_displayed():
                    driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                    WebDriverWait(driver, 3).until_not(
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
                    )
            except:
                break
        
        # Final wait for page stability
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        print("\n" + "="*60)
        print("✅ All payment methods captured!")
        print("="*60 + "\n")
        
        # Now select CASH for final payment
        print("💰 Selecting CASH for final payment...")
        select.select_by_value("CASH")
        print("✅ Selected 'CASH' payment method")
        
        # Wait for page to be ready
        WebDriverWait(driver, 10).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        # Take final screenshot with CASH selected
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        screenshot_name = f"payment_CASH_final_{timestamp}.png"
        screenshot_path = os.path.join(screenshots_folder, screenshot_name)
        driver.save_screenshot(screenshot_path)
        print(f"📸 Final CASH screenshot: {screenshot_path}")
        
        # Wait for specified timeout before proceeding
        print(f"⏳ Waiting {timeout} seconds before proceeding...")
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        # Click Receive Payment button
        print("💰 Clicking 'Receive Payment' button...")
        
        # Try multiple methods to click the button
        try:
            # Method 1: JavaScript click (most reliable for overlays)
            receive_payment_btn = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "submitOtherPayment"))
            )
            driver.execute_script("arguments[0].click();", receive_payment_btn)
            print("✅ 'Receive Payment' button clicked!")
            
        except Exception as e1:
            print(f"⚠️  click failed: {e1}")
            try:
                # Method 2: Standard click
                receive_payment_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "submitOtherPayment"))
                )
                receive_payment_btn.click()
                print("✅ 'Receive Payment' button clicked!")
                
            except Exception as e2:
                print(f"⚠️ Standard click failed: {e2}")
                # Method 3: Try by button text
                receive_payment_btn = driver.find_element(By.XPATH, "//button[@id='submitOtherPayment']//span[text()='Receive Payment']/..")
                driver.execute_script("arguments[0].click();", receive_payment_btn)
                print("✅ 'Receive Payment' button clicked (XPath)!")
        
        # Wait for confirmation dialog
        WebDriverWait(driver, 5).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        
        # Handle SweetAlert2 confirmation dialog
        print("\n⏳ Waiting for confirmation dialog...")
        
        try:
            # Wait for SweetAlert2 popup to appear
            confirmation_dialog = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
            )
            print("✅ Confirmation dialog appeared!")
            
            # Take screenshot of confirmation dialog
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_name = f"confirmation_dialog_{timestamp}.png"
            screenshot_path = os.path.join(screenshots_folder, screenshot_name)
            driver.save_screenshot(screenshot_path)
            print(f"📸 Confirmation dialog screenshot: {screenshot_path}")
            
            # Extract and display confirmation details
            try:
                confirmation_text = driver.find_element(By.CSS_SELECTOR, ".swal2-html-container").text
                print(f"\n📋 Confirmation Details:\n{confirmation_text}\n")
            except:
                print("⚠️ Could not extract confirmation text")
            
            # Click "Yes" button to confirm
            print("✅ Clicking 'Yes' to confirm payment...")
            
            # Try multiple selectors for the "Yes" button
            yes_button_selectors = [
                (By.CSS_SELECTOR, "button.swal2-confirm"),
                (By.XPATH, "//button[contains(@class, 'swal2-confirm')]"),
                (By.XPATH, "//button[text()='Yes']"),
            ]
            
            clicked = False
            for by, selector in yes_button_selectors:
                try:
                    yes_button = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((by, selector))
                    )
                    yes_button.click()
                    print("✅ 'Yes' button clicked!")
                    clicked = True
                    break
                except:
                    continue
            
            if not clicked:
                # Try JavaScript click as fallback
                yes_button = driver.find_element(By.CSS_SELECTOR, "button.swal2-confirm")
                driver.execute_script("arguments[0].click();", yes_button)
                print("✅ 'Yes' button clicked!")
            
            # Wait for processing
            WebDriverWait(driver, 10).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            
            # Take screenshot after confirmation
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_name = f"payment_confirmed_{timestamp}.png"
            screenshot_path = os.path.join(screenshots_folder, screenshot_name)
            driver.save_screenshot(screenshot_path)
            print(f"📸 Payment confirmation screenshot: {screenshot_path}")
            
        except Exception as e:
            print(f"⚠️ Error handling confirmation dialog: {e}")
            # Take error screenshot
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_name = f"confirmation_error_{timestamp}.png"
            screenshot_path = os.path.join(screenshots_folder, screenshot_name)
            driver.save_screenshot(screenshot_path)
            print(f"📸 Error screenshot saved: {screenshot_path}")
        
        print("\n✅ Payment process completed!")
        return True
        
    except Exception as e:
        print(f"❌ Error in payment handling: {e}")
        # Take error screenshot
        screenshot_name = f"payment_error_{int(time.time())}.png"
        screenshot_path = os.path.join(screenshots_folder, screenshot_name)
        driver.save_screenshot(screenshot_path)
        print(f"📸 Error screenshot saved: {screenshot_path}")
        return False


def confirmation_process(driver, button_text=None, button_type="confirm", screenshots_folder=None, timeout=10):
  
    try:
        print(f"⏳ Waiting for confirmation dialog...")
        
        # Wait for dialog
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".swal2-popup.swal2-show"))
        )
        print("✅ Dialog appeared")
        
        # Screenshot before
        if screenshots_folder:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            driver.save_screenshot(os.path.join(screenshots_folder, f"swal2_dialog_{timestamp}.png"))
        
        # Get button class
        button_class = {"confirm": "swal2-confirm", "cancel": "swal2-cancel", "deny": "swal2-deny"}.get(button_type, "swal2-confirm")
        
        # Build selectors with contains logic
        if button_text:
            selectors = [
                (By.XPATH, f"//button[contains(@class, '{button_class}') and contains(., '{button_text}')]"),
                (By.CSS_SELECTOR, f"button.{button_class}")
            ]
        else:
            selectors = [(By.CSS_SELECTOR, f"button.{button_class}")]
        
        # Click button
        clicked = False
        for by, selector in selectors:
            try:
                button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((by, selector)))
                button.click()
                print(f"✅ Button clicked: '{button.text}'")
                clicked = True
                break
            except:
                continue
        
        # JavaScript fallback
        if not clicked:
            driver.execute_script(f"document.querySelector('button.{button_class}').click()")
            print(f"✅ Button clicked (JS)")
        
        time.sleep(2)
        
        # Screenshot after
        if screenshots_folder:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            driver.save_screenshot(os.path.join(screenshots_folder, f"swal2_confirmed_{timestamp}.png"))
        
        print("✅ Dialog confirmed")
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        if screenshots_folder:
            driver.save_screenshot(os.path.join(screenshots_folder, f"swal2_error_{int(time.time())}.png"))
        return False
    

def get_order_number(driver, wait):
    try:
        # Wait for table and first row
        wait.until(EC.presence_of_element_located((By.ID, 'manageOrdersTable_wrapper')))
        wait.until(EC.presence_of_element_located((By.XPATH, "//table[@id='manageOrdersTable']//tbody//tr[1]")))

        # Sort by Date/Time
        date_time_header = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//table[@id='manageOrdersTable']//thead//th[@data-dt-column='1']"))
        )

        print("Clicking Date/Time column to sort by most recent...")
        date_time_header.click()
        time.sleep(0.5)  # let sort animation/refresh trigger

        # Capture current first-row text
        old_first_row_text = driver.find_element(
            By.XPATH, "//table[@id='manageOrdersTable']//tbody//tr[1]//td[2]"
        ).text.strip()
        print(f"Old first row: '{old_first_row_text}'")

        def first_row_changed(d):
            try:
                new_text = d.find_element(
                    By.XPATH, "//table[@id='manageOrdersTable']//tbody//tr[1]//td[2]"
                ).text.strip()
                print(f"[DEBUG] Checking row... new='{new_text}', old='{old_first_row_text}'")
                return new_text != "" and new_text != old_first_row_text
            except Exception as inner_e:
                print(f"[DEBUG] Ignored error while checking first row: {inner_e}")
                return False

        try:
            wait.until(lambda d: first_row_changed(d))
            print("✅ First row changed successfully.")
        except TimeoutException:
            print("⚠️ Timeout: Table did not refresh or first row did not change. Proceeding with current first row.")

        # Wait until first row is visible and stable
        first_row = wait.until(
            EC.visibility_of_element_located((By.XPATH, "//table[@id='manageOrdersTable']//tbody//tr[1]"))
        )

        order_number = first_row.find_element(By.XPATH, "./td[1]").text.strip()
        date_time = first_row.find_element(By.XPATH, "./td[2]").text.strip()

        print(f"\nMost recent order:")
        print(f"  Order Number: {order_number}")
        print(f"  Date/Time: {date_time}")

        return order_number if order_number else None

    except Exception as e:
        print(f"❌ Error retrieving order number: {e}")
        traceback.print_exc()
        return None

    
# ==================== CONSTANTS ====================
# Define constants at the top for easy configuration
DEFAULT_TIMEOUT = 20
LONG_TIMEOUT = 30
SHORT_WAIT = 1
MEDIUM_WAIT = 2
LONG_WAIT = 5
TEST_TABLE_NUMBER = "34"

def wait_for_page_ready(driver, element_id=None, timeout=DEFAULT_TIMEOUT):
    
    # Wait for document ready state
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script('return document.readyState') == 'complete'
    )
    
    # Wait for specific element if provided
    if element_id:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
    
    time.sleep(SHORT_WAIT)  # Small buffer for animations

def find_table_by_number(driver, table_number, timeout=15):
   
    try:
        table = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((
                By.XPATH,
                f'//*[@id="setupRestaurantTables"]//div[contains(text(), "{table_number}")]'
            ))
        )
        print(f" Found table by text: {table_number}")
        return table
    except TimeoutException:
        print(f" Could not find table by text, trying data attribute...")
    
    try:
        table = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((
                By.XPATH,
                f'//*[@id="setupRestaurantTables"]//div[@data-table="{table_number}"]'
            ))
        )
        print(f"Found table by data attribute: {table_number}")
        return table
    except TimeoutException:
        print(f" Could not find table by data attribute, using positional fallback...")
    
    # Strategy 3: Fallback to original positional locator
    table = WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((
            By.XPATH,
            '//*[@id="setupRestaurantTables"]/div[67]/div'
        ))
    )
    print(f"Found table using positional locator (less reliable)")
    return table


def find_and_click_cancel(driver, timeout=15):

    cancel_strategies = [
        # Strategy 1: By text (most reliable)
        (By.XPATH, "//button[contains(text(), 'Cancel') or contains(text(), 'CANCEL')]"),
        # Strategy 2: By class (common pattern)
        (By.XPATH, "//*[@id='editCustomerModal']//button[contains(@class, 'btn-secondary')]"),
        # Strategy 3: Original positional (fallback)
        (By.XPATH, '//*[@id="editCustomerModal"]/div/div/div[3]/button[2]'),
    ]
    for i, (by, locator) in enumerate(cancel_strategies, 1):
        try:
            cancel_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((by, locator))
            )
            driver.execute_script("arguments[0].click();", cancel_btn)
            print(f"✅ Cancel clicked using strategy {i}")
            return True
        except TimeoutException:
            if i < len(cancel_strategies):
                print(f"⚠️ Strategy {i} failed, trying next...")
                continue
            else:
                raise Exception("Could not find cancel button with any strategy")
    
    return False

def safe_fill_input(driver, field_id, value, timeout=15):
 
    try:
        field = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, field_id))
        )
        field.clear()
        time.sleep(0.3)  # Small delay after clear
        human_like_typing(field, str(value))
        time.sleep(SHORT_WAIT)
        return True
    except TimeoutException:
        print(f" Could not find or interact with field: {field_id}")
        return False
    
def safe_click_table(driver, table_number, max_retries=3,log_file_path=None):
    for attempt in range(max_retries):
        try:
            WebDriverWait(driver, 20).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            table_element, table_num = choose_gray_table(driver, table_number=table_number)
            
            # Scroll and wait for visibility
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", table_element)
            WebDriverWait(driver, 10).until(lambda d: table_element.is_displayed())
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(table_element))
            
            # Click using JavaScript to avoid interception
            driver.execute_script("arguments[0].click();", table_element)
            
            log_action(f"Successfully clicked table {table_num}", log_file_path=log_file_path)
            return table_element, table_num
            
        except StaleElementReferenceException:
            if attempt < max_retries - 1:
                log_action(f"Stale element, retrying... (Attempt {attempt + 1})", log_file_path=log_file_path)
                time.sleep(2)
            else:
                raise

def process_additional_order(driver, table_number, log_file_path):

    try:
        log_action(f"Starting additional order process for table {table_number}", log_file_path=log_file_path)
        # 1. Wait for the order screen / UI to appear after table click
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "orderScreen"))  # adjust locator
        )
        # 2. Perform order selection: e.g., click “Add Item”, pick item, etc.
        add_item_btn = driver.find_element(By.ID, "btnAddItem")  # adjust locator
        add_item_btn.click()
        log_action(f"Clicked Add Item for table {table_number}", log_file_path=log_file_path)

        # 3. Fill in required fields, submit order
        # Example:
        item_dropdown = driver.find_element(By.ID, "selectItem")  # adjust
        item_dropdown.click()
        driver.find_element(By.XPATH, "//option[text()='ItemName']").click()  # adjust
        log_action("Selected item for additional order", log_file_path=log_file_path)

        submit_order_btn = driver.find_element(By.ID, "btnSubmitOrder")  # adjust
        submit_order_btn.click()
        log_action(f"Submitted additional order for table {table_number}", log_file_path=log_file_path)

        # 4. Wait for confirmation or success message
        WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.CLASS_NAME, "orderSuccessMessage"))  # adjust
        )
        log_action(f"Additional order process completed for table {table_number}", log_file_path=log_file_path)

    except Exception as e:
        log_error(f"Error in process_additional_order for table {table_number}: {traceback.format_exc()}",
                  log_file_path=log_file_path, driver=driver)
        raise  # or handle appropriately


def select_table_when_available(driver, table_number, log_file_path,
                                 available_class_indicator="available",
                                 cancel_reason="Table not in available state; retry later",
                                 max_wait_seconds=300,
                                 poll_interval_seconds=5):
 
    try:
        # Wait for table tiles present
        WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'table-tile'))
        )

        table_elem, found_table_number = choose_gray_table(driver, table_number=str(table_number))
        found_table_number = extract_table_number(table_elem)

        def is_available(elem):
            cls = elem.get_attribute("class")
            return (available_class_indicator in cls.split())

        if is_available(table_elem):
            log_action(f"Table {found_table_number} is available → proceeding", log_file_path=log_file_path)
            table_elem.click()
            process_additional_order(driver, found_table_number, log_file_path=log_file_path)
            return True

        # Not available: cancel and then wait for availability
        log_action(f"Table {found_table_number} not available → cancelling now", log_file_path=log_file_path)

        # Cancel action
        cancel_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "btnCancelTable"))
        )
        cancel_btn.click()
        log_action(f"Clicked Cancel Table button for table {found_table_number}", log_file_path=log_file_path)

        reason_input = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "cancelReasonInput"))
        )
        reason_input.clear()
        reason_input.send_keys(cancel_reason)
        log_action(f"Entered cancel reason for table {found_table_number}", log_file_path=log_file_path)

        confirm_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "confirmCancelTableBtn"))
        )
        confirm_btn.click()
        log_action(f"Confirmed cancellation for table {found_table_number}", log_file_path=log_file_path)

        # Now poll until table becomes available
        start_time = time.time()
        while time.time() - start_time < max_wait_seconds:  # Optionally refresh or re-locate tables list
           
            # Wait for page to refresh or table list to update
            time.sleep(poll_interval_seconds)
            try:
                table_elem, _ = choose_gray_table(driver, table_number=str(table_number))
            except Exception as e:
                log_action(f"Retry locating table {table_number}: {e}", log_file_path=log_file_path)
                continue

            if is_available(table_elem):
                log_action(f"Table {found_table_number} is now available → proceeding", log_file_path=log_file_path)
                table_elem.click()
                process_additional_order(driver, found_table_number, log_file_path=log_file_path)
                return True

            log_action(f"Table {found_table_number} still not available; waiting...", log_file_path=log_file_path)

        # Timeout
        log_error(f"Timeout waiting for table {found_table_number} to become available after cancellation", log_file_path=log_file_path, driver=driver)
        return False

    except Exception as e:
        error_message = f"Error in select_table_when_available (table {table_number}): {traceback.format_exc()}"
        log_error(error_message, log_file_path=log_file_path, driver=driver)
        return False
    
def attach_screenshot(driver, name="step"):
    import os
    screenshot_path = os.path.join("Reports", "screenshots", f"{name}.png")
    os.makedirs(os.path.dirname(screenshot_path), exist_ok=True)
    driver.save_screenshot(screenshot_path)
    print(f"[INFO] Saved screenshot: {screenshot_path}")

# Create New Service
def select_location(driver, select_id, search_text, log_file_path):
    
    try:
        # Find and click the Select2 container
        select2_container = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, f"//select[@id='{select_id}']/following-sibling::span[contains(@class, 'select2-container')]"))
        )
        
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select2_container)
        time.sleep(0.5)
        select2_container.click()
        time.sleep(1)
        
        # Find search input and type
        search_input = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".select2-search__field"))
        )
        search_input.clear()
        search_input.send_keys(search_text)
        time.sleep(1)
        
        # Click the matching option
        option = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, f"//li[contains(@class, 'select2-results__option') and contains(text(), '{search_text}')]"))
        )
        option.click()
        time.sleep(0.5)
        
        print(f"✓ Selected: {search_text}")
        return True
        
    except Exception as e:
        print(f"✗ Failed to select: {search_text}")
        print(f"   Error: {str(e)}")
        return False


def select_branch_booking(driver, branch_value=None, branch_text=None, wait_time=15, log_file_path="branch_log.txt", max_retries=3):
   
    def log_action(message):
        with open(log_file_path, "a") as f:
            f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")
        print(message)
    
    attempt = 0
    while attempt < max_retries:
        try:
            wait = WebDriverWait(driver, wait_time, poll_frequency=0.5)
            dropdown_element = wait.until(EC.element_to_be_clickable((By.NAME, "BranchId")))
            
            # Scroll into view
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_element)
            time.sleep(0.2)
            
            select = Select(dropdown_element)
            
            if branch_value:
                select.select_by_value(branch_value)
                log_action(f"Selected branch by value: {branch_value}")
            elif branch_text:
                select.select_by_visible_text(branch_text)
                log_action(f"Selected branch by text: {branch_text}")
            else:
                log_action("No branch value or text provided")
                return False
            
            return True
        
        except (TimeoutException, ElementClickInterceptedException, NoSuchElementException) as e:
            log_action(f"Attempt {attempt+1} failed: {str(e)}")
            time.sleep(0.5)
            attempt += 1

        except Exception as e:
            log_action(f"Unexpected error: {str(e)}")
            return False
    
    log_action("Failed to select branch after multiple attempts")
    return False


def click_with_flyout_fallback(
    driver, log_file_path, screenshots_folder,
    direct_locator, flyout_toggle_locator,
    flyout_panel_locator, flyout_item_locator,
    screenshot_prefix="action"
):
   
    try:
        WebDriverWait(driver, 30).until(lambda d: d.execute_script('return document.readyState') == 'complete')

        # Try direct click
        elem = WebDriverWait(driver, 30).until(EC.element_to_be_clickable(direct_locator))
        driver.execute_script("arguments[0].scrollIntoView(true);", elem)
        driver.execute_script("arguments[0].click();", elem)
        log_action(f"Clicked directly: {direct_locator}", log_file_path=log_file_path)
        WebDriverWait(driver, 30).until(lambda d: d.execute_script('return document.readyState') == 'complete')
        driver.save_screenshot(os.path.join(screenshots_folder, f"{screenshot_prefix}_direct.png"))
        return True

    except (TimeoutException, WebDriverException, ElementClickInterceptedException) as e:
        log_action(f"Direct click failed ({direct_locator}): {e}. Attempting flyout fallback", log_file_path=log_file_path)

        try:
            # Click the flyout toggle
            toggle = WebDriverWait(driver, 30).until(EC.element_to_be_clickable(flyout_toggle_locator))
            driver.execute_script("arguments[0].scrollIntoView(true);", toggle)
            driver.execute_script("arguments[0].click();", toggle)
            log_action(f"Clicked flyout toggle: {flyout_toggle_locator}", log_file_path=log_file_path)

            # Wait for the flyout to appear
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located(flyout_panel_locator))
            log_action(f"Flyout panel is visible: {flyout_panel_locator}", log_file_path=log_file_path)
            time.sleep(1)

            # Click the item in the flyout
            flyout_item = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable(flyout_item_locator)
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", flyout_item)
            driver.execute_script("arguments[0].click();", flyout_item)
            log_action(f"Clicked flyout item: {flyout_item_locator}", log_file_path=log_file_path)
            WebDriverWait(driver, 30).until(lambda d: d.execute_script('return document.readyState') == 'complete')
            driver.save_screenshot(os.path.join(screenshots_folder, f"{screenshot_prefix}_flyout.png"))
            return True

        except Exception as e2:
            log_error(f"Flyout fallback failed ({flyout_item_locator}): {traceback.format_exc()}", log_file_path=log_file_path, driver=driver)
            return False

# SCD Booking-Bulk
def fill_all_rows_with_same_values(
    driver, client_name, branch_name, log_file_path=None
):
    wait = WebDriverWait(driver, 10)

    # Find all rows
    rows = wait.until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#dataTable tbody tr"))
    )

    for i, row in enumerate(rows):
        try:
            # Select Client
            client_dropdown = row.find_element(By.CSS_SELECTOR, "select.dd-clientname")
            select_client = Select(client_dropdown)
            select_client.select_by_visible_text(client_name)

            # Select Branch - handle multiple branch dropdowns in row
            branch_dropdowns = row.find_elements(By.CSS_SELECTOR, "select.ob-select")
            for branch_dropdown in branch_dropdowns:
                select_branch = Select(branch_dropdown)
                select_branch.select_by_visible_text(branch_name)

            log_action(
                f"Row {i+1}: Selected client: {client_name} and branch: {branch_name}",
                log_file_path=log_file_path,
            )

        except Exception as e:
            log_action(f"Error in row {i+1}: {str(e)}", log_file_path=log_file_path)

        time.sleep(2)
