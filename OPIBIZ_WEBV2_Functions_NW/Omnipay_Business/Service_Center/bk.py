import webbrowser
import os
import shutil
import re
from pathlib import Path
from datetime import datetime
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import warnings

# Suppress warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=pytest.PytestDeprecationWarning)

# -------------------------------
# Global Variables
# -------------------------------
RTM = {
    "QA-SCDWEB-PB41-S1": {
        "bdd_scenario": "User Login",
        "testing_elements": "Login page, username, password, login button",
        "test_script": "test_login"
    },
    "QA-SCDWEB-PB41-S2": {
        "bdd_scenario": "Branch Selection",
        "testing_elements": "Branch dropdown, selection options, confirmation",
        "test_script": "test_branch_selection"
    },
    "QA-SCDWEB-PB41-S3": {
        "bdd_scenario": "User Logout",
        "testing_elements": "Logout button, session termination, redirect",
        "test_script": "test_logout"
    },
    "QA-SCDWEB-PB41-S4": {
        "bdd_scenario": "Main Dashboard",
        "testing_elements": "Data validation, submission",
        "test_script": "test_main_dashboard"
    },
    "QA-SCDWEB-PB41-S5": {
        "bdd_scenario": "Final Log Process",
        "testing_elements": "Final log forms, data validation, submission",
        "test_script": "test_final_log"
    }
}

test_results = []
report_stats = {"PASSED": 0, "FAILED": 0, "SKIPPED": 0}
test_logs = {}

# -------------------------------
# Fixtures
# -------------------------------
@pytest.fixture(scope="session")
def driver():
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    service = ChromeService(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.get("http://vm-app-dev01:9001/")
    driver.maximize_window()
    WebDriverWait(driver, 15)
    yield driver
    driver.quit()

@pytest.fixture
def context():
    return {}

@pytest.fixture
def log_file_path(request):
    module_name = request.node.parent.name
    base_log_dir = os.path.join(os.getcwd(), "Files", "logs")
    module_log_dir = os.path.join(base_log_dir, module_name)
    os.makedirs(module_log_dir, exist_ok=True)
    log_file = os.path.join(module_log_dir, f"{module_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    return log_file

# -------------------------------
# Utility Functions
# -------------------------------
def qa_sort_key(id_value):
    if not id_value:
        return 999
    match = re.search(r'(\d+)', str(id_value))
    return int(match.group(1)) if match else 999

def save_screenshot(driver_instance, test_name, status="failed"):
    screenshots_dir = os.path.join(os.getcwd(), "screenshots", status)
    os.makedirs(screenshots_dir, exist_ok=True)
    filename = f"{test_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    path = os.path.join(screenshots_dir, filename)
    driver_instance.save_screenshot(path)
    return path

def log_action(message, success=None, error=False, log_file=None):
    if error:
        success = False
    elif success is None:
        keywords = ["error", "failed", "exception", "not found", "unable", "timeout"]
        success = not any(k in message.lower() for k in keywords)

    status = "PASSED" if success else "FAILED"
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{current_time}] {message} - {status}"

    print(log_message)

    if log_file:
        try:
            with open(log_file, "a", encoding="utf-8") as f:
                f.write(log_message + "\n")
        except Exception as e:
            print(f"Failed to write log to file '{log_file}': {e}")

    return log_message

# -------------------------------
# Pytest Hooks
# -------------------------------
@pytest.hookimpl(tryfirst=True)
def pytest_sessionstart(session):
    print("\n=================== TEST SESSION START ===================\n")
    
    # Clean screenshots folder
    screenshots_dir = os.path.join(os.getcwd(), "screenshots")
    if os.path.exists(screenshots_dir):
        shutil.rmtree(screenshots_dir)
    os.makedirs(screenshots_dir, exist_ok=True)
    
    # Create logs folder
    logs_dir = os.path.join(os.getcwd(), "Files", "logs")
    os.makedirs(logs_dir, exist_ok=True)
    
    # Remove old reports
    for report_name in ["bdd_test_report.html", "bdd_test_report.xlsx"]:
        old_report = Path.cwd() / report_name
        if old_report.exists():
            old_report.unlink()
    
    print("Test session initialization completed.\n")

def pytest_configure(config):
    if hasattr(config, "_metadata"):
        config._metadata = {
            "Project": "SCD WEB - BDD Automation",
            "Framework": "Pytest-BDD + Selenium",
            "Tester": "Christian",
            "Test Approach": "BDD",
            "Browser": "Chrome"
        }

@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()

    if call.when != "call":
        return

    test_name = item.nodeid.split("::")[-1]
    test_id = "N/A"
    bdd_scenario = "N/A"
    testing_elements = "N/A"
    
    # Map test to RTM
    for qa_id, data in RTM.items():
        if data["test_script"].lower() in test_name.lower():
            test_id = qa_id
            bdd_scenario = data["bdd_scenario"]
            testing_elements = data["testing_elements"]
            break
    
    # Capture screenshot
    screenshot_path = None
    driver_instance = item.funcargs.get("driver")
    if driver_instance:
        status = "passed" if report.outcome == "passed" else "failed"
        screenshot_path = save_screenshot(driver_instance, test_name, status)
    
    # Update stats
    report_stats[report.outcome.upper()] = report_stats.get(report.outcome.upper(), 0) + 1
    
    # Get test logs
    log_file_path = item.funcargs.get("log_file_path")
    logs_content = ""
    if log_file_path and os.path.exists(log_file_path):
        try:
            with open(log_file_path, "r", encoding="utf-8") as f:
                logs_content = f.read()
        except:
            pass
    
    # Store result
    test_results.append({
        "ID": test_id,
        "BDD Scenario": bdd_scenario,
        "Testing Elements": testing_elements,
        "Test Script": test_name,
        "Status": report.outcome.upper(),
        "Result": "PASS" if report.outcome == "passed" else "FAIL",
        "Screenshot": screenshot_path,
        "Duration": f"{report.duration:.2f}s",
        "Logs": logs_content
    })

def pytest_sessionfinish(session, exitstatus):
    # Sort results
    test_results.sort(key=lambda x: qa_sort_key(x.get("ID")))
    
    # Generate reports
    generate_html_report()
    generate_excel_report()
    
    # Open HTML report
    html_report_file = os.path.join(os.getcwd(), "bdd_test_report.html")
    if os.path.exists(html_report_file):
        webbrowser.open_new_tab("file://" + os.path.abspath(html_report_file))
        print(f"\n✅ HTML report opened: {html_report_file}")

# -------------------------------
# HTML Report Generation
# -------------------------------
def generate_html_report():
    execution_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total = sum(report_stats.values())
    passed = report_stats.get("PASSED", 0)
    failed = report_stats.get("FAILED", 0)
    skipped = report_stats.get("SKIPPED", 0)
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>BDD Test Report - SCD WEB</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 40px;
            text-align: center;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}
        
        .header p {{
            font-size: 1.1em;
            opacity: 0.9;
        }}
        
        .dashboard {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            padding: 30px;
            background: #f8f9fa;
        }}
        
        .card {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            transition: transform 0.3s;
        }}
        
        .card:hover {{
            transform: translateY(-5px);
        }}
        
        .card h3 {{
            color: #667eea;
            font-size: 0.9em;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 10px;
        }}
        
        .card .value {{
            font-size: 2.5em;
            font-weight: bold;
            color: #333;
        }}
        
        .card.passed .value {{ color: #28a745; }}
        .card.failed .value {{ color: #dc3545; }}
        .card.skipped .value {{ color: #ffc107; }}
        .card.pass-rate .value {{ color: #17a2b8; }}
        
        .section {{
            padding: 30px;
        }}
        
        .section h2 {{
            color: #667eea;
            font-size: 1.8em;
            margin-bottom: 20px;
            border-bottom: 3px solid #667eea;
            padding-bottom: 10px;
        }}
        
        .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 15px;
            margin-bottom: 30px;
        }}
        
        .info-item {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            border-left: 4px solid #667eea;
        }}
        
        .info-item strong {{
            color: #667eea;
            display: inline-block;
            min-width: 150px;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            background: white;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        
        th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
            text-transform: uppercase;
            font-size: 0.9em;
            letter-spacing: 0.5px;
        }}
        
        td {{
            padding: 15px;
            border-bottom: 1px solid #e9ecef;
        }}
        
        tr:hover {{
            background: #f8f9fa;
        }}
        
        .status-passed {{
            background: #28a745;
            color: white;
            padding: 5px 15px;
            border-radius: 20px;
            font-weight: bold;
            display: inline-block;
        }}
        
        .status-failed {{
            background: #dc3545;
            color: white;
            padding: 5px 15px;
            border-radius: 20px;
            font-weight: bold;
            display: inline-block;
        }}
        
        .screenshot {{
            max-width: 150px;
            max-height: 100px;
            border-radius: 5px;
            cursor: pointer;
            transition: transform 0.3s;
            box-shadow: 0 2px 8px rgba(0,0,0,0.2);
        }}
        
        .screenshot:hover {{
            transform: scale(1.1);
        }}
        
        .modal {{
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.9);
            justify-content: center;
            align-items: center;
        }}
        
        .modal img {{
            max-width: 90%;
            max-height: 90%;
            border-radius: 10px;
        }}
        
        .close-modal {{
            position: absolute;
            top: 20px;
            right: 40px;
            color: white;
            font-size: 40px;
            font-weight: bold;
            cursor: pointer;
        }}
        
        .logs-section {{
            background: #1e1e1e;
            color: #d4d4d4;
            padding: 20px;
            border-radius: 8px;
            font-family: 'Courier New', monospace;
            font-size: 0.9em;
            max-height: 300px;
            overflow-y: auto;
            margin-top: 10px;
        }}
        
        .logs-section::-webkit-scrollbar {{
            width: 8px;
        }}
        
        .logs-section::-webkit-scrollbar-track {{
            background: #2d2d2d;
        }}
        
        .logs-section::-webkit-scrollbar-thumb {{
            background: #667eea;
            border-radius: 4px;
        }}
        
        .export-btn {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 12px 30px;
            border-radius: 25px;
            font-size: 1em;
            cursor: pointer;
            transition: all 0.3s;
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
            margin: 20px 30px;
        }}
        
        .export-btn:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
        }}
        
        .footer {{
            background: #f8f9fa;
            padding: 20px;
            text-align: center;
            color: #666;
            border-top: 1px solid #dee2e6;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎯 BDD Test Execution Report</h1>
            <p>SCD WEB - Restaurant Center | Behavior Driven Development</p>
        </div>
        
        <div class="dashboard">
            <div class="card">
                <h3>Total Tests</h3>
                <div class="value">{total}</div>
            </div>
            <div class="card passed">
                <h3>Passed</h3>
                <div class="value">{passed}</div>
            </div>
            <div class="card failed">
                <h3>Failed</h3>
                <div class="value">{failed}</div>
            </div>
            <div class="card pass-rate">
                <h3>Pass Rate</h3>
                <div class="value">{pass_rate:.1f}%</div>
            </div>
        </div>
        
        <div class="section">
            <h2>📋 Execution Summary</h2>
            <div class="info-grid">
                <div class="info-item">
                    <strong>Project:</strong> SCD Web - Restaurant Center
                </div>
                <div class="info-item">
                    <strong>Version:</strong> v1.2.0.R0002B
                </div>
                <div class="info-item">
                    <strong>Framework:</strong> Pytest + Selenium WebDriver
                </div>
                <div class="info-item">
                    <strong>Test Approach:</strong> Behavior Driven Development
                </div>
                <div class="info-item">
                    <strong>Tester:</strong> Christian
                </div>
                <div class="info-item">
                    <strong>Execution Date:</strong> {execution_date}
                </div>
                <div class="info-item">
                    <strong>Browser:</strong> Chrome (Selenium WebDriver)
                </div>
                <div class="info-item">
                    <strong>Environment:</strong> http://vm-app-dev01:9001/
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2>📊 Requirements Traceability Matrix (RTM)</h2>
            <table id="rtmTable">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>BDD Scenario</th>
                        <th>Testing Elements</th>
                        <th>Test Script</th>
                        <th>Status</th>
                        <th>Duration</th>
                        <th>Screenshot</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Add test results
    for result in test_results:
        status_class = "status-passed" if result["Result"] == "PASS" else "status-failed"
        screenshot_html = ""
        if result["Screenshot"] and os.path.exists(result["Screenshot"]):
            screenshot_html = f'<img src="file:///{os.path.abspath(result["Screenshot"])}" class="screenshot" onclick="openModal(this.src)" alt="Screenshot">'
        
        html_content += f"""
                    <tr>
                        <td><strong>{result["ID"]}</strong></td>
                        <td>{result["BDD Scenario"]}</td>
                        <td>{result["Testing Elements"]}</td>
                        <td>{result["Test Script"]}</td>
                        <td><span class="{status_class}">{result["Result"]}</span></td>
                        <td>{result["Duration"]}</td>
                        <td>{screenshot_html}</td>
                    </tr>
"""
    
    html_content += """
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2>📝 Test Execution Logs</h2>
"""
    
    # Add logs for each test
    for result in test_results:
        if result["Logs"]:
            html_content += f"""
            <h3 style="color: #667eea; margin-top: 20px;">{result["ID"]} - {result["BDD Scenario"]}</h3>
            <div class="logs-section">{result["Logs"].replace(chr(10), '<br>')}</div>
"""
    
    html_content += """
        </div>
        
        <button class="export-btn" onclick="exportToExcel()">📥 Export to Excel</button>
        
        <div class="footer">
            <p>Generated by BDD Test Automation Framework | © 2025 SCD WEB Project</p>
            <p>Powered by Pytest-BDD + Selenium WebDriver</p>
        </div>
    </div>
    
    <div id="modal" class="modal" onclick="closeModal()">
        <span class="close-modal">&times;</span>
        <img id="modalImg" src="" alt="Screenshot">
    </div>
    
    <script>
        function openModal(src) {
            document.getElementById('modal').style.display = 'flex';
            document.getElementById('modalImg').src = src;
        }
        
        function closeModal() {
            document.getElementById('modal').style.display = 'none';
        }
        
        function exportToExcel() {
            alert('Excel file has been generated: bdd_test_report.xlsx');
        }
    </script>
</body>
</html>"""
    
    # Save HTML report
    html_path = os.path.join(os.getcwd(), "bdd_test_report.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    
    print(f"✅ HTML report generated: {html_path}")

# -------------------------------
# Excel Report Generation
# -------------------------------
def generate_excel_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Add title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "BDD Test Execution Report - SCD WEB"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Add summary
    execution_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total = sum(report_stats.values())
    passed = report_stats.get("PASSED", 0)
    failed = report_stats.get("FAILED", 0)
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    summary_data = [
        ["Project:", "SCD Web - Restaurant Center"],
        ["Tester:", "Christian"],
        ["Framework:", "Pytest-BDD + Selenium"],
        ["Execution Date:", execution_date],
        ["Total Tests:", total],
        ["Passed:", passed],
        ["Failed:", failed],
        ["Pass Rate:", f"{pass_rate:.1f}%"]
    ]
    
    row = 3
    for item in summary_data:
        ws[f'A{row}'] = item[0]
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = item[1]
        row += 1
    
    # Add headers
    row += 2
    headers = ["ID", "BDD Scenario", "Testing Elements", "Test Script", "Status", "Duration", "Screenshot"]
    header_row = row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    row += 1
    max_screenshot_width = 0
    
    for result in test_results:
        ws.cell(row=row, column=1, value=result["ID"])
        ws.cell(row=row, column=2, value=result["BDD Scenario"])
        ws.cell(row=row, column=3, value=result["Testing Elements"])
        ws.cell(row=row, column=4, value=result["Test Script"])
        
        status_cell = ws.cell(row=row, column=5, value=result["Result"])
        if result["Result"] == "PASS":
            status_cell.font = Font(color="006100", bold=True)
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.font = Font(color="9C0006", bold=True)
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row, column=6, value=result["Duration"])
        
        # Insert screenshot
        if result["Screenshot"] and os.path.exists(result["Screenshot"]):
            try:
                with PILImage.open(result["Screenshot"]) as img:
                    max_height = 80
                    aspect_ratio = img.width / img.height
                    new_height = min(img.height, max_height)
                    new_width = int(new_height * aspect_ratio)
                    
                    img_resized = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                    resized_path = result["Screenshot"].replace(".png", "_resized.png")
                    img_resized.save(resized_path)
                    
                    excel_img = XLImage(resized_path)
                    ws.add_image(excel_img, f"G{row}")
                    ws.row_dimensions[row].height = new_height * 0.75
                    max_screenshot_width = max(max_screenshot_width, new_width)
            except Exception as e:
                print(f"Error inserting screenshot: {e}")
        
        row += 1
    
    # Auto-fit columns
    for col in range(1, 7):
        max_length = 0
        column = ws.column_dimensions[chr(64 + col)]
        for cell in ws[chr(64 + col)]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        column.width = min(max_length + 2, 50)
    
    if max_screenshot_width > 0:
        ws.column_dimensions["G"].width = max_screenshot_width / 7
    
    # Save Excel file
    excel_path = os.path.join(os.getcwd(), "bdd_test_report.xlsx")
    try:
        wb.save(excel_path)
        print(f"✅ Excel report generated: {excel_path}")
    except PermissionError:
        alt_path = os.path.join(os.getcwd(), f"bdd_test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(alt_path)
        print(f"✅ Excel report saved as: {alt_path}")